# return_process/views.py

from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView
import requests
from django.shortcuts import render, redirect
from .models import ReturnItem, ScanLog
from .api_clients import (
    NAVER_ACCOUNTS,
    approve_naver_return,
    get_product_order_details,
    dispatch_naver_exchange,
    get_hourly_sales,
    get_hourly_sales_for_period,
    
)
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from .forms import CustomUserCreationForm  # 커스텀 유저 생성 폼
import logging
from django.http import JsonResponse
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
import json
from django.shortcuts import get_object_or_404
from django.http import JsonResponse, HttpResponse
import openpyxl
from openpyxl import Workbook
import warnings
import datetime, uuid, hmac, hashlib
from django.conf import settings
from .utils import update_returns_logic
from openpyxl.utils import get_column_letter
import pandas as pd
from django.db.models.functions import Coalesce, TruncDate
from .models import ReturnItem
from django.db.models import DateTimeField,Count, Q,Sum,Count
import io
from django.views.decorators.http import require_GET



warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


logger = logging.getLogger(__name__)  # __name__은 현재 모듈 이름

class CustomLoginView(LoginView):
    template_name = 'return_process/login.html'
    success_url = reverse_lazy('반품목록')

class ReturnListView(LoginRequiredMixin, ListView):
    model = ReturnItem
    template_name = 'return_process/return_list.html'
    context_object_name = 'return_items'

@login_required
def return_list(request):
    status_map = {
        '미처리': '미처리',
        '스캔': '스캔',
        '수거완료': '수거완료',
        '검수완료': '검수완료',
        '반품완료': '반품완료',
        '재고반영': '재고반영',
        '처리완료': '처리완료'
    }

    # 최신 업데이트 기준으로 정렬된 아이템
    return_items = ReturnItem.objects.order_by('-last_update_date')

        # ---- (검색어 필터링) ----
    search_recipient_name = request.GET.get('recipient_name', '').strip()
    if search_recipient_name:
        # 수취인명 컬럼(recipient_name)을 icontains로 필터
        return_items = return_items.filter(recipient_name__icontains=search_recipient_name)


    # 총 반품건수
    total_count = return_items.count()

    # 상태별 카운트 (ORM 집계) - DB 상의 processing_status값 그대로 group by
    raw_status_counts = (
        return_items
        .values('processing_status')
        .annotate(count=Count('id'))
        .order_by('processing_status')
    )
    # raw_status_counts = [{'processing_status': '반품완료', 'count': 10}, ...]

    # 상태맵 사용해 사람이 읽기 좋은 상태명으로 교체
    # 예: DB에 '반품완료' 그대로 저장되어 있으면 그냥 동일, 
    #    만약 'returned' 같은 영어였다면 -> 한글 '반품완료'
    status_counts = []
    for sc in raw_status_counts:
        original_status = sc['processing_status']
        display_status = status_map.get(original_status, original_status)
        status_counts.append({
            'processing_status': display_status,
            'count': sc['count'],
        })

    return render(request, 'return_process/return_list.html', {
        'return_items': return_items,
        'total_count': total_count,
        'status_counts': status_counts,
    })


# 매핑 딕셔너리 상단에 위치 (기존 딕셔너리에 항목 추가)
receipt_type_map = {
    'RETURN': '반품',
    'CANCEL_REQUEST': '취소 요청',
    'RETURN_REQUEST': '반품 요청',
    'EXCHANGE_REQUEST': '교환 요청',
    'COMPENSATION_REQUEST': '보상 요청',
}

receipt_status_map = {
    'CANCEL_REQUEST': '취소 요청',
    'CANCEL_ACCEPT': '취소 수락',
    'CANCEL_REJECT': '취소 거부',
    'RETURNS_COMPLETED': '반품 완료',
    'RELEASE_STOP_UNCHECKED': '출고중지요청',
    'RETURNS_UNCHECKED': '반품접수',
    'VENDOR_WAREHOUSE_CONFIRM': '입고완료',
    'REQUEST_COUPANG_CHECK': '쿠팡확인요청',
    'RETURN_REQUEST': '반품 요청',
    'RETURN_ACCEPT': '반품 수락',
    'RETURN_REJECT': '반품 거부',
    'RETURN_COMPLETE': '반품 완료',
    'EXCHANGE_REQUEST': '교환 요청',
    'EXCHANGE_ACCEPT': '교환 수락',
    'EXCHANGE_REJECT': '교환 거부',
    'COLLECTING'	: '수거중',
    'COLLECT_DONE'    : '수거완료',
    'EXCHANGE_DONE	': '교환완료',
    'RETURN_REJECT	': '반품철회',
    'EXCHANGE_REJECT': '교환철회',
}

cancel_reason_category_map = {
    'CUSTOMER_CHANGE': '고객변심',
    'SELLER_FAULTY': '판매자 귀책',
    'DROPPED_DELIVERY': '배송 중 분실',
    'DAMAGED_DELIVERY': '배송 중 파손',
    'WRONG_DELIVERY': '오배송',  # 이미 여기에도 '오배송'이 있으나 상황에 따라 사용
    'PRODUCT_UNSATISFIED': '상품 불만족',
    'PRODUCT_DEFECT': '상품 불량',
}

status_code_mapping = {
    'WRONG_DELIVERED_PRODUCT': '오배송',
    'WRONG_PRODUCT_CONTENT': '상품 내용 불일치',
    'WRONG_DELAYED_DELIVERY': '배송 지연',
    'SIMPLE_INTENT_CHANGED': '단순 변심',
    'MISTAKE_ORDER': '주문 실수',
    'NO_LONGER_NEED_PRODUCT': '상품이 필요 없음',
    'PRODUCT_UNSATISFIED': '상품에 만족하지 않음',
    'PRODUCT_UNSATISFIED_COLOR': '색상 불만족',
    'PRODUCT_UNSATISFIED_SIZE': '사이즈 불만족',
    'BROKEN_AND_BAD': '상품 불량 및 파손',
    'PRODUCT_UNSATISFIED_ETC': '기타 불만족',
    'RETURN_REQUEST': '반품 요청',
    'RETURN_REJECT': '반품 철회',
    'RETURN_COMPLETED': '반품 완료',
    'EXCHANGE_REQUEST': '교환 요청',
    'EXCHANGE_REJECT': '교환 철회',
    'EXCHANGE_COMPLETED': '교환 완료',
    'COLLECT_WAITING': '회수 대기 중',
    'COLLECT_DONE': '회수 완료',
    'COLLECT_DELAYED': '회수 지연',
    'DELIVERED': '배송 완료',
    'DELIVERY_COMPLETION': '배송 완료',
    'DELIVERING': '배송 중',
    'SHIPPING': '배송 준비 중',
    'RETURN_DONE': '반품 완료',
    'PURCHASE_DECIDED': '구매 확정',

    # 추가한 매핑
    'COLOR_AND_SIZE': '색상 및 사이즈 변경',
    'WRONG_DELIVERY': '오배송',
}

product_order_status_map = {
    'PAYMENT_WAITING': '결제 대기',
    'PAYED': '결제 완료',
    'DELIVERING': '배송 중',
    'DELIVERED': '배송 완료',
    'PURCHASE_DECIDED': '구매 확정',
    'EXCHANGED': '교환완료',
    'CANCELED': '취소완료',
    'RETURNED': '반품완료',
    'CANCELED_BY_NOPAYMENT': '미결제 취소'
}


@login_required
def update_returns(request):
    update_returns_logic()
    messages.success(request, '반품 데이터 업데이트가 완료되었습니다.')
    return redirect('반품목록')


@login_required
def update_returns(request):
    update_returns_logic()
    messages.success(request, '반품 데이터 업데이트가 완료되었습니다.')
    return redirect('반품목록')

@login_required
def delete_return_item(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        item_ids = data.get('item_ids', [])

        if not item_ids:
            return JsonResponse({'success': False, 'message': '아이템 ID 목록이 필요합니다.'})

        deleted_count, _ = ReturnItem.objects.filter(id__in=item_ids).delete()
        if deleted_count > 0:
            return JsonResponse({'success': True, 'message': f'{deleted_count}개 항목 삭제 완료'})
        else:
            return JsonResponse({'success': False, 'message': '해당 ID의 레코드를 찾을 수 없습니다.'})
    else:
        return JsonResponse({'success': False, 'message': 'POST 요청만 가능합니다.'})


logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

@login_required
def get_column_index_by_name(header_row, target_name):
    for idx, cell in enumerate(header_row):
        if cell.value and cell.value.strip() == target_name.strip():
            return idx
    return None


@login_required
def upload_returns_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            logger.warning("엑셀 파일이 선택되지 않았습니다.")
            return JsonResponse({'success': False, 'message': '엑셀 파일을 선택해주세요.'})
        import unicodedata

        # 엑셀 파일명에 "API_반품" 문자열 포함 여부 확인
        file_name = unicodedata.normalize('NFC', excel_file.name)
        if "API_반품" not in file_name:
            logger.warning("잘못된 엑셀 파일명: %s", file_name)
            return JsonResponse({'success': False, 'message': '파일명을 확인하세요 API_반품이 파일명에 포함되어야 합니다.'})
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            logger.debug("셀러툴 엑셀 파일 로드 성공")
        except Exception as e:
            logger.error(f"엑셀 파일 로드 오류: {str(e)}")
            return JsonResponse({'success': False, 'message': '유효한 엑셀 파일이 아닙니다.'})
        
        ws = wb.active

        success_items = []
        error_items = []

        def format_korean_datetime(dt):
            if not dt:
                return ''
            formatted = dt.strftime("%Y년 %m월 %d일 %I:%M %p")
            formatted = formatted.replace("AM", "오전").replace("PM", "오후")
            logger.debug(f"format_korean_datetime: {dt} -> {formatted}")
            return formatted

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            values = [c.value for c in row]
            print(f"Row {row_idx} values: {values}")
            def cell_value(idx, is_numeric=False):
                if idx < len(row):
                    val = row[idx].value
                    if val is None:
                        return 0 if is_numeric else ''
                    if is_numeric:
                        return val if val != '' else 0
                    else:
                        return str(val).strip() if isinstance(val, str) else str(val)
                else:
                    return 0 if is_numeric else ''

            status = cell_value(0) or '미처리'
            claim_type = cell_value(1)
            store_name = cell_value(2)
            order_number = cell_value(3)
            collect_tracking_number = cell_value(4)
            collect_delivery_company = cell_value(5)
            recipient_name = cell_value(6)
            option_code = cell_value(7)
            product_name = cell_value(8)
            option_name = cell_value(9)
            quantity = cell_value(10, is_numeric=True)
            invoice_number = cell_value(11)
            # 하이픈 제거 추가
            invoice_number = invoice_number.replace('-', '') if invoice_number else invoice_number

            claim_status = cell_value(12)
            claim_reason = cell_value(13)
            customer_reason = cell_value(14)
            return_shipping_charge = cell_value(15, is_numeric=True)
            shipping_charge_payment_method = cell_value(16)
            processing_status = cell_value(17) or '미처리'
            note = cell_value(18)
            inspector = cell_value(19)
            product_issue = cell_value(20)
            raw_delivered_date = cell_value(21)
            raw_claim_request_date = cell_value(22)

            logger.debug(f"[Row {row_idx}] order_number={order_number}, raw_delivered_date={raw_delivered_date}, raw_claim_request_date={raw_claim_request_date}")

            if not order_number:
                error_items.append({
                    'row': row_idx,
                    'reason': '주문번호 없음'
                })
                continue

            # 스토어명 매핑
            store_name_map = {
                "노는 개 최고양": "노는개최고양",
                "수비다 SUBIDA": "수비다"
            }
            if store_name in store_name_map:
                store_name = store_name_map[store_name]

            # 특정 스토어일 경우 platform='Naver' 설정
            naver_stores = ["니뜰리히", "노는개최고양", "아르빙", "수비다"]
            if store_name in naver_stores:
                platform_val = 'Naver'
            else:
                platform_val = ''

            item, created = ReturnItem.objects.get_or_create(
                order_number=order_number,
                option_code=option_code
            )
            
            item.platform = platform_val
            item.processing_status = processing_status
            item.claim_type = claim_type
            item.store_name = store_name
            item.collect_tracking_number = collect_tracking_number
            item.collect_delivery_company = collect_delivery_company
            item.recipient_name = recipient_name
            item.option_code = option_code
            item.product_name = product_name
            item.option_name = option_name
            item.quantity = quantity
            item.invoice_number = invoice_number
            item.claim_status = claim_status
            item.claim_reason = claim_reason
            item.customer_reason = customer_reason
            item.return_shipping_charge = return_shipping_charge
            item.shipping_charge_payment_method = shipping_charge_payment_method
            item.note = note
            item.inspector = inspector
            item.product_issue = product_issue

            def parse_date(raw_val):
                if raw_val == '':
                    logger.debug(f"parse_date: empty string -> None")
                    return None
                try:
                    logger.debug(f"parse_date: attempting to parse {raw_val}")
                    parsed = datetime.datetime.strptime(raw_val, "%Y-%m-%d")
                    parsed = parsed.replace(hour=10, minute=0)
                    logger.debug(f"parse_date: {raw_val} -> {parsed}")
                    return parsed
                except ValueError as ve:
                    logger.warning(f"parse_date: {raw_val} 변환 실패: {ve}")
                    return None

            delivered_date = parse_date(raw_delivered_date)
            claim_request_date = parse_date(raw_claim_request_date)

            logger.debug(f"[Row {row_idx}] 파싱 결과 delivered_date={delivered_date}, claim_request_date={claim_request_date}")

            item.delivered_date = delivered_date
            item.claim_request_date = claim_request_date
            item.save()

            formatted_delivered = format_korean_datetime(item.delivered_date)
            formatted_claim_request = format_korean_datetime(item.claim_request_date)
            logger.debug(f"[Row {row_idx}] formatted_delivered={formatted_delivered}, formatted_claim_request={formatted_claim_request}")

            success_items.append({
                'order_number': item.order_number,
                'display_status': getattr(item, 'display_status', ''),
                'claim_type': item.claim_type,
                'store_name': item.store_name,
                'collect_tracking_number': item.collect_tracking_number,
                'collect_delivery_company': item.collect_delivery_company,
                'recipient_name': item.recipient_name,
                'option_code': item.option_code,
                'product_name': item.product_name,
                'option_name': item.option_name,
                'quantity': item.quantity,
                'invoice_number': item.invoice_number,
                'claim_status': item.claim_status,
                'claim_reason': item.claim_reason,
                'customer_reason': item.customer_reason,
                'return_shipping_charge': item.return_shipping_charge,
                'shipping_charge_payment_method': item.shipping_charge_payment_method,
                'processing_status': item.processing_status,
                'note': item.note,
                'inspector': item.inspector,
                'product_issue': item.product_issue,
                'delivered_date': formatted_delivered,
                'claim_request_date': formatted_claim_request,
                'display_status': getattr(item, 'display_status', ''),
            })

        logger.debug(f"업로드 완료: 성공={len(success_items)}개, 실패={len(error_items)}개")

        uploaded_order_numbers = [item['order_number'] for item in success_items]
        ReturnItem.objects.filter(order_number__in=uploaded_order_numbers).update(processing_status='미처리')

        return JsonResponse({
            'success': True,
            'items': success_items,
            'errors': error_items
        })
    return JsonResponse({'success': False, 'message': 'Invalid request'})

@login_required
def upload_courier_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            logger.warning("엑셀 파일이 선택되지 않았습니다.")
            return JsonResponse({'success': False, 'message': '엑셀 파일을 선택해주세요.'})
        import unicodedata

        # 파일명 정규화 후 "운송장리스트" 포함 여부 확인
        file_name_normalized = unicodedata.normalize('NFC', excel_file.name)
        if "운송장리스트" not in file_name_normalized:
            logger.warning("잘못된 엑셀 파일명: %s", file_name_normalized)
            return JsonResponse({'success': False, 'message': '파일명을 확인하세요, 운송장리스트라는 파일명이 포함되어야 합니다.'})
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            logger.debug("택배사 엑셀 파일 로드 성공")
        except Exception as e:
            logger.error(f"엑셀 파일 로드 오류: {str(e)}")
            return JsonResponse({'success': False, 'message': '유효한 엑셀 파일이 아닙니다.'})

        ws = wb.active

        header = next(ws.iter_rows(min_row=1, max_row=1))
        header_values = [cell.value for cell in header]
        logger.debug(f"택배사 엑셀 헤더: {header_values}")

        related_col_name = "관련운송장"
        new_collect_col_name = "운송장번호"
        delivered_date_col_name = "배송일자"  # 실제 헤더명에 맞게 변경 필요

        def get_col_index(col_name):
            if col_name in header_values:
                return header_values.index(col_name)
            else:
                raise Exception(f"'{col_name}' 컬럼을 찾을 수 없습니다. 엑셀 헤더를 확인하세요.")

        try:
            related_waybill_col = get_col_index(related_col_name)
            new_collect_col = get_col_index(new_collect_col_name)
            delivered_date_col = get_col_index(delivered_date_col_name)
        except Exception as e:
            logger.error(str(e))
            return JsonResponse({'success': False, 'message': str(e)})

        success_items = []
        error_items = []

        def parse_date(val):
            if isinstance(val, datetime.datetime):
                logger.debug(f"parse_date: {val}는 datetime 객체입니다.")
                parsed = val.replace(hour=10, minute=0)
                logger.debug(f"parse_date: 시간 설정 후 {parsed}")
                return parsed
            if isinstance(val, str) and val.strip():
                logger.debug(f"parse_date: '{val}' 문자열 파싱 시도")
                try:
                    parsed = datetime.datetime.strptime(val.strip(), "%Y-%m-%d")
                    parsed = parsed.replace(hour=10, minute=0)
                    logger.debug(f"parse_date: {val} -> {parsed}")
                    return parsed
                except ValueError as ve:
                    logger.warning(f"parse_date: '{val}' 변환 실패: {ve}")
                    return None
            logger.debug(f"parse_date: '{val}'는 빈 값 또는 알 수 없는 형식")
            return None

        def format_korean_datetime(dt):
            if not dt:
                return ''
            formatted = dt.strftime("%Y년 %m월 %d일 %I:%M %p")
            formatted = formatted.replace("AM", "오전").replace("PM", "오후")
            logger.debug(f"format_korean_datetime: {dt} -> {formatted}")
            return formatted

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            raw_related = row[related_waybill_col].value
            raw_collect_tracking = row[new_collect_col].value
            raw_delivered_date = row[delivered_date_col].value

            logger.debug(f"[Row {row_idx}] raw_related={raw_related}, raw_collect_tracking={raw_collect_tracking}, raw_delivered_date={raw_delivered_date}")

            if isinstance(raw_related, datetime.datetime):
                related_waybill = raw_related.strftime("%Y%m%d%H%M%S")
            else:
                related_waybill = str(raw_related).strip() if raw_related else None

            if isinstance(raw_collect_tracking, datetime.datetime):
                new_collect_tracking = raw_collect_tracking.strftime("%Y%m%d%H%M%S")
            else:
                new_collect_tracking = str(raw_collect_tracking).strip() if raw_collect_tracking else None

            if related_waybill:
                related_waybill = related_waybill.replace("-", "")
            if new_collect_tracking:
                new_collect_tracking = new_collect_tracking.replace("-", "")

            logger.debug(f"[Row {row_idx}] 배송일자 파싱 시작: {raw_delivered_date}")
            courier_delivered_date = parse_date(raw_delivered_date)
            logger.debug(f"[Row {row_idx}] 파싱 결과 courier_delivered_date={courier_delivered_date}")

            if not related_waybill:
                logger.debug(f"[Row {row_idx}] 관련운송장 없음. 스킵.")
                continue

            items = ReturnItem.objects.filter(invoice_number=related_waybill)

            if not items.exists():
                logger.warning(f"[Row {row_idx}] invoice_number={related_waybill} 매칭 데이터 없음.")
                error_items.append({
                    'row': row_idx,
                    'related_waybill': related_waybill,
                    'error': '매칭되는 데이터 없음'
                })
                continue

            logger.debug(f"[Row {row_idx}] {items.count()}개 항목 매칭됨.")

            for item in items:
                original_collect = item.collect_tracking_number
                original_delivered = item.delivered_date

                if new_collect_tracking:
                    item.collect_tracking_number = new_collect_tracking
                if courier_delivered_date:
                    item.delivered_date = courier_delivered_date

                item.save()

                logger.debug(f"[Row {row_idx}] 주문번호={item.order_number}, CollectTracking: {original_collect} -> {item.collect_tracking_number}, DeliveredDate: {original_delivered} -> {item.delivered_date}")

                formatted_delivered = format_korean_datetime(item.delivered_date)
                logger.debug(f"[Row {row_idx}] formatted_delivered={formatted_delivered}")

                success_items.append({
                    'order_number': item.order_number,
                    'display_status': item.display_status,
                    'claim_type': item.claim_type,
                    'store_name': item.store_name,
                    'collect_tracking_number': item.collect_tracking_number or '',
                    'collect_delivery_company': item.collect_delivery_company or '',
                    'recipient_name': item.recipient_name or '',
                    'option_code': item.option_code or '',
                    'product_name': item.product_name or '',
                    'option_name': item.option_name or '',
                    'quantity': item.quantity,
                    'invoice_number': item.invoice_number or '',
                    'claim_status': item.claim_status or '',
                    'claim_reason': item.claim_reason or '',
                    'customer_reason': item.customer_reason or '',
                    'return_shipping_charge': item.return_shipping_charge or '',
                    'shipping_charge_payment_method': item.shipping_charge_payment_method or '',
                    'processing_status': item.processing_status or '',
                    'note': item.note or '',
                    'inspector': item.inspector or '',
                    'product_issue': item.product_issue or '',
                    'delivered_date': formatted_delivered,
                    'claim_request_date': item.claim_request_date.isoformat() if item.claim_request_date else '',
                    'display_status': item.display_status
                })

        logger.debug(f"업로드 완료: 성공={len(success_items)}개, 실패={len(error_items)}개")
        return JsonResponse({'success': True, 'items': success_items, 'errors': error_items})

    return JsonResponse({'success': False, 'message': 'Invalid request'})


@login_required
def finalize_excel_import(request):
    """
    여기서 더 이상 모든 데이터를 '미처리'로 업데이트하지 않습니다.
    즉, 기존 데이터 상태를 손상시키지 않고,
    단지 업로드한 데이터가 반품목록에 반영되도록 하려면
    사실상 이 함수에서 할 작업이 없거나 최소화해야 합니다.
    """

    # 필요하다면 특정 조건으로 필터링하여 방금 업로드한 데이터만 상태 변경 가능
    # 예: ReturnItem.objects.filter(...).update(processing_status='미처리')
    # 하지만 문제를 피하기 위해 아무것도 하지 않고 바로 반품목록 페이지로 이동
    messages.success(request, '엑셀 업로드 데이터가 반품목록에 반영되었습니다.')
    return redirect('반품목록')



# views.py 변경 없음 (단, 날짜 파싱 부분 수정)
@csrf_exempt
def upload_reason_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return JsonResponse({'success': False, 'message': '엑셀 파일을 선택해주세요.'})

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
        except Exception as e:
            return JsonResponse({'success': False, 'message': '유효한 엑셀 파일이 아닙니다.'})

        ws = wb.active

        # 헤더 매핑:
        # R(18), D(4), E(5), F(6), K(11)
        order_number_col    = 1   # 주문번호
        claim_status_col    = 2   # 클레임상태
        claim_reason_col    = 3   # 클레임사유
        customer_reason_col = 4   # 고객사유
        claim_request_col   = 5   # 클레임요청일

        success_items = []
        error_items = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            order_number        = row[order_number_col - 1]
            claim_status        = row[claim_status_col - 1]
            claim_reason        = row[claim_reason_col - 1]
            customer_reason     = row[customer_reason_col - 1]
            claim_request_raw   = row[claim_request_col - 1]

            # 기존 코드에서 날짜 파싱 부분 수정
            if isinstance(claim_request_raw, datetime.datetime):
                claim_request_date = claim_request_raw
            elif isinstance(claim_request_raw, str) and claim_request_raw.strip():
                raw_str = claim_request_raw.strip()
                for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]:
                    try:
                        claim_request_date = datetime.datetime.strptime(raw_str, fmt)
                        break
                    except ValueError:
                        continue

            # 날짜를 한국어 포맷으로 변환
            if claim_request_date:
                claim_request_date_str = claim_request_date.strftime("%Y년 %m월 %d일 %I:%M %p")
                claim_request_date_str = claim_request_date_str.replace("AM", "오전").replace("PM", "오후")
            else:
                claim_request_date_str = None

            if not order_number:
                error_items.append({
                    'row': row_idx,
                    'error': '주문번호(R열) 없음'
                })
                continue

            items = ReturnItem.objects.filter(order_number=order_number)
            if not items.exists():
                error_items.append({
                    'row': row_idx,
                    'order_number': order_number,
                    'error': '매칭되는 주문번호 없음'
                })
                continue

            for item in items:
                item.claim_status = claim_status if claim_status else item.claim_status
                item.claim_reason = claim_reason if claim_reason else item.claim_reason
                item.customer_reason = customer_reason if customer_reason else item.customer_reason
                if claim_request_date:
                    item.claim_request_date = claim_request_date
                item.save()

                success_items.append({
                    'order_number': item.order_number,
                    'display_status': item.display_status,
                    'claim_type': item.claim_type,
                    'store_name': item.store_name,
                    'collect_tracking_number': item.collect_tracking_number or '',
                    'collect_delivery_company': item.collect_delivery_company or '',
                    'recipient_name': item.recipient_name or '',
                    'option_code': item.option_code or '',
                    'product_name': item.product_name or '',
                    'option_name': item.option_name or '',
                    'quantity': item.quantity,
                    'invoice_number': item.invoice_number or '',
                    'claim_status': item.claim_status or '',
                    'claim_reason': item.claim_reason or '',
                    'customer_reason': item.customer_reason or '',
                    'return_shipping_charge': item.return_shipping_charge or '',
                    'shipping_charge_payment_method': item.shipping_charge_payment_method or '',
                    'processing_status': item.processing_status or '',
                    'note': item.note or '',
                    'inspector': item.inspector or '',
                    'product_issue': item.product_issue or '',
                    'delivered_date': item.delivered_date.isoformat() if item.delivered_date else '',
                    'claim_request_date': item.claim_request_date.isoformat() if item.claim_request_date else '',
                    'display_status': item.display_status
                })

        return JsonResponse({'success': True, 'items': success_items, 'errors': error_items})
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


@login_required
def download_reason_template(request):
    """
    반품 사유 엑셀 예제 파일을 헤더(주문번호, 클레임상태, 클레임사유, 고객사유, 클레임요청일)만 담아 내려줍니다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "반품사유_업로드템플릿"

    headers = ["주문번호", "클레임상태", "클레임사유", "고객사유", "클레임요청일"]
    ws.append(headers)
    ws.freeze_panes = "A2"  # 헤더 고정

    now = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"반품사유_템플릿_{now}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def scan(request):
    # DB에서 로그인한 사용자(user=request.user)가 등록한 스캔 목록 읽어오기
    scanned_qs = ScanLog.objects.filter(user=request.user, matched=True).order_by('created_at')
    unmatched_qs = ScanLog.objects.filter(user=request.user, matched=False).order_by('created_at')

    # 템플릿에 넘길 리스트
    scanned_numbers = [obj.tracking_number for obj in scanned_qs]
    unmatched_numbers = [obj.tracking_number for obj in unmatched_qs]

    if request.method == 'POST':
        remove_number = request.POST.get('remove_number')
        if remove_number:
            # scanned
            removed = ScanLog.objects.filter(user=request.user,
                                             tracking_number=remove_number).delete()
            if removed[0] > 0:
                messages.success(request, f"{remove_number} 운송장번호를 제거했습니다.")
            else:
                messages.warning(request, f"{remove_number} 운송장번호를 찾을 수 없습니다.")

    return render(request, 'return_process/scan.html', {
        'scanned_numbers': scanned_numbers,
        'unmatched_numbers': unmatched_numbers,
    })


@csrf_exempt
def scan_submit(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            data = request.POST.dict()

        action = data.get('action', None)
        scanned_numbers = request.session.get('scanned_numbers', [])
        unmatched_numbers = request.session.get('unmatched_numbers', [])
        user = request.user

        if not action:
            return JsonResponse({'success': False, 'error': 'action parameter is missing'})

        if action == 'approve_returns':
            # 1) DB에서 user가 matched=True로 저장해둔 운송장번호 목록을 가져옴
            matched_logs = ScanLog.objects.filter(user=user, matched=True)
            scanned_numbers = [log.tracking_number for log in matched_logs]

            # 디버깅용
            print("DEBUG: scanned_numbers at approve_returns:", scanned_numbers)

            if scanned_numbers:
                # 2) 실제 DB에 존재하는 ReturnItem만 필터
                existing_items = ReturnItem.objects.filter(collect_tracking_number__in=scanned_numbers)
                existing_numbers = list(existing_items.values_list('collect_tracking_number', flat=True))
                print("DEBUG: existing_numbers found in DB:", existing_numbers)

                # 3) new_unmatched: 스캔은 했지만 ReturnItem에 없는 번호
                scanned_set = set(scanned_numbers)
                existing_set = set(existing_numbers)
                new_unmatched = list(scanned_set - existing_set)

                # 4) 실제 존재하는 아이템은 수거완료 처리(원하면 '검수완료'로 변경)
                items = ReturnItem.objects.filter(collect_tracking_number__in=existing_numbers)

                for item in items:
                    item.processing_status = '수거완료'
                    item.save()  # 오버라이드된 save() 메서드 실행 -> collected_at 자동 기록
                # 5) matched_logs 전부 삭제 (혹은 남겨두고 싶다면 다른 로직)
                matched_logs.delete()

                # 6) new_unmatched를 DB에 저장(미일치)하거나 로그 업데이트
                #    기존 미일치 로그로 만들기
                for num in new_unmatched:
                    # 이미 unmatched=True로 등록된게 없으면 생성
                    obj, created = ScanLog.objects.get_or_create(user=user, tracking_number=num)
                    if obj.matched:  # 만약 기존이 matched였다면
                        obj.matched = False
                        obj.save()

                if new_unmatched:
                    return JsonResponse({
                        'success': True,
                        'message': '전송 완료하였으나 일부 일치하지 않는 송장이 있습니다.',
                        'unmatched_numbers': new_unmatched
                    })
                else:
                    return JsonResponse({'success': True, 'message': '전송되었습니다.'})
            else:
                return JsonResponse({'success': False, 'message': '스캔한 운송장번호가 없습니다.'})

        elif action == 'recheck_unmatched':
            # 1) DB에서 user가 matched=False인 운송장번호를 확인
            unmatched_logs = ScanLog.objects.filter(user=user, matched=False)
            unmatched_numbers = [log.tracking_number for log in unmatched_logs]

            if unmatched_numbers:
                # 2) 실제 존재하는 ReturnItem
                existing_items = ReturnItem.objects.filter(collect_tracking_number__in=unmatched_numbers)
                existing_numbers = set(existing_items.values_list('collect_tracking_number', flat=True))

                re_matched = []
                for log in unmatched_logs:
                    if log.tracking_number in existing_numbers:
                        log.matched = True
                        log.save()
                        re_matched.append(log.tracking_number)

                if re_matched:
                    return JsonResponse({
                        'success': True,
                        'message': f"일치하지 않던 {len(re_matched)}건의 송장이 다시 확인되어 일치하는 목록으로 이동했습니다."
                    })
                else:
                    return JsonResponse({'success': True, 'message': '일치하는 송장이 없습니다.'})
            else:
                return JsonResponse({'success': True, 'message': '미일치 목록이 비어있습니다.'})

        # 여기서부터 새로운 액션 처리 로직 추가
        elif action == 'update_issue':
            ids = data.get('ids', [])
            product_issue = data.get('product_issue', '').strip()

            if not ids:
                return JsonResponse({'success': False, 'message': 'No item ids provided.'})
            
            # 로그: 전달받은 데이터 확인
            print("DEBUG: update_issue - ids:", ids, "product_issue:", product_issue)

            items = ReturnItem.objects.filter(id__in=ids)
            updated_count = len(items)
            print("DEBUG: 업데이트할 아이템 수:", updated_count)

            for item in items:
                item.product_issue = product_issue
                item.processing_status = '검수완료'
                item.save()
            
            if updated_count > 0:
                return JsonResponse({
                    'success': True,
                    'updated_count': updated_count,
                })
            else:
                return JsonResponse({'success': False, 'message': 'No items updated.'})
        elif action == 'update_order_number':
            item_id = data.get('id')
            new_orderno = data.get('order_number')
            try:
                item = ReturnItem.objects.get(pk=item_id)
                item.order_number = new_orderno
                item.save()
                return JsonResponse({'success': True})
            except ReturnItem.DoesNotExist:
                return JsonResponse({'success': False, 'message': '존재하지 않는 항목입니다.'})

        elif action == 'update_note':
            ids = data.get('ids', [])
            note = data.get('note', '').strip()
            if not ids:
                return JsonResponse({'success': False, 'message': 'No item ids provided.'})

            updated_count = ReturnItem.objects.filter(id__in=ids).update(note=note)
            if updated_count > 0:
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'message': 'No items updated.'})
        
        elif action == 'delete_items':
            ids = data.get('ids', [])
            
            # ids가 비어있는 경우
            if not ids:
                return JsonResponse({'success': False, 'message': '삭제할 항목이 선택되지 않았습니다.'}, status=400)
            
            queryset = ReturnItem.objects.filter(id__in=ids)
            to_delete_count = queryset.count()
            
            if to_delete_count == 0:
                return JsonResponse({'success': False, 'message': '선택된 항목이 존재하지 않습니다.'}, status=400)
            
            # 여기서 실제 삭제
            queryset.delete()
            
            return JsonResponse({'success': True, 'message': f'{to_delete_count}건 삭제 완료'})

        elif action == 'update_quantity':
            item_id = data.get('id', None)
            new_quantity = data.get('new_quantity', None)
            if item_id and new_quantity is not None:
                try:
                    item = ReturnItem.objects.get(id=item_id)
                    item.quantity = new_quantity
                    item.save()
                    return JsonResponse({'success': True})
                except ReturnItem.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Item not found'})
            else:
                return JsonResponse({'success': False, 'message': 'Invalid parameters'})
        elif action == 'update_all_claim_status':
            # 1) 수거완료 상태인 모든 아이템의 주문번호 조회
            items = ReturnItem.objects.filter(processing_status='수거완료')
            if not items.exists():
                return JsonResponse({
                    'success': False,
                    'message': '수거완료 상태인 아이템이 없습니다.'
                }, status=400)

            order_numbers = list(items.values_list('order_number', flat=True))

            # 2) 첫 번째 아이템으로부터 네이버 계정 정보 선택
            first = items.first()
            platform    = first.platform.lower()
            store_name  = first.store_name
            if platform == 'naver':
                target = next(
                    (acc for acc in NAVER_ACCOUNTS if store_name in acc['names']),
                    None
                )
                if not target:
                    return JsonResponse({
                        'success': False,
                        'message': 'NAVER 계정을 찾을 수 없음'
                    }, status=400)
                account_info = target
            else:
                return JsonResponse({
                    'success': False,
                    'message': '지원되지 않는 플랫폼'
                }, status=400)

            # 3) 50개씩 나눠서 API 호출
            MAX_IDS = 50
            all_details = []
            for i in range(0, len(order_numbers), MAX_IDS):
                batch = order_numbers[i:i+MAX_IDS]
                result = get_product_order_details(account_info, batch)
                if not result.get('success'):
                    return JsonResponse({
                        'success': False,
                        'message': result.get('message', 'API 호출 실패')
                    }, status=400)
                all_details.extend(result.get('details', []))

            # 4) 받은 상세정보로 ReturnItem.product_order_status 업데이트
            updated_count = 0
            for detail in all_details:
                oid    = detail.get('productOrderId')
                status = detail.get('productOrderStatus', 'N/A')
                obj = ReturnItem.objects.filter(
                    order_number=oid,
                    processing_status='수거완료'
                ).first()
                if obj:
                    obj.product_order_status = status
                    obj.save()
                    updated_count += 1

            return JsonResponse({
                'success': True,
                'updated_count': updated_count
            })
        elif action == 'get_details':
            item_id = data.get('id', None)
            if not item_id:
                return JsonResponse({'success': False, 'message': 'No item id provided.'})
            try:
                item = ReturnItem.objects.get(id=item_id)
                details = {
                    'order_number': item.order_number,
                    'display_status': item.display_status,
                    'claim_type': item.claim_type,
                    'store_name': item.store_name,
                    'collect_tracking_number': item.collect_tracking_number,
                    'collect_delivery_company': item.collect_delivery_company,
                    'recipient_name': item.recipient_name,
                    'option_code': item.option_code,
                    'product_name': item.product_name,
                    'option_name': item.option_name,
                    'quantity': item.quantity,
                    'invoice_number': item.invoice_number,
                    'claim_status': item.claim_status,
                    'claim_reason': item.claim_reason,
                    'customer_reason': item.customer_reason,
                    'return_shipping_charge': item.return_shipping_charge,
                    'shipping_charge_payment_method': item.shipping_charge_payment_method,
                    'processing_status': item.processing_status,
                    'note': item.note,
                    'inspector': item.inspector,
                    'product_issue': item.product_issue,
                    'delivered_date': item.delivered_date.isoformat() if item.delivered_date else '',
                    'claim_request_date': item.claim_request_date.isoformat() if item.claim_request_date else '',
                }
                return JsonResponse({'success': True, 'details': details})
            except ReturnItem.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Item not found.'})
        
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@csrf_exempt
@login_required
def check_number_submit(request):
    """
    스캔된 운송장번호 목록을 DB(ScanLog, ReturnItem)로 관리하는 예시.
    스캔 시점에:
      1) collect_tracking_number가 존재하면 ReturnItem의 상태를 '검수완료'(또는 '수거완료')로 변경
      2) ScanLog에 matched=True로 등록
    존재하지 않으면 matched=False로 등록
    """
    if request.method == 'POST':
        user = request.user
        data = json.loads(request.body.decode('utf-8'))
        numbers = data.get('numbers', [])

        # 이미 DB에 저장된 스캔 정보 읽어오기
        existing_scanned = set(
            ScanLog.objects.filter(user=user, matched=True)
                           .values_list('tracking_number', flat=True)
        )
        existing_unmatched = set(
            ScanLog.objects.filter(user=user, matched=False)
                           .values_list('tracking_number', flat=True)
        )

        matched_list = []
        unmatched_list = []

        for num in numbers:
            number = num.strip()
            if not number:
                continue  # 공백만 있는 경우 건너뜀

            # 실제로 ReturnItem(collect_tracking_number=number) 존재 여부 체크
            item_qs = ReturnItem.objects.filter(collect_tracking_number=number)
            exists = item_qs.exists()
            if exists:

                item_qs.update(processing_status='검수완료')

                # ====== ScanLog 등록 (matched=True) ======
                if number not in existing_scanned and number not in existing_unmatched:
                    ScanLog.objects.create(
                        user=user,
                        tracking_number=number,
                        matched=True
                    )
                    existing_scanned.add(number)

                matched_list.append(number)

            else:
                # ====== DB에 없음 → 미일치 처리 ======
                if number not in existing_scanned and number not in existing_unmatched:
                    ScanLog.objects.create(
                        user=user,
                        tracking_number=number,
                        matched=False
                    )
                    existing_unmatched.add(number)

                unmatched_list.append(number)

        return JsonResponse({
            'success': True,
            'matched_list': matched_list,
            'unmatched_list': unmatched_list
        })

    else:
        return JsonResponse({'success': False, 'message': 'Only POST allowed'}, status=405)   
            
@login_required
def download_unmatched(request):
    unmatched_logs = ScanLog.objects.filter(user=request.user, matched=False)
    # 미일치 리스트를 엑셀로 다운로드
    unmatched_numbers = [log.tracking_number for log in unmatched_logs]

    wb = Workbook()
    ws = wb.active
    ws.title = "Unmatched"

    # 헤더


    row = 1
    for num in unmatched_numbers:
        ws.cell(row=row, column=1, value=num)
        row += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="unmatched_numbers.xlsx"'
    wb.save(response)
    return response

import json
import requests
from django.shortcuts import render
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import ReturnItem

@login_required
def collected_items(request):
    """'수거완료' 상태인 아이템 목록을 보여주는 뷰."""
    # '수거완료' 상태인 아이템들 조회
    items = ReturnItem.objects.filter(processing_status='수거완료')

    # GET 요청: 템플릿 렌더링
    return render(request, 'return_process/collected_items.html', {
        'items': items,
    })


@login_required
def collected_items(request):
    items = ReturnItem.objects.filter(processing_status='수거완료')
    # 지원되는 스토어 목록
    store_code_map = {
        "니뜰리히": "wcsrr1",
        "수비다": "wce1wv",
        "노는개최고양": "w4g8ot",
        "아르빙": "w48val",
    }
    store_code_map_keys = list(store_code_map.keys())
    return render(request, 'return_process/collected_items.html', {
        'items': items,
        'store_code_map_keys': store_code_map_keys,
    })


@login_required
@require_GET
def biztalk_proxy(request, order_number):
    api_url = f"https://sell.smartstore.naver.com/o/v3/oa/biztalk/{order_number}/"
    
    # 테스트용: 제공해주신 쿠키 값을 직접 사용 (실제 운영 시에는 보안에 유의)
    cookies = {
        "NID_AUT": "/2ikLx0V6BZ5kjCX+R5MGj3Kw7ts6ecuCtb3Y7Xi3uleywShx+92uVLaejEJYmjP",
        "NID_SES": "AAAB3T4gvyZacpRqa7/stO/2Apkgb83sa76RPM7tShUlJmInqkmgU+4TOnUXopp1mQwLtMdZXp+oqpqFWM/3oIo7T4Hko4ejPUgstlc9wvFv9zysd3KlOyfCMNnpWx2fcp8SBHB8iIDc/5k44SoqPWBGQg5hwJctBa7j4Gp+UUypvMtbOTXQEdHGN6HLvFRWGNu060hZZiIo4CeEefg2sM/vJNtK2k9dvGUoE3EDYRplkRA2NyUl7IRAL+PTnzaWKpYkw8E9emS8erAACpeDRVNE65RnC4SFxA5/smKNNOL2ldhjCGGzhRUlREGk80t8TmJMaQjKORhAoNrvAdXUaM1tscp+IEt5XfIfDKh/GsiyjnBL5R6W451GeOUXtG6YIlTUURp36geqQrFDsL+Miq4JSmy0dMUm16m+29vqMC8dNLCjyv5Bjnja+Gw9iomJbhXMl10qcPaLSSpoJ9N87aADXa25VEWFtwTz/gsotRa9RbjjVEs193MDLPFAK9zxjg4cqYYzIDee0ehHCv+I5V4WtHbm2KUPqWocqalRVtBj9SjBi0+ftae5YZA9LRgw=="
    }
    
    try:
        response = requests.get(api_url, cookies=cookies)
        logger.debug("SmartStore API URL: %s", api_url)
        logger.debug("Response status code: %s", response.status_code)
        logger.debug("Response content: %s", response.text.strip()[:200])
        
        response.raise_for_status()
        
        # HTML 로그인 페이지 감지
        if response.text.strip().startswith("<!DOCTYPE html>"):
            logger.debug("HTML 로그인 페이지 반환 감지")
            return JsonResponse({
                "bSuccess": False,
                "message": "세션 등록 필요",
                "content": response.text[:500]
            }, status=200)
        
        try:
            data = response.json()
        except Exception as e:
            logger.exception("JSON parsing error: %s", e)
            return JsonResponse({'error': 'JSON decode error', 'content': response.text[:500]}, status=500)
        
        return JsonResponse(data)
    except Exception as e:
        logger.exception("Error calling SmartStore API: %s", e)
        return JsonResponse({'error': str(e)}, status=500)
    

@login_required
def inspect(request, item_id):
    item = ReturnItem.objects.get(id=item_id)
    if request.method == 'POST':
        inspection_result = request.POST.get('inspection_result')
        
        # inspection_result는 '오배송', '이상없음', '파손 및 불량' 중 하나라고 가정
        # product_issue 필드에 해당 값을 그대로 저장 가능
        item.product_issue = inspection_result

        # 검수 완료 시 processing_status를 '검수완료'로 업데이트
        item.processing_status = '검수완료'

        # 검수자 이름을 추가하고 싶다면 inspector 필드에도 저장 가능
        # item.inspector = request.user.username 등 필요한 값 설정
        
        item.save()
        return redirect('검수완료')  # 'inspected_items' 뷰의 URL name

    return render(request, 'return_process/inspect.html', {'item': item})

@login_required
def inspected_items(request):
    items = ReturnItem.objects.filter(processing_status='검수완료')

    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        action = data.get('action', None)  # <-- 안전하게 선언

        if action == 'update_all_claim_status':
            # --- 기존 코드 (수정 불가) ---
            if not items.exists():
                return JsonResponse({'success': False, 'message': '검수완료 상태인 아이템이 없습니다.'}, status=400)

            from collections import defaultdict
            grouped_by_account = defaultdict(list)
            for item in items:
                platform = item.platform.lower()
                store_name = item.store_name
                if platform == 'naver':
                    target_account = next((acc for acc in NAVER_ACCOUNTS if store_name in acc['names']), None)
                    if not target_account:
                        continue
                    account_key = (platform, store_name)
                    grouped_by_account[account_key].append(item.order_number)
                else:
                    pass

            MAX_IDS = 50
            total_updated_count = 0

            for (platform, store_name), order_numbers in grouped_by_account.items():
                account_info = next((acc for acc in NAVER_ACCOUNTS if store_name in acc['names']), None)
                if not account_info:
                    continue

                chunked_details = []
                for i in range(0, len(order_numbers), MAX_IDS):
                    batch = order_numbers[i:i+MAX_IDS]
                    batch_result = get_product_order_details(account_info, batch)
                    if not batch_result.get('success'):
                        return JsonResponse({'success': False, 'message': batch_result.get('message')}, status=400)
                    chunked_details.extend(batch_result.get('details', []))

                for detail in chunked_details:
                    order_id = detail.get('productOrderId')
                    status = detail.get('productOrderStatus', 'N/A')
                    item = ReturnItem.objects.filter(order_number=order_id, processing_status='검수완료').first()
                    if item:
                        item.product_order_status = status
                        item.save()
                        total_updated_count += 1

            return JsonResponse({'success': True, 'updated_count': total_updated_count})

        elif action == 'back_to_collected':
            ids = data.get('ids', [])
            ReturnItem.objects.filter(id__in=ids, processing_status='검수완료').update(processing_status='수거완료')
            return JsonResponse({'success': True})

        elif action == 'dispatch_exchange':
            print("[DEBUG] dispatch_exchange action 호출됨.")
            print(f"[DEBUG] request.POST or body data: {data}")

            product_order_id = data.get('product_order_id')
            re_delivery_method = data.get('re_delivery_method', 'DELIVERY')
            re_delivery_company = data.get('re_delivery_company') or None
            re_delivery_tracking_number = data.get('re_delivery_tracking_number') or None

            print(f"[DEBUG] 전달받은 product_order_id={product_order_id}, "
                  f"re_delivery_method={re_delivery_method}, "
                  f"re_delivery_company={re_delivery_company}, "
                  f"re_delivery_tracking_number={re_delivery_tracking_number}")

            # 디버깅: DB에 어떤 레코드가 있는지 한번 출력해보기
            count_all = ReturnItem.objects.filter(order_number=product_order_id).count()
            count_in_inspected = ReturnItem.objects.filter(order_number=product_order_id, processing_status='검수완료').count()
            print(f"[DEBUG] order_number={product_order_id} 전체 매칭 개수: {count_all}")
            print(f"[DEBUG] order_number={product_order_id} + 검수완료 매칭 개수: {count_in_inspected}")

            # item 조회
            item_obj = ReturnItem.objects.filter(order_number=product_order_id, processing_status='검수완료').first()
            if not item_obj:
                print("[DEBUG] 400 반환: 해당 주문번호(검수완료) 아이템을 찾을 수 없습니다.")
                return JsonResponse({'success': False, 'message': '해당 주문번호(검수완료) 아이템을 찾을 수 없습니다.'}, status=400)

            # >>> 여기서 실제 DB에서 찾은 item의 order_number를 찍어볼 수 있음 <<<
            print(f"[DEBUG] 찾은 item: ID={item_obj.id}, order_number={item_obj.order_number}")

            platform = item_obj.platform.lower()
            print(f"[DEBUG] 찾은 item의 platform={platform}, store_name={item_obj.store_name}")

            if platform != 'naver':
                print(f"[DEBUG] 400 반환: 네이버가 아닌 플랫폼({platform})")
                return JsonResponse({'success': False, 'message': f'네이버 플랫폼이 아니므로 재배송 처리를 지원하지 않습니다. ({platform})'}, status=400)

            store_name = item_obj.store_name
            account_info = next((acc for acc in NAVER_ACCOUNTS if store_name in acc['names']), None)
            if not account_info:
                print("[DEBUG] 400 반환: NAVER 계정을 찾을 수 없음")
                return JsonResponse({'success': False, 'message': 'NAVER 계정을 찾을 수 없음'}, status=400)

            # 여기서 dispatch_naver_exchange 함수 콜
            success, message = dispatch_naver_exchange(
                account_info=account_info,
                product_order_id=product_order_id,
                re_delivery_method=re_delivery_method,
                re_delivery_company=re_delivery_company,
                re_delivery_tracking_number=re_delivery_tracking_number
            )

            print(f"[DEBUG] 교환 재배송 처리 결과: success={success}, message={message}")

            # --- "기발송건" 로직 추가 ---
            if success:
                # 정상 성공: 로컬 상태를 '반품완료'로 변경
                item_obj.processing_status = '반품완료'
                item_obj.save()
                return JsonResponse({'success': True, 'message': message})
            else:
                # 실패했지만 메시지에 "기발송건입니다."가 있다면, 이미 발송이 된 것이므로
                if "기발송건입니다." in message:
                    item_obj.processing_status = '반품완료'
                    item_obj.save()
                    return JsonResponse({
                        'success': True,
                        'message': "이미 발송된 주문 건이므로 로컬상태를 '반품완료'로 변경했습니다."
                    })
                else:
                    # 실제 오류는 그대로 실패 처리
                    return JsonResponse({'success': False, 'message': f'교환 재배송 처리 실패: {message}'})

    # GET 요청: 전달할 컨텍스트에 store_code_map_keys 추가
    store_code_map = {
        "니뜰리히": "wcsrr1",
        "수비다": "wce1wv",
        "노는개최고양": "w4g8ot",
        "아르빙": "w48val",
    }
    store_code_map_keys = list(store_code_map.keys())

    return render(request, 'return_process/inspected_items.html', {
        'items': items,
        'store_code_map_keys': store_code_map_keys,
    })


def inspected_export_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        ids = data.get('ids', [])

        # items 가져오기
        items = ReturnItem.objects.filter(id__in=ids)

        # openpyxl로 엑셀 파일 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "선택 항목"

        # 헤더 작성
        ws.append([
            "ID", "주문번호", "클레임 종류", "검수 결과", 
            "클레임 사유", "고객 사유", "사유(노트)", "수거 송장번호",
            "수거 배송비", "배송비 지급 방식", "스토어명", 
            "수취인명", "연락처", "옵션코드", 
            "상품명", "옵션명", "수량", "배송 완료일", 
            "클레임 요청일", "업데이트 날짜"
        ])
        for item in items:
            ws.append([
                item.id,  # ID
                item.order_number,  # 주문번호
                item.claim_type,  # 클레임 종류
                item.product_issue,  # 검수 결과
                item.claim_reason,  # 클레임 사유
                item.customer_reason,  # 고객 사유
                item.note,  # 사유(노트)
                item.collect_tracking_number,  # 수거 송장번호
                item.return_shipping_charge,  # 수거 배송비
                item.shipping_charge_payment_method,  # 배송비 지급 방식
                item.store_name,  # 스토어명
                item.recipient_name,  # 수취인명
                item.recipient_contact,  # 연락처
                item.option_code,  # 옵션코드
                item.product_name,  # 상품명
                item.option_name,  # 옵션명
                item.quantity,  # 수량
                item.delivered_date.strftime("%Y-%m-%d") if item.delivered_date else "",  # 배송 완료일
                item.claim_request_date.strftime("%Y-%m-%d") if item.claim_request_date else "",  # 클레임 요청일
                item.inspected_at.strftime("%Y-%m-%d %H:%M:%S") if item.inspected_at else ""  # 업데이트 날짜
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=selected_items.xlsx'
        return response

    # POST 외 다른 메서드일 때 처리
    return JsonResponse({"error": "POST only"}, status=400)

@csrf_exempt
def send_shipping_sms(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        items = data.get('items', [])

        account = settings.SOLAPI_ACCOUNTS[0]
        api_key = account['api_key']
        api_secret = account['api_secret']
        sender = account['sender']

        url = "https://api.solapi.com/messages/v4/send-many"

        # 메시지 목록 구성
        message_list = []
        for item in items:
            recipient_contact = item['recipient_contact']
            store_name = item['store_name']
            return_shipping_charge = item['return_shipping_charge']

            text = f"""[반품 배송비 송금 안내드립니다]

안녕하세요, {store_name}입니다.
먼저 저희 제품을 이용해주셔서 진심으로 감사드리며, 반품 관련하여 안내드립니다.

고객님께서 반품 요청시 선택하신 **판매자에게 직접 송금**에 따라 아래 계좌로 송금 부탁드립니다.

🔹 송금 계좌
    • 은행명: 농협
    • 예금주: 양태영
    • 계좌번호: 301-5003-6206-81

🔹 송금 금액
    • {return_shipping_charge}원

송금 후 확인이 되면, 빠르게 반품 처리를 진행하도록 하겠습니다.
혹시 반품 관련하여 추가 문의사항이 있으시면, 02-338-9465로 연락 주시기 바랍니다.

항상 고객님의 만족을 위해 최선을 다하겠습니다.
감사합니다.
{store_name} 드림
"""
            
            message_list.append({
                "to": recipient_contact,
                "from": sender,
                "text": text,
                "type": "LMS"
            })

        payload = {"messages": message_list}

        # 인증에 필요한 date와 salt 생성
        # Google Apps Script 로직과 동일하게 UTC ISO8601 시간 사용
        date = datetime.datetime.utcnow().isoformat(timespec='seconds') + 'Z'  
        # isoformat(timespec='seconds') 사용 후 'Z'추가로 UTC 표기
        salt = uuid.uuid4().hex
        hmac_data = date + salt

        # HMAC-SHA256 서명 생성
        signature_bytes = hmac.new(
            api_secret.encode('utf-8'),
            hmac_data.encode('utf-8'),
            hashlib.sha256
        ).digest()

        # 바이트 배열을 소문자 hex 문자열로 변환
        signature = signature_bytes.hex()

        # Authorization 헤더 설정 (Google Apps Script 코드와 유사한 형식)
        authorization_header = f"HMAC-SHA256 apiKey={api_key}, date={date}, salt={salt}, signature={signature}"

        headers = {
            "Content-Type": "application/json",
            "Authorization": authorization_header
        }

        response = requests.post(url, headers=headers, data=json.dumps(payload))

        if response.status_code == 200:
            return JsonResponse({"success": True, "message": "문자 발송 성공"})
        else:
            return JsonResponse({"success": False, "message": f"문자 발송 실패: {response.text}"})
    else:
        return JsonResponse({"success": False, "message": "Invalid request method"}, status=405)


@login_required
def process_return(request, item_id):
    print("process_return 함수 호출됨")  # 디버깅용 로그
    item = get_object_or_404(ReturnItem, id=item_id)
    print("플랫폼:", item.platform, "스토어명:", item.store_name, "주문번호:", item.order_number)  # 디버깅용 로그

    # 스토어명 매핑
    store_name_map = {
        "노는 개 최고양": "노는개최고양",
        "수비다 SUBIDA": "수비다"
        # 필요 시 추가
    }

    # 매핑 후 스토어명
    mapped_name = store_name_map.get(item.store_name, item.store_name)

    # 네이버 스토어 목록
    naver_stores = ["니뜰리히", "노는개최고양", "아르빙", "수비다"]

    # 쿠팡인 경우
    if item.platform.lower() == 'coupang':
        print("쿠팡 반품 승인 로직 시작")  # 디버깅용 로그
        item.processing_status = '반품완료'
        item.save()
        return True, "쿠팡 반품완료 처리 성공"

    # 네이버 스토어인지 체크 (매핑된 스토어명으로 확인)
    elif mapped_name in naver_stores:
        print("네이버 스토어로 인식됨:", mapped_name)

        # NAVER_ACCOUNTS에서 매핑된 스토어명으로 계정 찾기
        target_account = next((acc for acc in NAVER_ACCOUNTS if mapped_name in acc['names']), None)
        if not target_account:
            print("해당 이름의 NAVER 계정을 찾지 못했습니다.")
            return False, "NAVER 계정을 찾을 수 없음"

        account_info = target_account

        # 네이버 반품 승인 API 호출 여부 결정 (예시)
        condition_call_api = (
            item.claim_type in ['RETURN', 'N/A', '반품'] and 
            item.product_order_status in ['DELIVERED', 'DELIVERING', 'PURCHASE_DECIDED', '배송 중', '배송 완료', '구매 확정', 'N/A']
        )

        if condition_call_api:
            print("네이버 반품 승인 시작")  # 디버깅용 로그
            success, message = approve_naver_return(account_info, item.order_number)
            print("반품 승인 결과:", success, message)  # 디버깅용 로그

            if not success:
                print("네이버 반품 승인 실패:", message)
                return False, message
        else:
            print("API 호출 없이 바로 반품완료 처리")
            message = "API 호출 없이 반품완료 처리"

        # 여기까지 왔다면 (API 성공 or API 스킵) -> 반품완료 처리
        order_prefix_12 = item.order_number[:12]
        print(f"DEBUG: order_prefix_12 = {order_prefix_12}")

        related_items = ReturnItem.objects.filter(order_number__startswith=order_prefix_12)
        print("연관 아이템들(12자리 기준):", related_items)

        for ri in related_items:
            ri.processing_status = '반품완료'
            ri.save()

        return True, "반품완료 처리 성공"

    else:
        # 그 외 스토어 (naver_stores 에도 없고, 쿠팡도 아님)
        print(f"그 외 스토어 ({item.store_name} -> {mapped_name}), 직접 반품완료 처리")
        item.processing_status = '반품완료'
        item.save()
        return True, f"{mapped_name} 반품완료 처리 성공"




@login_required
def process_return_bulk(request):
    """
    여러 아이템을 한꺼번에 '반품승인' 처리하는 뷰.
    내부적으로 process_return()을 호출.
    """
    if request.method == 'POST':
        item_ids = request.POST.getlist('item_ids')
        item_ids = [int(x) for x in item_ids if x.strip().isdigit()]

        success_list = []
        error_list = []

        for i_id in item_ids:
            success, msg = process_return(request, i_id)
            if success:
                success_list.append((i_id, msg))
            else:
                error_list.append((i_id, msg))

        if error_list:
            # 하나라도 오류가 있다면 실패로 처리
            return JsonResponse({
                'success': False,
                'errors': [{'id': eid, 'message': emsg} for eid, emsg in error_list]
            })
        else:
            return JsonResponse({
                'success': True,
                'message': '모든 반품승인 처리가 성공적으로 완료되었습니다.'
            })

    return JsonResponse({'success': False, 'message': 'Invalid request'})

@login_required
def returned_items(request):
    # 반품 완료된 상품 목록 표시
    items = ReturnItem.objects.filter(processing_status='반품완료')
    return render(request, 'return_process/returned_items.html', {'items': items})


@csrf_exempt
def update_claim_type_bulk(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except Exception as e:
            return JsonResponse({"success": False, "message": f"JSON parse error: {e}"}, status=400)

        action = data.get('action', '')
        if action != 'update_claim_type':
            return JsonResponse({"success": False, "message": "Invalid action"}, status=400)

        claim_type = data.get('claim_type', '').strip()
        ids = data.get('ids', [])

        if not ids:
            return JsonResponse({"success": False, "message": "No ids provided"}, status=400)

        # 실제 DB 모델 임포트 (예시)
        from .models import ReturnItem

        # claim_type이 N/A(대소문자 구분X) 이거나 빈 문자열('')인 아이템만 필터
        # Q(..., claim_type__iexact='N/A') OR Q(..., claim_type='')
        queryset = ReturnItem.objects.filter(
            id__in=ids
        ).filter(
            Q(claim_type__iexact='N/A') | Q(claim_type='')
        )

        updated_count = queryset.update(claim_type=claim_type)

        return JsonResponse({
            "success": True,
            "updated_count": updated_count
        })

    return JsonResponse({"success": False, "message": "Method not allowed"}, status=405)

    

@login_required
def update_stock_bulk(request):
    """
    체크박스 선택 후 '재고반영' 버튼을 누르면,
    선택된 ReturnItem들의 processing_status를 '재고반영'으로 업데이트하고
    '재고반영완료' 페이지로 이동
    """
    if request.method == 'POST':
        # 1) 체크된 아이템 id 리스트 가져오기
        item_ids = request.POST.getlist('item_ids', [])

        # 2) 이 예시에서는 별다른 로직 없이 그냥 status만 재고반영으로 업데이트
        ReturnItem.objects.filter(id__in=item_ids).update(processing_status='재고반영')

        # 3) 메시지 띄우고 싶다면:
        messages.success(request, f"총 {len(item_ids)}건의 아이템이 재고반영 되었습니다.")

        # 4) 재고반영완료 페이지로 이동
        return redirect('재고반영완료')
    else:
        # GET으로 접근했을 경우 (직접 URL 치고 들어오는 등)은
        # 적절히 다른 페이지로 보내거나 처리
        return redirect('returned_items')  

@login_required
def update_stock(request, item_id):
    item = get_object_or_404(ReturnItem, id=item_id)
    # 셀러툴 API 호출 로직
    url = 'https://api.sellertool.com/update_stock'
    data = {
        'product_id': item.product_id,
        'quantity': 1,  # 재고 수량 조정 로직에 따라 변경
    }
    try:
        response = requests.post(url, data=data)
        response.raise_for_status()  # HTTP 오류 발생 시 예외 발생
    except requests.exceptions.RequestException as e:
        messages.error(request, f"재고 업데이트 실패: {e}")
        return redirect('returned_items')  # 또는 적절한 페이지로 리다이렉트

    # API 호출 성공 시 processing_status 업데이트
    item.processing_status = '재고반영'
    item.save()
    messages.success(request, f"재고가 성공적으로 반영되었습니다: {item.order_number}")
    return redirect('stock_updated_items')


@login_required
def stock_updated_items(request):
    # DB에서 처리상태가 '재고반영'인 항목만 가져옴
    items = ReturnItem.objects.filter(processing_status='재고반영')
    return render(request, 'return_process/stock_updated_items.html', {'items': items})
@login_required
def update_complete_bulk(request):
    """
    체크박스 선택 후 '처리완료' 버튼을 누르면,
    선택된 ReturnItem들의 processing_status를 '처리완료'로 업데이트하고
    처리완료 목록 페이지로 이동
    """
    if request.method == 'POST':
        # 1) POST로 넘어온 체크박스 선택 아이템들 ID 목록
        item_ids = request.POST.getlist('item_ids', [])
        
        # 2) 한 번에 상태 업데이트
        ReturnItem.objects.filter(id__in=item_ids).update(processing_status='처리완료')

        # 3) 메시지 띄우기 (선택)
        messages.success(request, f"총 {len(item_ids)}건의 아이템이 처리완료 되었습니다.")
        
        # 4) 처리완료 목록 페이지로 이동
        return redirect('처리완료')  
        # ↑ urls.py에서 name='처리완료' 로 등록된 페이지로 이동

    else:
        # GET으로 접근하면 그냥 다른 페이지로 돌려보냄 (취향에 맞게)
        return redirect('stock_updated_items')  
    
@login_required
def completed_items(request):
    items = ReturnItem.objects.filter(processing_status='처리완료')
    return render(request, 'return_process/completed_items.html', {'items': items})

@login_required
def upload_returns(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        df = pd.read_excel(excel_file)
        for _, row in df.iterrows():
            ReturnItem.objects.update_or_create(
                order_id=row['주문번호'],
                defaults={
                    'customer_name': row['고객명'],
                    'product_name': row['상품명'],
                    'tracking_number': row['운송장번호'],
                    'status': 'pending',
                }
            )
        return redirect('return_list')
    return render(request, 'return_process/upload_returns.html')

def signup(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)  # 회원가입 후 자동 로그인
            return redirect('반품목록')  # 회원가입 후 이동할 페이지
    else:
        form = CustomUserCreationForm()
    return render(request, 'return_process/signup.html', {'form': form})

def home(request):
    if request.user.is_authenticated:
        return redirect('반품목록')  # 로그인된 사용자는 반품목록으로 이동
    else:
        return redirect('login')  # 로그인되지 않은 사용자는 로그인 페이지로 이동


@login_required
def download_returned_items(request):
    """
    체크박스 선택된 반품완료 목록을 엑셀 파일로 다운로드.
    """
    if request.method == 'POST':
        # 1) JSON 파싱
        try:
            data = json.loads(request.body.decode('utf-8'))
        except Exception as e:
            return JsonResponse({"error": f"JSON 파싱 오류: {str(e)}"}, status=400)
        
        item_ids = data.get('item_ids', [])
        if not item_ids:
            return JsonResponse({"error": "item_ids가 비어 있음"}, status=400)

        # 2) DB에서 선택된 id만 가져오기
        items = ReturnItem.objects.filter(id__in=item_ids)

        # 3) openpyxl 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "반품완료목록"

        # 4) 헤더 작성 (테이블에 보이는 순서대로)
        headers = [
            "검수 결과", "현재클레임상태", "클레임 종류", "클레임 사유",
            "고객 사유", "사유(노트)", "수거 배송비", "배송비 지급 방식",
            "주문번호", "스토어명", "수취인명", "연락처",
            "옵션코드", "상품명", "옵션명", "수량",
        ]
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # 5) 행 데이터 작성
        for row_num, item in enumerate(items, start=2):
            ws.cell(row=row_num, column=1,  value=item.product_issue or "미검수")
            ws.cell(row=row_num, column=2,  value=item.product_order_status or "N/A")
            ws.cell(row=row_num, column=3,  value=item.claim_type or "N/A")
            ws.cell(row=row_num, column=4,  value=item.claim_reason or "N/A")
            ws.cell(row=row_num, column=5,  value=item.customer_reason or "N/A")
            ws.cell(row=row_num, column=6,  value=item.note or "")
            ws.cell(row=row_num, column=7,  value=item.return_shipping_charge or "0")
            ws.cell(row=row_num, column=8,  value=item.shipping_charge_payment_method or "N/A")
            ws.cell(row=row_num, column=9,  value=item.order_number)
            ws.cell(row=row_num, column=10, value=item.store_name or "N/A")
            ws.cell(row=row_num, column=11, value=item.recipient_name or "N/A")
            ws.cell(row=row_num, column=12, value=item.recipient_contact or "N/A")
            ws.cell(row=row_num, column=13, value=item.option_code or "N/A")
            ws.cell(row=row_num, column=14, value=item.product_name or "N/A")
            ws.cell(row=row_num, column=15, value=item.option_name or "N/A")
            ws.cell(row=row_num, column=16, value=item.quantity or 0)

        # 6) 메모리에 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 7) HttpResponse로 바이너리 전송
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="returned_items.xlsx"'
        return response

    else:
        # GET 등 다른 메서드로 접근 시 에러 or 처리
        return JsonResponse({"error": "POST only"}, status=400)

# 대시보드
@login_required
def return_dashboard(request):
    """
    - 판매량(기간) : get_hourly_sales_for_period()
    - 반품(기간)   : ReturnItem (기준일= COALESCE(claim_request_date, last_update_date))
    - 반품률       : (반품수량 ÷ 판매수량) × 100
    - 클레임 사유 (일자별)
    - 클레임 사유 (전체기간 총합)
    """
    if request.method == 'POST':
        # 1) 폼 입력
        start_date_str = request.POST.get('start_date', '')
        end_date_str   = request.POST.get('end_date', '')
        codes_str      = request.POST.get('option_codes', '')

        option_codes = [c.strip() for c in codes_str.split(',') if c.strip()]

        # 2) 날짜 변환
        try:
            start_dt = pd.to_datetime(start_date_str)
            end_dt   = pd.to_datetime(end_date_str)
        except:
            context = {
                'daily_data': [],
                'daily_claims': [],
                'overall_reasons': [],
                'total_claims': 0,
                'return_rate': 0,
                'start_date': start_date_str,
                'end_date': end_date_str,
                'option_codes': codes_str
            }
            return render(request, 'return_process/return_dashboard.html', context)

        # start_dt > end_dt 시 swap
        if start_dt > end_dt:
            start_dt, end_dt = end_dt, start_dt

        # ─────────────────────────────────────────────
        # (A) 기간별 "판매량" 구하기
        # ─────────────────────────────────────────────
        sales_data = get_hourly_sales_for_period(
            option_codes=option_codes,
            start_date=start_dt,
            end_date=end_dt
        )
        # sales_data 예시:
        # [
        #   {"date":"2025-01-05", "daily_total_units":10, "detail":[...]},
        #   ...
        # ]

        # ─────────────────────────────────────────────
        # (B) 기간별 "반품(클레임)" 조회 (ReturnItem)
        # ─────────────────────────────────────────────
        # 1) effective_date = COALESCE(claim_request_date, last_update_date)
        # 2) 기간 필터 : [start_dt, end_dt]
        # 3) 옵션코드 필터 (있다면)
        qs = ReturnItem.objects.annotate(
            effective_date=Coalesce('claim_request_date', 'last_update_date', output_field=DateTimeField())
        ).filter(
            effective_date__range=[start_dt, end_dt]
        )
        if option_codes:
            qs = qs.filter(option_code__in=option_codes)

        # (B-1) 일자별 반품 수량
        grouped_return = qs.annotate(
            date_only=TruncDate('effective_date')
        ).values('date_only') \
         .annotate(daily_qty=Sum('quantity')) \
         .order_by('date_only')

        daily_return_map = {}
        for row in grouped_return:
            d = row['date_only']   # date obj
            qty = row['daily_qty'] or 0
            d_str = d.strftime('%Y-%m-%d')
            daily_return_map[d_str] = qty

        # (B-2) 일자별 "클레임 사유" count (id 기준)
        from collections import defaultdict
        grouped_reason = qs.annotate(
            date_only=TruncDate('effective_date')
        ).values('date_only', 'claim_reason') \
         .annotate(count=Count('id')) \
         .order_by('date_only', 'claim_reason')

        daily_claim_map = defaultdict(lambda: defaultdict(int))
        date_list_claim = set()

        for row in grouped_reason:
            d = row['date_only']
            reason = row['claim_reason'] or '기타/미입력'
            cnt = row['count']
            d_str = d.strftime('%Y-%m-%d')
            daily_claim_map[d_str][reason] += cnt
            date_list_claim.add(d_str)

        date_list_claim = sorted(list(date_list_claim))

        # ─────────────────────────────────────────────
        # (C) 판매 vs 반품 (일자별)
        # ─────────────────────────────────────────────
        daily_data = []
        total_sales_qty  = 0
        total_return_qty = 0

        for item in sales_data:
            d_str = item['date']
            s_qty = item['daily_total_units']
            r_qty = daily_return_map.get(d_str, 0)

            daily_data.append({
                'date': d_str,
                'sales_qty': s_qty,
                'return_qty': r_qty
            })
            total_sales_qty  += s_qty
            total_return_qty += r_qty

        # 전체 반품률(%) = (total_return_qty ÷ total_sales_qty) × 100
        if total_sales_qty > 0:
            return_rate = (total_return_qty / total_sales_qty) * 100
        else:
            return_rate = 0

        # ─────────────────────────────────────────────
        # (D) 일자별 클레임 사유 비중
        # ─────────────────────────────────────────────
        daily_claims = []
        for d_str in date_list_claim:
            reason_map = daily_claim_map[d_str]  # { reason: count, ...}
            total_cnt = sum(reason_map.values())

            reason_list = []
            if total_cnt > 0:
                for rr, cnt in reason_map.items():
                    ratio = (cnt / total_cnt) * 100
                    reason_list.append({
                        'reason': rr,
                        'count': cnt,
                        'ratio': ratio
                    })

            daily_claims.append({
                'date': d_str,
                'total': total_cnt,
                'reasons': reason_list
            })

        # ─────────────────────────────────────────────
        # (E) "전체 기간" 클레임 사유 총합
        # ─────────────────────────────────────────────
        #   ex) [ {'claim_reason':'색상 변경', 'count':1}, ... ]
        overall_reason_qs = qs.values('claim_reason') \
                              .annotate(count=Count('id')) \
                              .order_by('claim_reason')
        total_claims = sum(row['count'] for row in overall_reason_qs)

        overall_reasons = []
        if total_claims > 0:
            for row in overall_reason_qs:
                rr = row['claim_reason'] or '기타/미입력'
                cnt = row['count']
                ratio = (cnt / total_claims) * 100
                overall_reasons.append({
                    'reason': rr,
                    'count': cnt,
                    'ratio': ratio
                })

        # ─────────────────────────────────────────────
        # (F) 템플릿 컨텍스트
        # ─────────────────────────────────────────────
        context = {
            # 판매 vs 반품
            'daily_data': daily_data,
            'total_sales_qty': total_sales_qty,
            'total_return_qty': total_return_qty,
            'return_rate': return_rate,

            # 일자별 클레임 사유
            'daily_claims': daily_claims,

            # 전체 기간 클레임 사유
            'overall_reasons': overall_reasons,
            'total_claims': total_claims,

            # 폼 입력
            'start_date': start_date_str,
            'end_date': end_date_str,
            'option_codes': codes_str
        }
        return render(request, 'return_process/return_dashboard.html', context)

    else:
        # GET
        return render(request, 'return_process/return_dashboard.html')
  
from decouple import config
from django.views import View  # FBV가 아닌 CBV로 예시
from .utils import (
    generate_sellertool_signature,
    convert_return_item_to_formdata
)
from concurrent.futures import ThreadPoolExecutor, as_completed
from django.db import transaction

class SendReturnItemsView(View):
    def post(self, request, *args, **kwargs):
        logger.info("POST /send-return-items/ 요청 시작.")

        # 1) 요청 바디 파싱
        try:
            data = json.loads(request.body)
            ids = data.get('ids', [])
        except Exception as e:
            logger.exception("JSON 파싱 실패: %s", e)
            return JsonResponse({'success': False, 'message': 'Invalid JSON.'}, status=400)

        if not ids:
            logger.warning("전송할 아이디 미제공.")
            return JsonResponse({'success': False, 'message': 'No item ids provided.'}, status=400)

        return_items = list(ReturnItem.objects.filter(id__in=ids))
        if not return_items:
            logger.warning("해당 ID의 ReturnItem 없음: %s", ids)
            return JsonResponse({'success': False, 'message': 'No matching items.'}, status=404)

        # 2) SellerTool 인증값 생성 (공통)
        ST_API_KEY = config('ST_API_KEY', default=None)
        ST_SECRET_KEY = config('ST_SECRET_KEY', default=None)
        timestamp, signature = generate_sellertool_signature(ST_API_KEY, ST_SECRET_KEY)
        headers = {
            "x-sellertool-apiKey": ST_API_KEY,
            "x-sellertool-timestamp": timestamp,
            "x-sellertool-signiture": signature,
            "Content-Type": "application/json",
        }
        url = "https://sellertool-api-server-function.azurewebsites.net/api/return-exchanges/from-system"

        # 3) 병렬 전송 작업 정의
        def send_one(item):
            form_data = convert_return_item_to_formdata(item)
            body = {"formDatas": [form_data]}
            try:
                resp = requests.post(url, headers=headers, json=body, timeout=5)
                resp.raise_for_status()
                result = resp.json()
                failure_contents = result.get("content", {}).get("failureContents", [])
                if not result.get("error") and not failure_contents:
                    return {'id': item.id, 'success': True}
                reason = (failure_contents[0].get("reason")
                          if failure_contents else result.get("error", "Unknown error"))
                return {'id': item.id, 'success': False, 'reason': reason}
            except requests.exceptions.RequestException as e:
                logger.error("Item %s 전송 중 HTTP 오류: %s", item.id, e)
                return {'id': item.id, 'success': False, 'reason': f"HTTP error: {e}"}
            except ValueError as e:
                logger.error("Item %s 응답 파싱 실패: %s", item.id, e)
                return {'id': item.id, 'success': False, 'reason': f"Invalid JSON: {e}"}
            except Exception as e:
                logger.exception("Item %s 전송 중 예외: %s", item.id, e)
                return {'id': item.id, 'success': False, 'reason': str(e)}

        # 4) ThreadPoolExecutor로 병렬 전송
        max_workers = min(10, len(return_items))
        results = []
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(send_one, item): item for item in return_items}
            for fut in as_completed(futures):
                results.append(fut.result())

        # 5) 성공/실패 분리 & DB 대량 업데이트
        success_ids = [r['id'] for r in results if r['success']]
        failure_details = [r for r in results if not r['success']]

        if success_ids:
            # bulk_update를 위한 객체 생성
            success_items = [
                ReturnItem(id=item.id, processing_status='검수완료', product_issue='보류')
                for item in return_items if item.id in success_ids
            ]
            with transaction.atomic():
                ReturnItem.objects.bulk_update(success_items, ['processing_status', 'product_issue'])
            logger.info("성공 처리된 아이템 %d개 bulk_update 완료.", len(success_items))

        logger.info("최종 응답 반환: 성공 %d건, 실패 %d건", len(success_ids), len(failure_details))
        return JsonResponse({
            'total': len(return_items),
            'success_count': len(success_ids),
            'failure_count': len(failure_details),
            'failures': failure_details,
        })
