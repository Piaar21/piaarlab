# rankings/views.py

import json
import logging
from django.db.models import Avg, Subquery, OuterRef,Q
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db.models.functions import TruncMinute
from django.db import IntegrityError
from django.views.decorators.csrf import csrf_protect
from django.contrib import messages  # 메시지 추가
from django.db.models import Sum
from django.db.models.functions import TruncDate
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from django.forms import modelformset_factory
from django.urls import reverse  # reverse 함수 임포트 추가
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.csrf import csrf_exempt
from .forms import ProductForm, TrafficForm, UserProfileForm, AdForm, ExcelUploadForm, TaskForm
from .models import Product, Keyword, Ranking, Traffic, Task, UserProfile, Ad
from .resources import ProductResource
from tablib import Dataset
from celery import shared_task  # 추가
from .api_clients import get_naver_rank,naver_client_id, naver_client_secret  # Naver API 함수 임포트
import openpyxl  # 엑셀 파일 처리를 위한 openpyxl 임포트
from datetime import datetime, timedelta, date
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from django.utils.encoding import smart_str
from openpyxl import load_workbook  # 엑셀 파일 파싱을 위해
from sales_management.models import NaverDailySales
import re
from django.contrib import messages


logger = logging.getLogger(__name__)


@login_required
def ranking_list(request):
    keywords = Keyword.objects.all()
    context = {
        'keywords': keywords
    }
    return render(request, 'rankings/ranking_list.html', context)

@login_required
def product_add(request):
    if request.method == 'POST':
        # 엑셀 파일이 업로드된 경우
        if 'upload_excel' in request.POST:
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, "엑셀 파일을 선택해주세요.")
                return redirect('rankings:product_add')  # 네임스페이스 포함

            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active

                # 헤더 매핑 정의
                header_mapping = {
                    '카테고리': 'category',
                    '상품명': 'name',
                    '순위 조회 키워드': 'search_keyword',  # 추가된 부분
                    '단일상품링크': 'single_product_link',
                    '단일상품 MID값': 'single_product_mid',
                    '원부 링크': 'original_link',
                    '원부 MID': 'original_mid',
                    '스토어명': 'store_name',
                    '담당자': 'manager',
                }

                # 첫 번째 행을 헤더로 가져오고, 필드 이름 매핑
                headers = [cell.value for cell in ws[1]]
                mapped_headers = [header_mapping.get(header, None) for header in headers]

                # 데이터 읽기
                data_list = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_data = {}
                    for idx, cell_value in enumerate(row):
                        field_name = mapped_headers[idx]
                        if field_name:  # 매핑된 필드만 저장
                            row_data[field_name] = cell_value if cell_value is not None else ''  # 수정된 부분
                    data_list.append(row_data)

                # 세션에 데이터 저장
                request.session['excel_data'] = data_list
                messages.success(request, "엑셀 데이터를 성공적으로 불러왔습니다. 확인 후 등록해주세요.")
                return redirect('rankings:product_add')

            except Exception as e:
                logger.error(f"엑셀 파일 처리 중 오류 발생: {e}")
                messages.error(request, "엑셀 파일을 읽는 중 오류가 발생했습니다. 파일 형식을 확인하세요.")
                return redirect('rankings:product_add')
            
        # 2) 폼으로 등록하는 로직
        elif 'register_products' in request.POST:
            categories = request.POST.getlist('category[]')
            names = request.POST.getlist('name[]')
            search_keywords = request.POST.getlist('search_keyword[]')
            single_product_links = request.POST.getlist('single_product_link[]')
            single_product_mids = request.POST.getlist('single_product_mid[]')
            original_links = request.POST.getlist('original_link[]')
            original_mids = request.POST.getlist('original_mid[]')
            store_names = request.POST.getlist('store_name[]')
            managers = request.POST.getlist('manager[]')

            total_count = len(names)
            created_count = 0

            for i in range(total_count):
                category = categories[i]
                name = names[i]
                search_keyword = search_keywords[i] if i < len(search_keywords) else ''
                single_product_link = single_product_links[i]
                single_product_mid = single_product_mids[i]
                original_link = original_links[i]
                original_mid = original_mids[i]
                store_name = store_names[i]
                manager = managers[i] if i < len(managers) else ''

                # 필수 값 검증
                if not category or not name:
                    messages.warning(request, f"{i+1}번째 상품: 필수 항목(카테고리, 상품명) 누락으로 건너뜁니다.")
                    continue

                # (A) single_product_mid 중복 체크
                if single_product_mid:
                    if Product.objects.filter(single_product_mid=single_product_mid).exists():
                        messages.error(request, 
                            f"{i+1}번째 상품: single_product_mid값 [{single_product_mid}]가 이미 존재합니다. 건너뜁니다.")
                        continue

                # # (B) original_mid 중복 체크
                # if original_mid:
                #     if Product.objects.filter(original_mid=original_mid).exists():
                #         messages.error(request, 
                #             f"{i+1}번째 상품: original_mid값 [{original_mid}]가 이미 존재합니다. 건너뜁니다.")
                #         continue

                # (C) 중복이 없으면 DB에 생성
                Product.objects.create(
                    category=category,
                    name=name,
                    search_keyword=search_keyword,
                    single_product_link=single_product_link,
                    single_product_mid=single_product_mid,
                    original_link=original_link,
                    original_mid=original_mid,
                    store_name=store_name,
                    manager=manager,
                )
                created_count += 1

            # 세션에서 엑셀 데이터 삭제
            if 'excel_data' in request.session:
                del request.session['excel_data']

            messages.success(request, f"상품 등록 완료! (총 {created_count}건 생성)")
            return redirect('rankings:product_list')

    else:
        form = ProductForm()

    # 세션에 엑셀 데이터가 있으면 컨텍스트에 전달
    excel_data = request.session.get('excel_data', [])
    return render(request, 'rankings/product_add.html', {
        'form': form, 
        'excel_data': excel_data
    })

@login_required
def download_product_sample_excel(request):
    # 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample"

    # 헤더 설정
    headers = ['카테고리', '상품명', '순위 조회 키워드', '단일상품링크', '단일상품 MID값', '원부 링크', '원부 MID', '스토어명', '담당자']
    ws.append(headers)

    # 샘플 데이터 추가
    sample_data = [
        ['건강식품', '철분제', '철분제', 'http://example.com/product1', '123456', 'http://example.com/original1', '654321', '스토어1', '담당자1'],
        # 추가 샘플 데이터 필요 시 여기에 추가
    ]
    for row in sample_data:
        ws.append(row)

    # 열 너비 조정 및 정렬
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2
        for cell in column_cells:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 엑셀 파일로 응답
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=product_sample.xlsx'
    wb.save(response)
    return response


def product_list(request):
    products = Product.objects.all()
    context = {'products': products}
    return render(request, 'rankings/product_list.html', context)

from collections import defaultdict


@login_required
def dashborad_get_sales_data(request):
    task_id = request.GET.get('task_id')
    if not task_id:
        logger.error("task_id is missing in request")
        return JsonResponse({'error': 'task_id is required'}, status=400)
    
    try:
        task = Task.objects.get(id=task_id)
    except Task.DoesNotExist:
        logger.error("Task with id %s not found", task_id)
        return JsonResponse({'error': 'Task not found'}, status=404)
    
    if not task.single_product_mid:
        logger.error("Task id %s does not have single_product_mid", task_id)
    logger.debug("single_product_mid repr: %r (type=%s)",
             task.single_product_mid, type(task.single_product_mid))
    
    # 시작 날짜: task.available_start_date가 있다면 해당 날짜의 10일 전, 없으면 오늘 기준 10일 전
    if task.available_start_date:
        start_date = task.available_start_date - timedelta(days=10)
    else:
        start_date = timezone.now().date() - timedelta(days=10)
    # 종료 날짜: 어제
    end_date = timezone.now().date() - timedelta(days=0)
    
    logger.debug("Task id: %s, single_product_mid: %s", task_id, task.single_product_mid)
    logger.debug("Fetching sales records where product_id=%s between %s and %s",
                 task.single_product_mid, start_date, end_date)
    
    try:
        # 날짜 필드는 문자열 형태('YYYY-MM-DD')로 저장되어 있다고 가정하고 비교
        sales_records = NaverDailySales.objects.filter(
            product_id=task.single_product_mid,
            date__gte=start_date.strftime("%Y-%m-%d"),
            date__lte=end_date.strftime("%Y-%m-%d")
        ).order_by('date')
        record_count = sales_records.count()
        logger.debug("Found %s sales records", record_count)
    except Exception as e:
        logger.error("Error fetching sales records: %s", e)
        return JsonResponse({'error': 'Error fetching sales records'}, status=500)
    
    # 날짜별로 매출액과 판매수량을 합산하여 그룹화
    aggregated = defaultdict(lambda: {'sales_revenue': 0, 'sales_qty': 0})
    for record in sales_records:
        if not record.date:
            logger.warning("Sales record id %s has no date", record.id)
            continue
        try:
            # record.date는 'YYYY-MM-DD' 형태의 문자열라고 가정
            date_str = record.date.strip()
            aggregated[date_str]['sales_revenue'] += record.sales_revenue
            aggregated[date_str]['sales_qty'] += record.sales_qty
        except Exception as e:
            logger.error("Error processing sales record id %s: %s", record.id, e)
            continue
    
    # 지정한 기간의 모든 날짜를 생성하고, 해당 날짜의 집계 데이터가 있으면 사용하고 없으면 0으로 채움
    result = []
    current_date = start_date
    while current_date <= end_date:
        date_str = current_date.strftime("%Y-%m-%d")
        data = aggregated.get(date_str, {'sales_revenue': 0, 'sales_qty': 0})
        result.append({
            'date': date_str,
            'sales_revenue': data['sales_revenue'],
            'sales_qty': data['sales_qty'],
        })
        current_date += timedelta(days=1)
    
    return JsonResponse({'sales': result})

@login_required
def task_register(request):
    from datetime import datetime
    # 이제 사용자 프로필에서 API 정보를 받지 않고, api_clients에 정의된 naver_client_id, naver_client_secret 사용
    if request.method == 'POST':
        if 'register' in request.POST:
            product_ids = request.POST.getlist('product_ids')
            if not product_ids:
                messages.warning(request, '상품을 선택해주세요.')
                return redirect('rankings:product_list')
            products = Product.objects.filter(id__in=product_ids)
            traffics = Traffic.objects.all()
            initial_data = build_initial_data_from_post(request, product_ids)  # 폼 재구성을 위한 initial_data 생성

            for product_id in product_ids:
                try:
                    product = Product.objects.get(id=product_id)
                except Product.DoesNotExist:
                    continue  # 존재하지 않는 상품은 건너뜁니다.

                # 폼에서 데이터 가져오기
                category = request.POST.get(f'category_{product_id}', '네이버')
                keyword_name = request.POST.get(f'keyword_{product_id}')
                url_type = request.POST.get(f'url_type_{product_id}')
                url = request.POST.get(f'url_{product_id}')
                memo = request.POST.get(f'memo_{product_id}')
                product_name = request.POST.get(f'product_name_{product_id}')
                store_name = request.POST.get(f'store_name_{product_id}', '').strip()
                traffic_id = request.POST.get(f'traffic_{product_id}')
                traffic = Traffic.objects.get(id=traffic_id) if traffic_id else None

                # 필수 필드 검증
                missing_fields = []
                if not category:
                    missing_fields.append('카테고리')
                if not keyword_name:
                    missing_fields.append('순위조회키워드')
                if not url_type:
                    missing_fields.append('URL 타입')
                if not url:
                    missing_fields.append('URL')
                if not product_name:
                    missing_fields.append('상품명')

                # ticket_count를 정수로 변환
                ticket_count_str = request.POST.get(f'ticket_count_{product_id}')
                if ticket_count_str:
                    try:
                        ticket_count = int(ticket_count_str)
                    except ValueError:
                        missing_fields.append('이용권 수 (숫자여야 합니다)')
                        ticket_count = None
                else:
                    missing_fields.append('이용권 수')
                    ticket_count = None

                # 날짜 필드를 datetime.date로 변환
                available_start_date_str = request.POST.get(f'available_start_date_{product_id}')
                available_end_date_str = request.POST.get(f'available_end_date_{product_id}')
                
                try:
                    available_start_date = datetime.strptime(available_start_date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    missing_fields.append('이용가능 시작일자 (올바른 날짜 형식이어야 합니다)')
                    available_start_date = None

                try:
                    available_end_date = datetime.strptime(available_end_date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    missing_fields.append('이용가능 종료일자 (올바른 날짜 형식이어야 합니다)')
                    available_end_date = None

                if missing_fields:
                    error_message = f"상품 ID {product_id}: 다음 필드를 확인해주세요: {', '.join(missing_fields)}"
                    initial_data = build_initial_data_from_post(request, product_ids)
                    return render(request, 'rankings/task_register.html', {
                        'error': error_message,
                        'products': products,
                        'traffics': traffics,
                        'initial_data': initial_data,
                    })

                # Keyword 인스턴스 가져오기 또는 생성
                keyword_obj, created = Keyword.objects.get_or_create(name=keyword_name)

                # Naver API를 사용하여 시작 순위 가져오기
                # user 대신 api_clients에 정의된 API 정보를 활용합니다.
                try:
                    start_rank = get_naver_rank(keyword_name, url)
                    print(f"Start rank for keyword '{keyword_name}': {start_rank}")
                except Exception as e:
                    print(f"Error in get_naver_rank: {e}")
                    error_message = f"Naver API 호출 중 오류 발생: {str(e)}"
                    initial_data = build_initial_data_from_post(request, product_ids)
                    return render(request, 'rankings/task_register.html', {
                        'error': error_message,
                        'products': products,
                        'traffics': traffics,
                        'initial_data': initial_data,
                    })

                if start_rank == -1 or start_rank > 1000:
                    start_rank = None  # 검색 결과가 없거나 1000위 초과면 None으로 설정

                # 현재 순위와 어제 순위는 등록 시 동일하게 설정
                yesterday_rank = start_rank
                current_rank = start_rank
                difference_rank = 0  # 처음 등록 시 차이는 0

                # Task 생성
                try:
                    task = Task.objects.create(
                        product=product,
                        category=category,
                        keyword=keyword_obj,
                        url=url,
                        start_rank=start_rank,
                        yesterday_rank=yesterday_rank,
                        current_rank=current_rank,
                        difference_rank=difference_rank,
                        memo=memo,
                        product_name=product_name,
                        ticket_count=ticket_count,
                        available_start_date=available_start_date,
                        available_end_date=available_end_date,
                        traffic=traffic,
                        single_product_link=product.single_product_link,  # 추가
                        single_product_mid=product.single_product_mid,    # 추가
                        store_name=product.store_name,  # 추가
                    )
                    # Ranking 객체 생성
                    Ranking.objects.create(
                        task=task,
                        product=task.product,
                        keyword=keyword_obj,
                        rank=start_rank if start_rank is not None else 1000,
                        date_time=timezone.now(),
                    )
                except Exception as e:
                    error_message = f"작업 생성 중 오류 발생: {str(e)}"
                    initial_data = build_initial_data_from_post(request, product_ids)
                    return render(request, 'rankings/task_register.html', {
                        'error': error_message,
                        'products': products,
                        'traffics': traffics,
                        'initial_data': initial_data,
                    })

            messages.success(request, '작업이 성공적으로 등록되었습니다.')
            return redirect('rankings:dashboard')
        else:
            messages.error(request, '잘못된 요청입니다.')
            return redirect('rankings:product_list')
    else:
        # GET 요청 처리
        selected_product_ids = request.GET.getlist('selected_products')
        if not selected_product_ids:
            messages.warning(request, '상품을 선택해주세요.')
            return redirect('rankings:product_list')
        products = Product.objects.filter(id__in=selected_product_ids)
        initial_data = {}
        for product in products:
            initial_data[product.id] = {
                'keyword': product.search_keyword or '',
                'url_type': '',
                'url': '',
                'memo': '',
                'product_name': product.name,
                'store_name': '',
                'traffic_id': '',
                'ticket_count': '',
                'available_start_date': '',
                'available_end_date': '',
            }
        traffics = Traffic.objects.all()
        context = {
            'products': products,
            'traffics': traffics,
            'initial_data': initial_data,
        }
        return render(request, 'rankings/task_register.html', context)
    


@login_required
def task_upload_excel_data(request):
    logger.info("▶▶ task_upload_excel_data 호출")

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST 요청이 아닙니다.'})

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return JsonResponse({'success': False, 'error': '엑셀 파일이 전송되지 않았습니다.'})

    try:
        wb = load_workbook(filename=excel_file, data_only=True)
        ws = wb.active

        # 1) 엑셀 내 모든 셀의 하이퍼링크 제거 + URL로 교체
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    # 셀에 하이퍼링크가 존재하는 경우, 아래 중 원하는 로직으로 사용
                    # (1) 클릭 시 이동하는 실제 링크로 바꿔치기
                    cell.value = cell.hyperlink.target  
                    # (2) 링크 제거
                    cell.hyperlink = None

        # 2) 헤더 읽기 및 컬럼 인덱스 찾기
        headers = [str(cell.value).strip() for cell in ws[1]]
        logger.debug(f"  ▶ 읽어온 헤더: {headers}")

        # 필수 컬럼 체크
        for col_name in ['상품명', '단일', '원부']:
            if col_name not in headers:
                return JsonResponse({
                    'success': False,
                    'error': f'엑셀 헤더에 "{col_name}" 컬럼이 없습니다.'
                })
        name_col      = headers.index('상품명')
        single_col    = headers.index('단일')
        original_col  = headers.index('원부')
        keyword_col   = headers.index('순위조회키워드') if '순위조회키워드' in headers else None
        url_col       = headers.index('URL')               if 'URL' in headers               else None
        memo_col      = headers.index('메모')              if '메모' in headers              else None
        ticket_col    = headers.index('이용권 수')         if '이용권 수' in headers         else None
        start_col     = headers.index('이용가능 시작일자')  if '이용가능 시작일자' in headers  else None
        end_col       = headers.index('이용가능 종료일자') if '이용가능 종료일자' in headers else None

        tasks_to_create = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 2) 엑셀값 꺼내기
            name_excel     = row[name_col]
            single_excel   = row[single_col]
            original_excel = row[original_col]
            keyword_name   = row[keyword_col] if keyword_col is not None else ''
            ranking_url    = row[url_col]     if url_col     is not None else ''
            memo           = row[memo_col]    if memo_col    is not None else ''
            ticket_count   = row[ticket_col]  if ticket_col  is not None else 0
            start_date     = row[start_col]   if start_col   is not None else None
            end_date       = row[end_col]     if end_col     is not None else None

            logger.debug(f"  ▶ 행 {idx}: name={name_excel!r}, single={single_excel!r}, original={original_excel!r}")

            # 3) Product 조회 (상품명 + 링크 조합)
            product = None
            if name_excel and single_excel:
                product = Product.objects.filter(
                    name=name_excel.strip().lower(),
                    single_product_link=single_excel.strip().lower()
                ).first()
                if product is None:
                    logger.warning(f"  ✖ 행 {idx}: 상품({name_excel!r}) + 단일 링크({single_excel!r}) 매칭 실패")

            if not product and name_excel and original_excel:
                product = Product.objects.filter(
                    name=name_excel.strip().lower(),
                    original_link=original_excel.strip().lower()
                ).first()
                if product is None:
                    logger.warning(f"  ✖ 행 {idx}: 상품({name_excel!r}) + 원부 링크({original_excel!r}) 매칭 실패")
        



            if not product:
                logger.warning(f"  ✖ 행 {idx}: 상품({name_excel!r}) + 링크 매칭 실패 (single:{single_excel!r}, original:{original_excel!r})")
                continue

            # 4) 트래픽 조회 (엑셀의 '트래픽명' 컬럼이 있을 경우)
            traffic_obj = None
            if '트래픽명' in headers:
                traffic_name = row[headers.index('트래픽명')]
                if traffic_name:
                    traffic_obj = Traffic.objects.filter(name=traffic_name.strip()).first()
                    if not traffic_obj:
                        logger.warning(f"  ✖ 행 {idx}: 트래픽명({traffic_name!r}) 매칭 실패")

            # 5) Ranking 로직
            start_rank = get_naver_rank(keyword_name, ranking_url) if keyword_name and ranking_url else None
            if start_rank == -1 or (start_rank and start_rank > 1000):
                start_rank = None

            keyword_obj, _ = Keyword.objects.get_or_create(name=keyword_name)

            # 6) Task 생성 데이터 축적
            tasks_to_create.append({
                'product':              product,
                'category':             product.category,   # 기존 FK/필드 사용
                'keyword':              keyword_obj,
                'url':                  ranking_url,
                'start_rank':           start_rank,
                'yesterday_rank':       start_rank,
                'current_rank':         start_rank,
                'difference_rank':      0,
                'memo':                 memo,
                'product_name':         name_excel,
                'ticket_count':         ticket_count or 0,
                'available_start_date': start_date,
                'available_end_date':   end_date,
                'traffic':              traffic_obj,
            })

        logger.info(f"▶ 총 생성할 Task 개수: {len(tasks_to_create)}")

        # 7) 실제 DB 반영
        for data in tasks_to_create:
            task = Task.objects.create(**data)
            Ranking.objects.create(
                task=task,
                product=task.product,
                keyword=task.keyword,
                rank=task.start_rank or 1000,
                date_time=timezone.now()
            )
            logger.info(f"  ✔ Task(ID={task.id}) 및 Ranking 생성 완료")

        return JsonResponse({'success': True})

    except Exception as e:
        logger.error(f"▶ 예외 발생: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)})




@login_required
def download_bulk_traffic_sample_excel(request):
    # 오늘 날짜 기준 내일과 10일 뒤 날짜 계산
    today = timezone.localdate()
    start_date = today + timedelta(days=1)
    end_date = today + timedelta(days=10)

    # GET 파라미터로 traffic_id 확인 및 Traffic 객체 조회
    traffic_id = request.GET.get('traffic_id')
    logger.debug(f"Received traffic_id: {traffic_id}")
    traffic = None
    if traffic_id:
        try:
            traffic = Traffic.objects.get(id=traffic_id)
            logger.debug(f"Loaded Traffic: ID {traffic.id}, type: {traffic.type}")
        except Traffic.DoesNotExist:
            logger.error(f"Traffic with id {traffic_id} does not exist.")
            traffic = None
    else:
        logger.info("No traffic_id provided in request.")

    # 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active

    # 헤더 설정 (트래픽명 컬럼은 엑셀에 포함하지 않음)
    headers = [
        '상품ID', 
        '상품명', 
        '카테고리', 
        '순위조회키워드', 
        'URL', 
        '메모', 
        '이용권수', 
        '이용가능 시작일자', 
        '이용가능 종료일자',
        '스토어명',
    ]
    ws.append(headers)

    # DB에 등록된 모든 상품 조회
    products = Product.objects.all()

    for product in products:
        if traffic:
            logger.debug(f"Processing Product ID {product.id} with Traffic type: {traffic.type}")
            if traffic.type == '단일':
                url = product.single_product_link or ''
            elif traffic.type == '원부':
                url = product.original_link or ''
            else:
                url = ''
            logger.info(f"Product ID {product.id} - URL: {url} (Traffic type: {traffic.type})")
        else:
            # Traffic 정보가 없는 경우 기본 단일 링크 사용
            url = product.single_product_link or ''
            logger.info(f"Product ID {product.id} - Traffic 정보가 없습니다. 기본 URL 사용: {url}")

        row = [
            product.id,                          # 상품ID
            product.name or '',                  # 상품명
            "네이버",                            # 카테고리 (항상 네이버)
            product.search_keyword or '',        # 순위조회키워드
            url,                                 # URL (모달에서 선택한 트래픽 기준)
            '',                                  # 메모 (빈 값)
            1,                                   # 이용권수 (항상 1)
            start_date.strftime('%Y-%m-%d'),      # 이용가능 시작일자 (내일)
            end_date.strftime('%Y-%m-%d'),        # 이용가능 종료일자 (10일 뒤)
            product.store_name or '',                                    # 스토어명 (빈 값)
            
        ]
        ws.append(row)

    # 열 너비 자동 조정
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = max_length + 2

    # 엑셀 파일 응답 생성
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bulk_task_sample.xlsx'
    wb.save(response)
    return response

def build_initial_data_from_post(request, product_ids):
    initial_data = {}
    for product_id in product_ids:
        initial_data[int(product_id)] = {
            'keyword': request.POST.get(f'keyword_{product_id}', ''),
            'url_type': request.POST.get(f'url_type_{product_id}', ''),
            'url': request.POST.get(f'url_{product_id}', ''),
            'memo': request.POST.get(f'memo_{product_id}', ''),
            'product_name': request.POST.get(f'product_name_{product_id}', ''),
            'traffic_id': request.POST.get(f'traffic_{product_id}', ''),
            'ticket_count': request.POST.get(f'ticket_count_{product_id}', ''),
            'available_start_date': request.POST.get(f'available_start_date_{product_id}', ''),
            'available_end_date': request.POST.get(f'available_end_date_{product_id}', ''),
        }
    return initial_data




def build_initial_data_from_post(request, product_ids):
    initial_data = {}
    for product_id in product_ids:
        initial_data[int(product_id)] = {
            'keyword': request.POST.get(f'keyword_{product_id}', ''),
            'url_type': request.POST.get(f'url_type_{product_id}', ''),
            'url': request.POST.get(f'url_{product_id}', ''),
            'memo': request.POST.get(f'memo_{product_id}', ''),
            'product_name': request.POST.get(f'product_name_{product_id}', ''),
            'traffic_id': request.POST.get(f'traffic_{product_id}', ''),
            'ticket_count': request.POST.get(f'ticket_count_{product_id}', ''),
            'available_start_date': request.POST.get(f'available_start_date_{product_id}', ''),
            'available_end_date': request.POST.get(f'available_end_date_{product_id}', ''),
        }
    return initial_data

def get_products_with_latest_task(request, selected_product_ids):
    return Product.objects.filter(id__in=selected_product_ids).annotate(
        latest_task_id=Subquery(
            Task.objects.filter(product=OuterRef('pk')).order_by('-created_at').values('id')[:1]
        ),
        yesterday_rank_annotated=Subquery(
            Task.objects.filter(id=OuterRef('latest_task_id')).values('yesterday_rank')
        ),
        current_rank_annotated=Subquery(
            Task.objects.filter(id=OuterRef('latest_task_id')).values('current_rank')
        ),
        start_rank_annotated=Subquery(
            Task.objects.filter(id=OuterRef('latest_task_id')).values('start_rank')
        ),
        last_checked_date_annotated=Subquery(
            Task.objects.filter(id=OuterRef('latest_task_id')).values('last_checked_date')
        )
    )

@login_required
def dashboard(request):
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.utils import timezone
    from .models import Task
    import logging

    logger = logging.getLogger(__name__)
    today = timezone.now().date()
    tasks_list = Task.objects.filter(
        available_end_date__gte=today
    ).order_by('created_at').select_related('traffic', 'product')

    page = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 50)
    per_page_options = [50, 100, 150, 300]

    try:
        per_page = int(per_page)
        if per_page not in per_page_options:
            per_page = 50
    except ValueError:
        per_page = 50

    paginator = Paginator(tasks_list, per_page)
    try:
        tasks = paginator.page(page)
    except PageNotAnInteger:
        tasks = paginator.page(1)
    except EmptyPage:
        tasks = paginator.page(paginator.num_pages)

    context = {
        'tasks': tasks,
        'per_page_options': per_page_options,
        'selected_per_page': per_page,
    }
    return render(request, 'rankings/dashboard.html', context)

def compute_month_volume_and_growth(est_data):
    """
    est_data: 
      - {"results": [ {'period': 'YYYY-MM-DD', 'ratio': <float>, 'estimatedVolume': <int>}, ... ]}
      또는 
      - { "방석": [ {'period': 'YYYY-MM-DD', 'ratio': <float>, 'estimatedVolume': <int>}, ... ] }
      
    오늘을 기준으로:
      - recent period: today - 15일 ~ today - 1일
      - previous period: today - 29일 ~ today - 15일
    두 기간의 estimatedVolume 합계를 비교하여 성장률을 산출하고,
    최근 기간의 총 합계와 성장률을 반환합니다.
    """
    import datetime
    from django.utils import timezone

    today = timezone.now().date()
    recent_start = today - datetime.timedelta(days=15)
    recent_end = today - datetime.timedelta(days=1)
    previous_start = today - datetime.timedelta(days=29)
    previous_end = today - datetime.timedelta(days=15)
    
    # est_data의 구조에 따라 results 리스트를 가져옴
    if "results" in est_data:
        results = est_data["results"]
    else:
        # 키워드가 키로 되어 있는 경우, 첫 번째 키의 값을 사용
        key = list(est_data.keys())[0]
        results = est_data[key]
    
    recent_sum = 0
    previous_sum = 0
    for r in results:
        period_str = r.get("period")
        try:
            period_date = datetime.datetime.strptime(period_str, "%Y-%m-%d").date()
        except Exception:
            continue
        volume = r.get("estimatedVolume", 0)
        if recent_start <= period_date <= recent_end:
            recent_sum += volume
        if previous_start <= period_date < recent_start:
            previous_sum += volume

    if previous_sum == 0:
        growth_rate = "+0%"
    else:
        rate = ((recent_sum - previous_sum) / previous_sum) * 100.0
        sign = "+" if rate >= 0 else ""
        growth_rate = f"{sign}{rate:.1f}%"
    
    return (recent_sum, growth_rate)

def compute_recent_7day_sales_and_growth(product_mid, baseline_date):
    """
    product_mid를 사용해 NaverDailySales 테이블에서 매출 데이터를 조회하여,
    현재 창과 기준 창의 매출액을 비교해 성장률을 계산합니다.
    
    - 현재 창: 오늘 기준 7일 전부터 (오늘 - 1)일까지
    - 기준 창: Task.available_start_date 기준 7일 전부터 (available_start_date - 1)일까지
    
    만약 오늘이 baseline_date와 동일하면, 두 창이 같으므로 성장률은 0%로 반환합니다.
    
    반환:
      (current_window_total, growth_rate_str)
    """
    import datetime
    from django.utils import timezone

    if not product_mid:
        return (0, "0%")
    
    today = timezone.now().date()
    # 현재 창: 오늘 기준 7일 전 ~ (오늘 - 1일)
    current_start = today - datetime.timedelta(days=7)
    current_end = today - datetime.timedelta(days=1)
    
    # 기준 창: available_start_date 기준 7일 전 ~ (available_start_date - 1일)
    baseline_start = baseline_date - datetime.timedelta(days=8)
    baseline_end = baseline_date - datetime.timedelta(days=2)
    
    # 조회: 날짜는 "YYYY-MM-DD" 문자열 비교 (DB에 저장된 형식과 일치해야 함)
    current_records = NaverDailySales.objects.filter(
        product_id=product_mid,
        date__gte=current_start.strftime("%Y-%m-%d"),
        date__lte=current_end.strftime("%Y-%m-%d")
    ).order_by('date')
    baseline_records = NaverDailySales.objects.filter(
        product_id=product_mid,
        date__gte=baseline_start.strftime("%Y-%m-%d"),
        date__lte=baseline_end.strftime("%Y-%m-%d")
    ).order_by('date')
    
    # 각 창의 매출액 합산
    current_sum = sum(rec.sales_revenue for rec in current_records)
    baseline_sum = sum(rec.sales_revenue for rec in baseline_records)
    
    # 만약 오늘이 기준 날짜와 같다면, 두 창은 동일하므로 성장률은 0%
    if today == baseline_date:
        return (current_sum, "0%")
    
    # 기준 창 매출이 0이면 성장률은 0%로 처리
    if baseline_sum == 0:
        return (current_sum, "0%")
    
    rate = ((current_sum - baseline_sum) / baseline_sum) * 100.0
    sign = "+" if rate >= 0 else ""
    growth_rate = f"{sign}{rate:.1f}%"
    return (current_sum, growth_rate)

def compute_month_volume_and_growth_daily(daily_data):
    """
    daily_data: 일별 검색량 데이터 리스트.
      예: [
            {'period': '2023-01-01', 'searchVolume': 1050},
            {'period': '2023-01-02', 'searchVolume': 1050},
            ...
          ]
    
    오늘을 기준으로:
      - Recent period: today - 15일 ~ today - 1일
      - Previous period: today - 29일 ~ today - 15일
    각 기간의 searchVolume 합계를 계산하여 성장률 = ((recent_sum - previous_sum) / previous_sum) * 100을 산출합니다.
    (이전 기간 매출이 0이면 성장률은 "0%")
    
    반환:
        (recent_sum, growth_rate_str)
    """
    import datetime
    from django.utils import timezone

    today = timezone.now().date()
    recent_start = today - datetime.timedelta(days=15)
    recent_end = today - datetime.timedelta(days=1)
    previous_start = today - datetime.timedelta(days=29)
    previous_end = today - datetime.timedelta(days=15)
    
    recent_sum = 0
    previous_sum = 0
    for record in daily_data:
        try:
            period_date = datetime.datetime.strptime(record.get("period"), "%Y-%m-%d").date()
        except Exception:
            continue
        volume = record.get("searchVolume", 0)
        if recent_start <= period_date <= recent_end:
            recent_sum += volume
        elif previous_start <= period_date < recent_start:
            previous_sum += volume

    if previous_sum == 0:
        growth_rate = "0%"
    else:
        rate = ((recent_sum - previous_sum) / previous_sum) * 100.0
        sign = "+" if rate >= 0 else ""
        growth_rate = f"{sign}{rate:.1f}%"
    
    return recent_sum, growth_rate

def get_daily_search_volume_from_rel_keywords(keywords, start_date, end_date):
    """
    주어진 기간(start_date ~ end_date) 동안의 일별 검색량을, 
    get_rel_keywords 함수로부터 얻은 월간 총 검색수(totalSearchCount)를 균등 분배하여 반환합니다.
    
    계산식:
      daily_search_volume = totalSearchCount / (number of days in period)
    
    반환 예시:
      {
         '방석': [
              { 'period': '2023-01-01', 'searchVolume': 1050 },
              { 'period': '2023-01-02', 'searchVolume': 1050 },
              ...
         ]
      }
    """
    import datetime
    from datetime import timedelta
    import logging
    logger = logging.getLogger(__name__)
    
    # 광고 API를 통해 월간 총 검색수를 추출
    ad_df = get_rel_keywords(keywords)
    if ad_df is None or ad_df.empty:
        logger.error("광고 API 결과가 없거나 잘못되었습니다.")
        return None
    desired_keyword = keywords[0]
    filtered_df = ad_df[ad_df['relKeyword'] == desired_keyword]
    if filtered_df.empty:
        logger.error("광고 API 결과에서 해당 키워드 (%s)가 존재하지 않습니다.", desired_keyword)
        return None
    total_search = int(filtered_df.iloc[0]['totalSearchCount'])
    logger.debug("Desired keyword: %s, total monthly search count: %s", desired_keyword, total_search)
    
    # 기간 계산 (inclusive)
    start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
    end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
    num_days = (end_dt - start_dt).days + 1
    if num_days <= 0:
        logger.error("Invalid date range: start_date=%s, end_date=%s", start_date, end_date)
        return None
    
    daily_volume = int(round(total_search / num_days))
    
    daily_results = []
    current_date = start_dt
    while current_date <= end_dt:
        period_str = current_date.strftime("%Y-%m-%d")
        daily_results.append({
            "period": period_str,
            "searchVolume": daily_volume
        })
        current_date += timedelta(days=1)
    
    return { desired_keyword: daily_results }

@login_required
def get_kpi_data(request):
    """
    GET 파라미터로 전달된 task_id를 기반으로,
    해당 Task의 KPI(1달간 검색량, 최근 7일 매출액, status, 현재 순위 및 차이)를 계산하여 JSON으로 반환.
    이 뷰는 모달창이 열릴 때 AJAX 호출로 KPI 데이터를 가져오는 데 사용됩니다.
    """
    from django.http import JsonResponse
    from .models import Task
    from .api_clients import get_rel_keywords  # get_rel_keywords 함수 사용
    from django.utils import timezone
    import datetime
    import logging

    logger = logging.getLogger(__name__)
    
    task_id = request.GET.get('task_id')
    if not task_id:
        return JsonResponse({'error': 'Task ID is required.'}, status=400)
    
    try:
        task = Task.objects.get(id=task_id)
    except Task.DoesNotExist:
        return JsonResponse({'error': 'Task not found.'}, status=404)

    kpi = {}

    # KPI 3) 현재 상태
    kpi['status'] = task.status

    # KPI 4) 현재 순위 및 시작 순위 차이
    current_r = task.current_rank if task.current_rank else 1000
    start_r = task.start_rank if task.start_rank else 1000
    diff = current_r - start_r
    sign = "+" if diff < 0 else ""
    kpi['currentRank'] = current_r
    kpi['rankDiff'] = f"{sign}{diff}"

    # KPI 1) 1달간 검색량 및 키워드 규모 분류
    if task.keyword:
        today_date = timezone.now().date()
        # 최근 30일 데이터: 오늘 기준 지난 29일부터 오늘까지
        start_date = (today_date - datetime.timedelta(days=29)).strftime("%Y-%m-%d")
        end_date = today_date.strftime("%Y-%m-%d")
        daily_data_dict = get_daily_search_volume_from_rel_keywords([task.keyword.name], start_date, end_date)
        if daily_data_dict and task.keyword.name in daily_data_dict:
            daily_data = daily_data_dict[task.keyword.name]
            total_volume = sum(item.get("searchVolume", 0) for item in daily_data)
        else:
            total_volume = 0
        kpi['monthVolume'] = total_volume
        # 월간 검색량을 기준으로 키워드 규모 분류
        if total_volume < 2000:
            kpi['monthVolumeGrowth'] = "소형 키워드"
        elif total_volume < 30000:
            kpi['monthVolumeGrowth'] = "중형 키워드"
        else:
            kpi['monthVolumeGrowth'] = "대형 키워드"
    else:
        kpi['monthVolume'] = 0
        kpi['monthVolumeGrowth'] = "0%"

    # KPI 2) 최근 7일 매출액 및 성장률 (기존 함수 사용)
    if task.single_product_mid and task.available_start_date:
        recent_sales, sales_growth = compute_recent_7day_sales_and_growth(task.single_product_mid, task.available_start_date)
        kpi['recent7daySales'] = recent_sales
        kpi['recent7daySalesGrowth'] = sales_growth
    else:
        kpi['recent7daySales'] = 0
        kpi['recent7daySalesGrowth'] = "0%"

    return JsonResponse({'success': True, 'data': kpi})


@login_required
def completed_tasks_list(request):
    completed_tasks_list = Task.objects.filter(is_completed=True).order_by('-available_end_date')

    # 페이지네이션 설정
    page = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 10)
    per_page_options = [10, 30, 50, 100]

    # per_page 값 검증
    try:
        per_page = int(per_page)
        if per_page not in per_page_options:
            per_page = 10
    except ValueError:
        per_page = 10

    paginator = Paginator(completed_tasks_list, per_page)

    try:
        tasks = paginator.page(page)
    except PageNotAnInteger:
        tasks = paginator.page(1)
    except EmptyPage:
        tasks = paginator.page(paginator.num_pages)

    context = {
        'tasks': tasks,
        'per_page_options': per_page_options,
        'selected_per_page': per_page,
    }

    return render(request, 'rankings/completed_tasks_list.html', context)

def process_task_registration(request, product_ids, products, traffics, initial_data):
    for product_id in product_ids:
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            continue  # 다음 상품으로 이동

        # 나머지 코드 진행
        category = request.POST.get(f'category_{product_id}', '네이버')
        keyword_name = request.POST.get(f'keyword_{product_id}')
        url_type = request.POST.get(f'url_type_{product_id}')
        url = request.POST.get(f'url_{product_id}')
        memo = request.POST.get(f'memo_{product_id}')
        product_name = request.POST.get(f'product_name_{product_id}')
        traffic_id = request.POST.get(f'traffic_{product_id}')
        traffic = Traffic.objects.get(id=traffic_id) if traffic_id else None

        # 필수 필드 검증
        missing_fields = []
        if not category:
            missing_fields.append('category')
        if not keyword_name:
            missing_fields.append('keyword')
        if not url_type:
            missing_fields.append('url_type')
        if not url:
            missing_fields.append('url')
        if not product_name:
            missing_fields.append('product_name')

        # ticket_count를 정수로 변환
        ticket_count_str = request.POST.get(f'ticket_count_{product_id}')
        if ticket_count_str:
            try:
                ticket_count = int(ticket_count_str)
            except ValueError:
                missing_fields.append('ticket_count (숫자여야 합니다)')
                ticket_count = None
        else:
            missing_fields.append('ticket_count')
            ticket_count = None

        # 날짜 필드를 datetime.date로 변환
        available_start_date_str = request.POST.get(f'available_start_date_{product_id}')
        available_end_date_str = request.POST.get(f'available_end_date_{product_id}')

        try:
            available_start_date = datetime.strptime(available_start_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            missing_fields.append('available_start_date (올바른 날짜 형식이어야 합니다)')
            available_start_date = None

        try:
            available_end_date = datetime.strptime(available_end_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            missing_fields.append('available_end_date (올바른 날짜 형식이어야 합니다)')
            available_end_date = None

        if missing_fields:
            error_message = f"Product ID {product_id}: Missing or invalid fields: {', '.join(missing_fields)}"
            return render(request, 'rankings/task_register.html', {'error': error_message, 'products': products, 'traffics': traffics, 'initial_data': initial_data})

        # Keyword 인스턴스 가져오기 또는 생성
        keyword_obj, created = Keyword.objects.get_or_create(name=keyword_name)

        # Naver API를 사용하여 시작 순위 가져오기
        try:
            start_rank = get_naver_rank(keyword_name, url)
            print(f"Start rank for keyword '{keyword_name}': {start_rank}")
        except Exception as e:
            print(f"Error in get_naver_rank: {e}")
            error_message = f"Naver API 호출 중 오류 발생: {str(e)}"
            return render(request, 'rankings/task_register.html', {'error': error_message, 'products': products, 'traffics': traffics, 'initial_data': initial_data})

        if start_rank == -1 or start_rank > 1000:
            start_rank = None  # 검색 결과에 없거나 1000위 초과인 경우 None으로 설정

        # 현재 순위와 어제 순위는 작업 등록 시점에서는 동일하게 설정
        yesterday_rank = start_rank
        current_rank = start_rank
        difference_rank = 0  # 처음 등록 시 차이는 0

        # Task 생성
        try:
            task = Task.objects.create(
                product=product,
                category=category,
                keyword=keyword_obj,
                url=url,
                start_rank=start_rank,
                yesterday_rank=yesterday_rank,
                current_rank=current_rank,
                difference_rank=difference_rank,
                memo=memo,
                product_name=product_name,
                ticket_count=ticket_count,
                available_start_date=available_start_date,
                available_end_date=available_end_date,
                traffic=traffic,
            )
        except Exception as e:
            error_message = f"작업 생성 중 오류 발생: {str(e)}"
            return render(request, 'rankings/task_register.html', {'error': error_message, 'products': products, 'traffics': traffics, 'initial_data': initial_data})

    messages.success(request, '작업이 성공적으로 등록되었습니다.')
    return redirect('rankings:dashboard')  # 작업 등록 후 대시보드로 이동

@login_required
def task_reregister(request):

    if request.method == 'POST':
        if 'register' in request.POST:
            # 작업 등록 로직
            product_ids = request.POST.getlist('product_ids')
            if not product_ids:
                return redirect('rankings:product_list')
            products = Product.objects.filter(id__in=product_ids)
            traffics = Traffic.objects.all()
            initial_data = build_initial_data_from_post(request, product_ids)
            # process_task_registration 함수 호출
            return process_task_registration(request, product_ids, products, traffics, initial_data)
        else:
            # 선택된 작업으로 폼 초기화
            selected_task_ids = request.POST.getlist('selected_tasks')
            if not selected_task_ids:
                messages.warning(request, '작업을 선택해주세요.')
                return redirect('rankings:completed_tasks_list')
            tasks = Task.objects.filter(id__in=selected_task_ids)
            products = [task.product for task in tasks]
            traffics = Traffic.objects.all()
            context = {
                'products': products,
                'traffics': traffics,
            }
            return render(request, 'rankings/task_register.html', context)
    else:
        # GET 요청 처리
        selected_task_ids = request.GET.get('task_ids')
        if not selected_task_ids:
            messages.warning(request, '작업을 선택해주세요.')
            return redirect('rankings:completed_tasks_list')
        task_ids = selected_task_ids.split(',')
        tasks = Task.objects.filter(id__in=task_ids)
        products = [task.product for task in tasks]
        traffics = Traffic.objects.all()
        # initial_data 구성
        initial_data = {}
        for task in tasks:
            product_id = task.product.id  # 정수형 유지
            # URL 타입 결정
            if task.url == task.product.single_product_link:
                url_type = 'single'
            elif task.url == task.product.original_link:
                url_type = 'original'
            else:
                url_type = ''
            initial_data[product_id] = {  # 키를 정수형으로 사용
                'keyword': task.keyword.name,
                'url_type': url_type,
                'url': task.url,
                'memo': task.memo,
                'traffic_id': str(task.traffic.id) if task.traffic else '',
                'ticket_count': task.ticket_count,
                'available_start_date': task.available_start_date.strftime('%Y-%m-%d') if task.available_start_date else '',
                'available_end_date': task.available_end_date.strftime('%Y-%m-%d') if task.available_end_date else '',
            }
        context = {
            'products': products,
            'traffics': traffics,
            'initial_data': initial_data,
        }
        return render(request, 'rankings/task_register.html', context)

@login_required
def product_bulk_edit(request):
    ProductFormSet = modelformset_factory(
        Product,
        form=ProductForm,
        fields=[
            'id',
            'category',
            'name',
            'search_keyword',
            'single_product_link',
            'single_product_mid',
            'original_link',
            'original_mid',
            'store_name',
            'manager',
        ],
        extra=0
    )

    if request.method == 'POST':
        # 폼셋이 제출된 경우
        formset = ProductFormSet(request.POST)
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                instance.user = request.user
                instance.save()
            messages.success(request, "선택한 상품이 성공적으로 수정되었습니다.")
            return redirect('rankings:product_list')
        else:
            messages.error(request, "폼에 오류가 있습니다. 확인해주세요.")
    else:
        # GET 요청 처리: 선택된 상품 IDs를 가져옵니다.
        selected_ids = request.GET.getlist('selected_products')
        if not selected_ids:
            messages.error(request, "수정할 상품을 선택해주세요.")
            return redirect('rankings:product_list')
        products = Product.objects.filter(id__in=selected_ids)
        formset = ProductFormSet(queryset=products)

    return render(request, 'rankings/product_bulk_edit.html', {'formset': formset})



@login_required
def average_ranking(request):
    task_id = request.GET.get('task_id')
    if not task_id:
        return redirect('rankings:dashboard')

    task = get_object_or_404(Task, id=task_id)

    rankings = Ranking.objects.filter(
        product=task.product,
        keyword=task.keyword,
    ).annotate(
        time=TruncMinute('date_time')
    ).order_by('time')
    print(f"Rankings count: {rankings.count()}")

    dates = [ranking.time.strftime('%Y-%m-%d %H:%M') for ranking in rankings]
    ranks = [ranking.rank for ranking in rankings]

    context = {
        'task': task,
        'dates': json.dumps(dates),
        'ranks': json.dumps(ranks),
    }

    return render(request, 'rankings/average_ranking.html', context)

@login_required
def issues(request):
    # GET 파라미터에서 현재 페이지 번호와 페이지 당 항목 수를 가져옵니다.
    page = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 10)

    # 페이지 당 항목 수 옵션 설정
    per_page_options = [10, 30, 50, 100]

    # 유효한 페이지 당 항목 수인지 확인하고, 아니면 기본값 10으로 설정
    try:
        per_page = int(per_page)
        if per_page not in per_page_options:
            per_page = 10
    except ValueError:
        per_page = 10

    # current_rank가 None이거나 1000위 초과인 Task를 필터링합니다.
    problematic_tasks = Task.objects.filter(
        Q(current_rank__isnull=True) | Q(current_rank__gt=1000)
    )
    # Paginator를 사용하여 쿼리셋을 페이지네이션합니다.
    paginator = Paginator(problematic_tasks, per_page)

    try:
        tasks = paginator.page(page)
    except PageNotAnInteger:
        # 페이지 번호가 정수가 아니면 첫 번째 페이지를 보여줍니다.
        tasks = paginator.page(1)
    except EmptyPage:
        # 페이지 번호가 범위를 벗어나면 마지막 페이지를 보여줍니다.
        tasks = paginator.page(paginator.num_pages)

    context = {
        'tasks': tasks,
        'per_page_options': per_page_options,
        'selected_per_page': per_page,
    }
    return render(request, 'rankings/issues.html', context)

@login_required
def traffic_bulk_edit(request):
    TrafficFormSet = modelformset_factory(Traffic, form=TrafficForm, extra=0)
    if request.method == 'POST':
        if 'edit_traffics' in request.POST:
            # 선택된 트래픽을 가져와 폼셋 생성 (초기 로딩 시)
            selected_ids = request.POST.getlist('selected_traffics')
            if not selected_ids:
                messages.error(request, "수정할 트래픽을 선택해주세요.")
                return redirect('rankings:traffic_list')
            traffics = Traffic.objects.filter(id__in=selected_ids)
            formset = TrafficFormSet(queryset=traffics)
        else:
            # 폼셋 제출 시 데이터 처리
            formset = TrafficFormSet(request.POST)
            if formset.is_valid():
                instances = formset.save(commit=False)
                for instance in instances:
                    instance.user = request.user
                    instance.save()
                messages.success(request, "선택한 트래픽이 성공적으로 수정되었습니다.")
                return redirect('rankings:traffic_list')
            else:
                messages.error(request, "폼에 오류가 있습니다. 확인해주세요.")
    else:
        messages.error(request, "잘못된 접근입니다.")
        return redirect('rankings:traffic_list')
    return render(request, 'rankings/traffic_bulk_edit.html', {'formset': formset})



@login_required
def task_update(request):
    if request.method == 'POST':
        next_url = request.POST.get('next', reverse('rankings:dashboard'))
        logger.debug(f"Received POST request with next_url: {next_url}")

        task_ids = [key.split('_')[1] for key in request.POST.keys() if key.startswith('category_')]
        logger.debug(f"Task IDs to update: {task_ids}")

        for task_id in task_ids:
            try:
                task = get_object_or_404(Task, id=task_id)
                new_product_id = request.POST.get(f'product_id_{task_id}')
                new_keyword_name = request.POST.get(f'keyword_{task_id}')
                new_url = request.POST.get(f'url_{task_id}')
                new_category = request.POST.get(f'category_{task_id}')
                new_memo = request.POST.get(f'memo_{task_id}')

                # 새로운 날짜 값 가져오기
                new_available_start_date = request.POST.get(f'available_start_date_{task_id}')
                new_available_end_date = request.POST.get(f'available_end_date_{task_id}')

                # 날짜 형식 변환
                date_format = "%Y-%m-%d"
                if new_available_start_date:
                    new_available_start_date = datetime.strptime(new_available_start_date, date_format).date()
                if new_available_end_date:
                    new_available_end_date = datetime.strptime(new_available_end_date, date_format).date()

                # 필수 필드 검증
                if not new_keyword_name:
                    messages.error(request, f"작업 ID {task_id}의 키워드가 입력되지 않았습니다.")
                    continue
                if not new_url:
                    messages.error(request, f"작업 ID {task_id}의 URL이 입력되지 않았습니다.")
                    continue

                # 상품이 변경되었는지 확인
                if new_product_id and int(new_product_id) != task.product.id:
                    # 기존 작업 종료 처리
                    task.available_end_date = timezone.localdate()
                    task.save()

                    # 새로운 작업 생성
                    new_product = get_object_or_404(Product, id=new_product_id)
                    keyword_obj, _ = Keyword.objects.get_or_create(name=new_keyword_name)

                    # **available_start_date를 내일 날짜로 설정**
                    new_task = Task.objects.create(
                        product=new_product,
                        category=new_category,
                        keyword=keyword_obj,
                        url=new_url,
                        memo=new_memo,
                        product_name=new_product.name,
                        single_product_link=new_product.single_product_link,  # 추가
                        single_product_mid=new_product.single_product_mid,    # 추가
                        ticket_count=task.ticket_count,
                        available_start_date=timezone.localdate() + timedelta(days=1),
                        available_end_date=new_available_end_date or task.available_end_date,
                        traffic=task.traffic,
                    )
                    # 순위 조회 및 저장 (필요 시)
                else:
                    # 기존 작업 수정
                    old_keyword_name = task.keyword.name
                    old_url = task.url
                    task.category = new_category
                    keyword_obj, _ = Keyword.objects.get_or_create(name=new_keyword_name)
                    task.keyword = keyword_obj
                    task.url = new_url
                    task.memo = new_memo

                    if task.product:
                        task.product_name = task.product.name
                        task.single_product_link = task.product.single_product_link
                        task.single_product_mid = task.product.single_product_mid

                    # **available_start_date를 폼에서 제출된 값으로 업데이트**
                    if new_available_start_date:
                        task.available_start_date = new_available_start_date

                    # **available_end_date도 동일하게 처리**
                    if new_available_end_date:
                        task.available_end_date = new_available_end_date

                    task.save()
                    # 키워드 또는 URL이 변경되었는지 확인하여 순위 조회 수행 (필요 시)
            except Exception as e:
                logger.error(f"Error updating task {task_id}: {e}")
                messages.error(request, f"작업 ID {task_id} 업데이트 중 오류가 발생했습니다.")

        logger.debug(f"Redirecting to: {next_url}")
        return redirect(next_url)
    else:
        try:
            task_ids = request.GET.get('task_ids')
            next_url = request.GET.get('next', reverse('rankings:dashboard'))
            products = Product.objects.all()  # 모든 상품 가져오기

            if task_ids:
                task_id_list = task_ids.split(',')
                tasks = Task.objects.filter(id__in=task_id_list)
                if not tasks.exists():
                    messages.error(request, '선택한 작업을 찾을 수 없습니다.')
                    return redirect('rankings:task_list')
            else:
                messages.error(request, '수정할 작업을 선택해주세요.')
                return redirect('rankings:task_list')

            context = {
                'tasks': tasks,
                'products': products,  # 컨텍스트에 products 추가
                'next_url': next_url,  # 컨텍스트에 next_url 추가
            }
            return render(request, 'rankings/task_edit.html', context)
        except Exception as e:
            logger.error(f"Error in task_update GET: {e}")
            messages.error(request, '요청을 처리하는 중 오류가 발생했습니다.')
            return redirect('rankings:dashboard')


@csrf_protect
def product_delete(request):
    if request.method == 'POST':
        selected_product_ids = request.POST.getlist('selected_products')
        if selected_product_ids:
            Product.objects.filter(id__in=selected_product_ids).delete()
        return redirect('rankings:product_list')


@login_required
def traffic_register(request):
    if request.method == 'POST':
        if 'upload_excel' in request.POST:
            # 엑셀 업로드 처리
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, "엑셀 파일을 선택해주세요.")
                return redirect('rankings:traffic_register')
            # 엑셀 파일 읽기
            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                traffics_data = []
                
                for row in sheet.iter_rows(min_row=2, values_only=True):  # 헤더 제외
                    # 엑셀 파일의 열 순서에 따라 데이터 매핑
                    name = row[0]
                    price = row[1]
                    method = row[2]
                    days = row[3]
                    inflow_count = row[4]
                    link = row[5].strip() if row[5] else ''  # 링크가 없으면 빈 문자열 처리
                    vendor = row[6] if len(row) > 6 else ''  # 업체명 (추가된 부분)
                    type_value = row[7] if len(row) > 7 else ''  # 노출방식 (추가된 부분)
                    # 필수 값 검증
                    if not (name and price and method and days and inflow_count):
                        continue  # 필수 값이 없으면 건너뜁니다

                    # 데이터 타입 변환 및 유효성 검사
                    try:
                        price = int(price)
                        days = int(days)
                        inflow_count = int(inflow_count)
                    except ValueError:
                        continue  # 값이 유효하지 않으면 건너뜁니다

                        # 링크에 스킴 추가
                    if link and not link.startswith(('http://', 'https://')):
                        link = 'http://' + link

                    # 트래픽 데이터를 리스트에 추가
                    traffics_data.append({
                        'name': name,
                        'price': price,
                        'method': method,
                        'days': days,
                        'inflow_count': inflow_count,
                        'link': link,
                        'vendor': vendor,  
                        'type': type_value,  # 추가된 부분

                    })

                # traffics_data를 컨텍스트에 전달하여 폼에 표시
                return render(request, 'rankings/traffic_register.html', {'traffics_data': traffics_data})

            except Exception as e:
                print(f"Error processing the file: {e}")
                messages.error(request, f"엑셀 파일 처리 중 오류 발생: {e}")
                return redirect('rankings:traffic_register')

        elif 'register_traffic' in request.POST:
            # 폼 데이터 처리
            names = request.POST.getlist('name')
            prices = request.POST.getlist('price')
            methods = request.POST.getlist('method')
            days_list = request.POST.getlist('days')
            inflow_counts = request.POST.getlist('inflow_count')
            links = request.POST.getlist('link')
            vendors = request.POST.getlist('vendor')  
            types = request.POST.getlist('type')  # 추가된 부분: 노출방식
            for i in range(len(names)):
                name = names[i]
                price = prices[i]
                method = methods[i]
                days = days_list[i]
                inflow_count = inflow_counts[i]
                link = links[i].strip() if links[i] else ''  # 링크가 없으면 빈 문자열로 저장
                vendor = vendors[i].strip() if vendors[i] else ''  # 링크가 없으면 빈 문자열로 저장
                type_value = types[i].strip() if types[i] else ''  # 추가된 부분: 노출방식

                # 필수 값 검증
                if not (name and price and method and days and inflow_count):
                    continue  # 필수 값이 없으면 건너뜁니다

                # 데이터 타입 변환
                try:
                    price = int(price)  # 정수형으로 변환하여 소숫점 제거
                    days = int(days)
                    inflow_count = int(inflow_count)
                except ValueError:
                    continue  # 값이 유효하지 않으면 건너뜁니다

                # 링크에 스킴 추가
                if link and not link.startswith(('http://', 'https://')):
                    link = 'http://' + link

                # 트래픽 생성
                Traffic.objects.create(
                    name=name,
                    price=price,
                    method=method,
                    days=days,
                    inflow_count=inflow_count,
                    link=link or '',
                    vendor=vendor,  # 추가된 부분
                    type=type_value,  # 추가된 부분

                )
            messages.success(request, "트래픽을 성공적으로 등록했습니다.")
            return redirect('rankings:traffic_list')

    else:
        return render(request, 'rankings/traffic_register.html')



@login_required
def traffic_delete(request):
    if request.method == 'POST':
        selected_traffic_ids = request.POST.getlist('selected_traffics')
        if selected_traffic_ids:
            Traffic.objects.filter(id__in=selected_traffic_ids).delete()
            messages.success(request, "선택한 트래픽을 삭제했습니다.")
        else:
            messages.warning(request, "삭제할 트래픽을 선택해주세요.")
    return redirect('rankings:traffic_list')

@login_required
def download_traffic_sample_excel(request):
    # 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active

    # 헤더 설정
    headers = ['트래픽명', '금액', '방식', '일자', '유입수', '링크', '업체명']
    ws.append(headers)

    # 데이터 추가
    data = ['피날레', 30000, '리워드', 10, 10000, 'www.example.com','Sample Vendor']
    ws.append(data)

    # 열 너비 조정 및 정렬
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = length + 2

    # 엑셀 파일로 응답
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=traffic_sample.xlsx'
    wb.save(response)
    return response

@login_required
def task_action(request):
    if request.method == 'POST':
        selected_task_ids = request.POST.getlist('selected_tasks')
        action = request.POST.get('action')
        next_url = request.POST.get('next', 'rankings:dashboard')  # next 파라미터를 가져옴
        if selected_task_ids:
            tasks = Task.objects.filter(id__in=selected_task_ids)
            if action == 'edit':
                task_ids = ','.join(selected_task_ids)
                return redirect(f"{reverse('rankings:task_update')}?task_ids={task_ids}")
            elif action == 'complete':
                today = timezone.now().date()
                tasks.update(available_end_date=today)
                messages.success(request, "선택한 작업을 완료했습니다.")
            elif action == 'extend':
                for task in tasks:
                    # 새로운 작업 생성
                    new_task = Task.objects.create(
                        product=task.product,
                        category=task.category,
                        keyword=task.keyword,
                        url=task.url,
                        start_rank=task.start_rank,
                        yesterday_rank=task.yesterday_rank,
                        current_rank=task.current_rank,
                        difference_rank=task.difference_rank,
                        status=task.status,
                        memo=task.memo,
                        ticket_count=task.ticket_count,
                        product_name=task.product_name,
                        original_link=task.original_link,
                        original_mid=task.original_mid,
                        traffic=task.traffic,
                        is_completed=task.is_completed,
                        ending_rank=task.ending_rank,
                        available_start_date=task.available_end_date + timedelta(days=1),
                        available_end_date=task.available_end_date + timedelta(days=10),
                        is_extended=True,
                        original_task=task,
                    )
                messages.success(request, "선택한 작업을 10일 연장했습니다.")
            elif action == 'undo_extend':
                # 연장된 작업을 찾아서 삭제
                extended_tasks = Task.objects.filter(original_task__in=tasks, is_extended=True)
                extended_tasks.delete()
                messages.success(request, "선택한 작업의 연장을 취소했습니다.")
            elif action == 'delete':
                tasks.delete()
                messages.success(request, "선택한 작업을 삭제했습니다.")
        else:
            messages.warning(request, "작업을 선택해주세요.")
        return redirect(next_url)

def task_extend(request):
    if request.method == 'POST':
        task_ids = request.POST.get('task_ids').split(',')
        new_end_date = request.POST.get('new_end_date')
        if new_end_date:
            Task.objects.filter(id__in=task_ids).update(available_end_date=new_end_date)
            messages.success(request, "선택한 작업의 종료일자를 연장했습니다.")
        else:
            messages.warning(request, "새로운 종료일자를 입력해주세요.")
        return redirect('rankings:dashboard')
    else:
        task_ids = request.GET.get('task_ids').split(',')
        tasks = Task.objects.filter(id__in=task_ids)
        return render(request, 'rankings/task_extend.html', {'tasks': tasks, 'task_ids': ','.join(task_ids)})

def product_select(request):
    task_id = request.GET.get('task_id')
    products = Product.objects.all()
    return render(request, 'rankings/product_select.html', {'products': products, 'task_id': task_id})

@login_required
def task_edit(request):
    if request.method == 'POST':
        selected_task_ids = request.POST.getlist('selected_tasks')
        if not selected_task_ids:
            messages.error(request, '수정할 작업을 선택해주세요.')
            return redirect('rankings:task_list')
        selected_task_ids = task_ids.split(',')
        tasks = Task.objects.filter(id__in=selected_task_ids)
        traffics = Traffic.objects.all()
        context = {'tasks': tasks, 'traffics': traffics}
        return render(request, 'rankings/task_edit.html', context)
    else:
        return redirect('rankings:task_list')  # GET 요청 시 적절한 페이지로 리디렉션


def update_rankings():
    today = timezone.now().date()
    tasks = Task.objects.filter(is_completed=False, available_start_date__lte=today)

    for task in tasks:
        try:
            # 순위 업데이트 로직
            current_rank = get_naver_rank(task.keyword.name, task.url)
            # 추가: current_rank가 -1이거나 1000보다 크면 None으로 설정
            if current_rank == -1 or current_rank > 1000:
                current_rank = None

            task.yesterday_rank = task.current_rank
            task.current_rank = current_rank
            task.difference_rank = (
                (task.yesterday_rank - task.current_rank)
                if task.yesterday_rank is not None and task.current_rank is not None
                else 0
            )
        except Exception as e:
            print(f"Error updating task {task.id}: {e}")
            # 예외 발생 시에도 계속 진행하여 날짜 조건을 확인하도록 함
            pass

        # 오늘 날짜가 available_end_date보다 지난 경우 is_completed를 True로 업데이트
        if today > task.available_end_date:
            if not task.is_completed:  # 이미 완료된 작업은 재처리하지 않음
                task.is_completed = True
                task.ending_rank = task.current_rank

        task.save()

from .tasks import update_task_status

@login_required
def update_all_rankings(request):
    if request.method == 'POST':
        

        today = timezone.now().date()
        # “시작 전날부터” 포함하려면 today + 1일
        include_from = today + timedelta(days=1)

        tasks = Task.objects.filter(
            is_completed=False,
            available_start_date__lte=include_from,
            available_end_date__gte=today,
        )
        today = timezone.now().date()
        for task in tasks:
            try:
                current_rank = get_naver_rank(task.keyword.name, task.url)
                if current_rank == -1 or current_rank > 1000:
                    current_rank = None
            except Exception as e:
                logger.error(f"작업 ID {task.id}의 순위 업데이트 중 오류 발생: {e}")
                # 예외 발생 시에도 기존 current_rank 값을 유지하도록 함
                current_rank = task.current_rank

            difference_rank = (
                (task.current_rank - current_rank)
                if task.current_rank is not None and current_rank is not None
                else None
            )
            task.yesterday_rank = task.current_rank
            task.current_rank = current_rank
            task.difference_rank = difference_rank
            task.last_checked_date = timezone.now()
            
            # 오늘 날짜가 available_end_date보다 지난 경우, 예외 발생 여부와 상관없이 is_completed 업데이트
            if today > task.available_end_date:
                if not task.is_completed:
                    task.is_completed = True
                    task.ending_rank = task.current_rank
            task.save()
            
            if task.product:
                Ranking.objects.create(
                    product=task.product,
                    keyword=task.keyword,
                    rank=current_rank,
                    date_time=timezone.now(),
                    task=task,
                )
            else:
                logger.error(f"Task ID {task.id} has no associated product; skipping Ranking creation.")
            
            # Ranking 업데이트 후 각 작업의 상태를 업데이트
            update_task_status(task)
            
        messages.success(request, "모든 작업의 순위가 업데이트되었습니다.")
        return redirect('rankings:dashboard')
    else:
        return redirect('rankings:dashboard')
    
def traffic_create(request):
    if request.method == 'POST':
        form = TrafficForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '트래픽이 성공적으로 등록되었습니다.')
            return redirect('rankings:traffic_list')
    else:
        form = TrafficForm()
    return render(request, 'rankings/traffic_register.html', {'form': form})


def traffic_update(request, pk):
    traffic = get_object_or_404(Traffic, pk=pk)
    if request.method == 'POST':
        form = TrafficForm(request.POST, instance=traffic)
        if form.is_valid():
            form.save()
            messages.success(request, '트래픽이 성공적으로 수정되었습니다.')
            return redirect('rankings:traffic_list')
    else:
        form = TrafficForm(instance=traffic)
    return render(request, 'rankings/traffic_register.html', {'form': form})


from django.contrib.auth.decorators import login_required

@login_required
def traffic_cost_summary(request):
    # 필터링을 위한 GET 파라미터 가져오기
    selected_product_id = request.GET.get('product')
    selected_traffic_id = request.GET.get('traffic')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # 날짜 범위 설정
    today = datetime.today().date()  # 오늘 날짜
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        start_date = today - timedelta(days=14)
        end_date = today + timedelta(days=14)

    date_range = [start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]

    # 각 날짜별로 비용을 계산할 딕셔너리 생성
    date_costs = {date: 0 for date in date_range}

    # 현재 활성화된 작업 가져오기 (필터 적용)
    tasks = Task.objects.filter(
        available_start_date__lte=end_date,
        available_end_date__gte=start_date,
    )

    if selected_product_id:
        tasks = tasks.filter(product_id=selected_product_id)
    if selected_traffic_id:
        tasks = tasks.filter(traffic_id=selected_traffic_id)

    total_cost = 0  # 총 비용 초기화

    for task in tasks:
        # 작업의 실제 시작일과 종료일 계산
        task_start = max(task.available_start_date, start_date)
        task_end = min(task.available_end_date, end_date)

        # 작업에 연결된 트래픽 가져오기
        traffic = task.traffic
        if traffic and traffic.price and traffic.days:
            # 트래픽의 일일 비용 계산
            total_days = traffic.days
            daily_cost = traffic.price / total_days
            daily_cost = int(daily_cost)  # 소수점 이하 절삭

            # 작업 기간 내의 각 날짜에 비용 누적
            for n in range((task_end - task_start).days + 1):
                current_date = task_start + timedelta(days=n)
                if current_date in date_costs:
                    date_costs[current_date] += daily_cost
                    total_cost += daily_cost  # 총 비용에 일일 비용 추가

    # 날짜별 비용 데이터를 리스트로 변환하여 날짜 순으로 정렬
    date_costs = sorted(date_costs.items())

    # 날짜와 비용을 각각의 리스트로 분리
    date_labels = [date.strftime('%Y-%m-%d') for date, _ in date_costs]
    cost_data = [cost for _, cost in date_costs]

    # JSON 형식으로 변환
    date_labels_json = json.dumps(date_labels)
    cost_data_json = json.dumps(cost_data)

    # 모든 상품과 트래픽을 가져오기 (필터 선택을 위해)
    products = Product.objects.all()
    traffics = Traffic.objects.all()

    context = {
        'date_labels_json': date_labels_json,
        'cost_data_json': cost_data_json,
        'products': products,
        'traffics': traffics,
        'selected_product_id': int(selected_product_id) if selected_product_id else '',
        'selected_traffic_id': int(selected_traffic_id) if selected_traffic_id else '',
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'total_cost': total_cost,
    }
    return render(request, 'rankings/traffic_cost_summary.html', context)


# 로그인한 사용자만 접근 가능
# @login_required
# def my_data_view(request):
#     # 로그인한 사용자 본인의 데이터만 가져오기
#     user_data = MyModel.objects.all()
#     return render(request, 'my_template.html', {'user_data': user_data})

@login_required
def dashboard_view(request):
    products = Product.objects.all()
    traffic_data = Traffic.objects.all()
    # 기타 필요한 데이터 조회
    context = {
        'products': products,
        'traffic_data': traffic_data,
        # 기타 데이터
    }
    return render(request, 'dashboard.html', context)

@login_required
def product_list_view(request):
    if request.method == "POST":
        selected_products = request.POST.getlist('selected_products')
        if not selected_products:
            messages.error(request, "상품을 선택해주세요.")
            return redirect('rankings:product_list')

        if 'register_traffic' in request.POST:
            # 트래픽 등록 로직
            pass
        elif 'delete_products' in request.POST:
            # 선택한 상품 삭제 로직
            Product.objects.filter(id__in=selected_products).delete()
            messages.success(request, "선택한 상품을 삭제했습니다.")
            return redirect('rankings:product_list')

    # GET 요청 처리
    products_list = Product.objects.all()

    # 페이지네이션 설정
    page = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 100)
    per_page_options = [100, 300, 500, 1000]

    # per_page 값 검증
    try:
        per_page = int(per_page)
        if per_page not in per_page_options:
            per_page = 100
    except ValueError:
        per_page = 100

    paginator = Paginator(products_list, per_page)

    try:
        products = paginator.page(page)
    except PageNotAnInteger:
        # 페이지 번호가 정수가 아니면 첫 번째 페이지를 보여줍니다.
        products = paginator.page(1)
    except EmptyPage:
        # 페이지 번호가 범위를 벗어나면 마지막 페이지를 보여줍니다.
        products = paginator.page(paginator.num_pages)

    context = {
        'products': products,
        'per_page_options': per_page_options,
        'selected_per_page': per_page,
        'traffics': Traffic.objects.all(),  # Traffic 객체들을 추가
    }

    return render(request, 'rankings/product_list.html', context)

@login_required
def traffic_list(request):
    if request.method == 'POST':
        selected_traffics = request.POST.getlist('selected_traffics')
        if not selected_traffics:
            messages.error(request, "트래픽을 선택해주세요.")
            return redirect('rankings:traffic_list')

        if 'edit_traffics' in request.POST:
            # 트래픽 수정 로직
            return redirect('rankings:traffic_bulk_edit')  # 필요한 경우 수정
        elif 'delete_traffics' in request.POST:
            # 선택한 트래픽 삭제 로직
            Traffic.objects.filter(id__in=selected_traffics).delete()
            messages.success(request, "선택한 트래픽을 삭제했습니다.")
            return redirect('rankings:traffic_list')

    # GET 요청 처리
    traffics_list = Traffic.objects.all().order_by('-id')

    # 페이지네이션 설정
    page = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 100)
    per_page_options = [100, 300, 500, 1000]

    # per_page 값 검증
    try:
        per_page = int(per_page)
        if per_page not in per_page_options:
            per_page = 100
    except ValueError:
        per_page = 100

    paginator = Paginator(traffics_list, per_page)

    try:
        traffics = paginator.page(page)
    except PageNotAnInteger:
        traffics = paginator.page(1)
    except EmptyPage:
        traffics = paginator.page(paginator.num_pages)

    context = {
        'traffics': traffics,
        'per_page_options': per_page_options,
        'selected_per_page': per_page,
    }
    return render(request, 'rankings/traffic_list.html', context)

@login_required
def product_create_view(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save(commit=False)
            product.user = request.user  # 현재 로그인한 사용자 지정
            product.save()
            return redirect('product_list')
    else:
        form = ProductForm()
    return render(request, 'product_form.html', {'form': form})

@login_required
def completed_tasks_delete(request):
    if request.method == 'POST':
        selected_task_ids = request.POST.getlist('selected_tasks')
        if selected_task_ids:
            Task.objects.filter(id__in=selected_task_ids).delete()
            messages.success(request, "선택한 완료된 작업을 삭제했습니다.")
        else:
            messages.warning(request, "삭제할 작업을 선택해주세요.")
    return redirect('rankings:completed_tasks_list')

@login_required
def my_page(request):
    user_profile, created = UserProfile.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        if 'modify' in request.POST:
            # '수정' 버튼 클릭 시
            form = UserProfileForm(instance=user_profile)
            settings_completed = False
        else:
            # 폼 제출 시
            form = UserProfileForm(request.POST, instance=user_profile)
            if form.is_valid():
                form.save()
                messages.success(request, "셋팅이 완료되었습니다.")
                return redirect('rankings:my_page')
            else:
                messages.error(request, "입력한 정보를 확인해주세요.")
                settings_completed = False
    else:
        settings_completed = bool(user_profile.naver_client_id and user_profile.naver_client_secret)
        form = None if settings_completed else UserProfileForm(instance=user_profile)

    # 저장 후에도 NAVER_CLIENT_ID와 NAVER_CLIENT_SECRET를 템플릿에 전달
    client_id = user_profile.naver_client_id if user_profile.naver_client_id else ''
    client_secret = user_profile.naver_client_secret if user_profile.naver_client_secret else ''

    context = {
        'form': form,
        'settings_completed': settings_completed,
        'client_id': client_id,
        'client_secret': client_secret,
    }
    return render(request, 'rankings/my_page.html', context)

from django.db.models import Q



@login_required
def ad_single_summary(request):
    ads = Ad.objects.all()

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    today = date.today()
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = today - timedelta(days=30)
            end_date = today
    else:
        start_date = today - timedelta(days=30)
        end_date = today

    # 기간 필터링 적용
    ads = ads.filter(
        Q(start_date__lte=end_date, end_date__gte=start_date) |
        Q(start_date__gte=start_date, end_date__lte=end_date)
    )

    # 계산 로직 구현
    ad_data = []
    for ad in ads:
        # 필요한 계산 수행
        profit = ad.margin - ad.cost
        margin_rate = (ad.margin / ad.sales) * 100 if ad.sales else 0
        profit_rate = (profit / ad.sales) * 100 if ad.sales else 0
        roas = (ad.sales / ad.cost) * 100 if ad.cost else 0
        roi = (profit / ad.cost) * 100 if ad.cost else 0

        ad_data.append({
            'id': ad.id,
            'start_date': ad.start_date,
            'end_date': ad.end_date,
            'channel': ad.channel,
            'ad_name': ad.name,
            'category': ad.category,
            'product_name': ad.product,
            'sales': ad.sales,
            'margin': ad.margin,
            'cost': ad.cost,
            'profit': profit,
            'margin_rate': margin_rate,
            'profit_rate': profit_rate,
            'roas': roas,
            'roi': roi,
            'memo': ad.memo,
            'page_link': ad.page_link,
            'company': ad.company,
        })

    # 페이지네이션 설정
    page = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 10)
    per_page_options = [10, 30, 50, 100]

    # per_page 값 검증
    try:
        per_page = int(per_page)
        if per_page not in per_page_options:
            per_page = 10
    except ValueError:
        per_page = 10

    paginator = Paginator(ad_data, per_page)

    try:
        ads_page = paginator.page(page)
    except PageNotAnInteger:
        ads_page = paginator.page(1)
    except EmptyPage:
        ads_page = paginator.page(paginator.num_pages)

    context = {
        'ads': ads_page,
        'start_date': start_date,
        'end_date': end_date,
        'per_page_options': per_page_options,
        'selected_per_page': per_page,
    }
    return render(request, 'rankings/ad_single_summary.html', context)



@login_required
def ad_create(request):
    if request.method == 'POST':
        if 'upload_excel' in request.POST:
            # 엑셀 업로드 처리
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, "엑셀 파일을 선택해주세요.")
                return redirect('rankings:ad_create')

            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                ads_data = []
                error_occurred = False  # 에러 발생 여부를 추적
                for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # 헤더 제외
                    # 엑셀 파일의 열 순서에 따라 데이터 매핑
                    start_date = row[0]
                    end_date = row[1]
                    channel = row[2]
                    name = row[3]
                    category = row[4]
                    product = row[5]
                    sales = row[6]
                    margin = row[7]
                    cost = row[8]
                    memo = row[9]
                    page_link = row[10]
                    company = row[11]

                    # 필수 값 검증
                    missing_fields = []
                    if not start_date:
                        missing_fields.append('시작일')
                    if not end_date:
                        missing_fields.append('종료일')
                    if not channel:
                        missing_fields.append('채널')
                    if not name:
                        missing_fields.append('광고명')
                    if missing_fields:
                        error_message = f"{idx}행: 필수 필드 누락 - {', '.join(missing_fields)}"
                        messages.error(request, error_message)
                        error_occurred = True
                        continue  # 다음 행으로 이동

                    # 날짜 형식 변환 및 유효성 검사
                    try:
                        if isinstance(start_date, datetime):
                            start_date = start_date.date()
                        elif isinstance(start_date, date):
                            pass
                        elif isinstance(start_date, str):
                            # 여러 날짜 형식 시도
                            for fmt in ('%m/%d/%y', '%Y-%m-%d', '%Y/%m/%d', '%m-%d-%Y'):
                                try:
                                    start_date = datetime.strptime(start_date, fmt).date()
                                    break
                                except ValueError:
                                    continue
                            else:
                                raise ValueError
                        else:
                            raise ValueError
                    except ValueError:
                        error_message = f"{idx}행: 시작일의 날짜 형식이 올바르지 않습니다."
                        messages.error(request, error_message)
                        error_occurred = True
                        continue  # 다음 행으로 이동

                    try:
                        if isinstance(end_date, datetime):
                            end_date = end_date.date()
                        elif isinstance(end_date, date):
                            pass
                        elif isinstance(end_date, str):
                            # 여러 날짜 형식 시도
                            for fmt in ('%m/%d/%y', '%Y-%m-%d', '%Y/%m/%d', '%m-%d-%Y'):
                                try:
                                    end_date = datetime.strptime(end_date, fmt).date()
                                    break
                                except ValueError:
                                    continue
                            else:
                                raise ValueError
                        else:
                            raise ValueError
                    except ValueError:
                        error_message = f"{idx}행: 종료일의 날짜 형식이 올바르지 않습니다."
                        messages.error(request, error_message)
                        error_occurred = True
                        continue  # 다음 행으로 이동

                    # 날짜를 문자열로 변환 (YYYY-MM-DD 형식)
                    start_date_str = start_date.strftime('%Y-%m-%d')
                    end_date_str = end_date.strftime('%Y-%m-%d')

                    # 광고 데이터를 리스트에 추가
                    ads_data.append({
                        'start_date': start_date_str,
                        'end_date': end_date_str,
                        'channel': channel,
                        'name': name,
                        'category': category,
                        'product': product,
                        'sales': sales if sales else '',
                        'margin': margin if margin else '',
                        'cost': cost if cost else '',
                        'memo': memo or '',
                        'page_link': page_link or '',
                        'company': company or '',
                    })
                if error_occurred:
                    # 오류가 발생한 경우, ads_data와 함께 폼을 렌더링하여 사용자에게 오류를 표시
                    return render(request, 'rankings/ad_form.html', {'ads_data': ads_data, 'today': timezone.now().strftime('%Y-%m-%d')})
                else:
                    # 오류가 없으면 ads_data를 컨텍스트에 전달하여 폼에 표시
                    return render(request, 'rankings/ad_form.html', {'ads_data': ads_data, 'today': timezone.now().strftime('%Y-%m-%d')})
            except Exception as e:
                messages.error(request, f"엑셀 파일 처리 중 오류 발생: {str(e)}")
                return render(request, 'rankings/ad_form.html', {'today': timezone.now().strftime('%Y-%m-%d')})

        else:
            # 폼 데이터 처리
            start_dates = request.POST.getlist('start_date')
            end_dates = request.POST.getlist('end_date')
            channels = request.POST.getlist('channel')
            names = request.POST.getlist('name')
            categories = request.POST.getlist('category')
            products = request.POST.getlist('product')
            sales_list = request.POST.getlist('sales')
            margins = request.POST.getlist('margin')
            costs = request.POST.getlist('cost')
            memos = request.POST.getlist('memo')
            page_links = request.POST.getlist('page_link')
            companies = request.POST.getlist('company')
            print("start_dates:", start_dates)
            print("end_dates:", end_dates)
            print("channels:", channels)
            print("names:", names)
            if not names:
                messages.error(request, "등록할 광고 데이터가 없습니다.")
                return redirect('rankings:ad_create')

            # 모든 광고 항목을 저장
            error_occurred = False
            for i in range(len(names)):
                # 필수 값 검증
                if not (start_dates[i] and end_dates[i] and channels[i] and names[i]):
                    messages.error(request, f"{i + 1}번째 행의 필수 값을 모두 입력해주세요.")
                    error_occurred = True
                    continue  # 필수 값이 없으면 건너뜁니다

                # 데이터 타입 변환 및 유효성 검사
                try:
                    sales = float(sales_list[i]) if sales_list[i] else 0
                    margin = float(margins[i]) if margins[i] else 0
                    cost = float(costs[i]) if costs[i] else 0
                except ValueError:
                    messages.error(request, f"{i + 1}번째 행의 숫자 필드에 올바른 값을 입력해주세요.")
                    error_occurred = True
                    continue  # 값이 유효하지 않으면 건너뜁니다

                # 광고 생성
                Ad.objects.create(
                    start_date=start_dates[i],
                    end_date=end_dates[i],
                    channel=channels[i],
                    name=names[i],
                    category=categories[i] or '',
                    product=products[i] or '',
                    sales=sales,
                    margin=margin,
                    cost=cost,
                    memo=memos[i] or '',
                    page_link=page_links[i] or '',
                    company=companies[i] or '',
                )
            if error_occurred:
                messages.error(request, "일부 광고 항목이 등록되지 않았습니다. 오류를 확인해주세요.")
                # 입력된 데이터를 폼에 다시 표시하기 위해 컨텍스트에 전달
                ads_data = []
                for i in range(len(names)):
                    ads_data.append({
                        'start_date': start_dates[i],
                        'end_date': end_dates[i],
                        'channel': channels[i],
                        'name': names[i],
                        'category': categories[i],
                        'product': products[i],
                        'sales': sales_list[i],
                        'margin': margins[i],
                        'cost': costs[i],
                        'memo': memos[i],
                        'page_link': page_links[i],
                        'company': companies[i],
                    })
                return render(request, 'rankings/ad_form.html', {'ads_data': ads_data, 'today': timezone.now().strftime('%Y-%m-%d')})
            else:
                messages.success(request, "광고를 성공적으로 등록했습니다.")
                return redirect('rankings:ad_single_summary')
    
    else:
        today = timezone.now().date().strftime('%Y-%m-%d')
        return render(request, 'rankings/ad_form.html', {'today': today})


     
@login_required
def ad_update(request, pk):
    ad = get_object_or_404(Ad, pk=pk)
    if request.method == 'POST':
        form = AdForm(request.POST, instance=ad)
        if form.is_valid():
            form.save()
            messages.success(request, "광고가 성공적으로 수정되었습니다.")
            return redirect('rankings:ad_single_summary')
    else:
        form = AdForm(instance=ad)
    return render(request, 'rankings/ad_form.html', {'form': form})

# views.py

@login_required
def ad_bulk_edit(request):
    AdFormSet = modelformset_factory(Ad, form=AdForm, extra=0)
    if request.method == 'POST':
        if 'edit_ads' in request.POST:
            # 선택된 광고 IDs를 가져옵니다.
            selected_ids = request.POST.getlist('ad_ids')
            if not selected_ids:
                messages.error(request, "수정할 광고를 선택해주세요.")
                return redirect('rankings:ad_single_summary')
            ads = Ad.objects.filter(id__in=selected_ids)
            # **각 필드의 초기값을 소수점 제거하여 설정**
            initial_data = []
            for ad in ads:
                initial_data.append({
                    'id': ad.id,
                    'start_date': ad.start_date,
                    'end_date': ad.end_date,
                    'channel': ad.channel,
                    'name': ad.name,
                    'category': ad.category,
                    'product': ad.product,
                    'sales': int(ad.sales),
                    'margin': int(ad.margin),
                    'cost': int(ad.cost),
                    'memo': ad.memo,
                    'page_link': ad.page_link,
                    'company': ad.company,
                })
            formset = AdFormSet(queryset=ads, initial=initial_data)
        else:
            # 폼셋이 제출된 경우
            formset = AdFormSet(request.POST)
            if formset.is_valid():
                # 저장하기 전에 소수점 이하 자릿수를 제거합니다.
                instances = formset.save(commit=False)
                for instance in instances:
                    instance.sales = int(instance.sales)
                    instance.margin = int(instance.margin)
                    instance.cost = int(instance.cost)
                    instance.save()
                messages.success(request, "선택한 광고가 성공적으로 수정되었습니다.")
                return redirect('rankings:ad_single_summary')
            else:
                messages.error(request, "폼에 오류가 있습니다. 확인해주세요.")
    else:
        messages.error(request, "잘못된 접근입니다.")
        return redirect('rankings:ad_single_summary')
    return render(request, 'rankings/ad_bulk_edit.html', {'formset': formset})




@login_required
def ad_delete(request, pk):
    # `Ad` 모델에서 `pk`에 해당하는 객체 가져오기
    ad = get_object_or_404(Ad, pk=pk)  # user 필터링은 필요에 따라 추가

    # POST 요청이 확인되면 객체 삭제
    if request.method == 'POST':
        ad.delete()
        messages.success(request, '광고가 성공적으로 삭제되었습니다.')
        return redirect('rankings:ad_single_summary')  # 리디렉션할 URL 네임스페이스 확인 필요

    # 삭제 확인을 위한 템플릿 렌더링
    return render(request, 'rankings/ad_confirm_delete.html', {'ad': ad})

def ad_upload(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            df = pd.read_excel(excel_file)

            # 데이터프레임의 각 행을 Ad 모델로 저장
            for index, row in df.iterrows():
                Ad.objects.create(
                    name=row['이름'],
                    start_date=row['시작일자'],
                    end_date=row['종료일자'],
                    # 나머지 필드들에 맞게 매핑
                    # 예: product=row['상품명'], cost=row['비용'], ...
                )
            return redirect('rankings:ad_single_summary')
    else:
        form = ExcelUploadForm()
    return render(request, 'rankings/ad_upload.html', {'form': form})

@login_required
def ad_delete_multiple(request):
    if request.method == 'POST':
        ad_ids = request.POST.getlist('ad_ids')
        if ad_ids:
            ads_to_delete = Ad.objects.filter(id__in=ad_ids)
            ads_to_delete.delete()
            messages.success(request, f"{len(ads_to_delete)}개의 광고가 삭제되었습니다.")
        else:
            messages.error(request, "삭제할 광고를 선택하세요.")
    return redirect('rankings:ad_single_summary')

@login_required
def download_sample_excel(request):
    # 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active

    # 헤더 설정
    headers = [
        '시작일', '종료일', '채널', '광고명', '카테고리', '상품명',
        '매출', '마진', '비용', '메모', '페이지 링크', '업체'
    ]
    ws.append(headers)

    # 데이터 추가
    data = [
        '11/1/24', '11/10/24', '인스타', '공구', '가구', '소파',
        1000000, 500000, 300000, '메모 내용', 'http://example.com', '업체A'
    ]
    ws.append(data)

    # 열 너비 조정 및 정렬
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = length + 2
        for cell in column_cells:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 엑셀 파일로 응답
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sample_ad_data.xlsx'
    wb.save(response)
    return response





def custom_404(request, exception):
    return render(request, 'errors/404.html', status=404)

def custom_500(request):
    return render(request, 'errors/500.html', status=500)



# views.py

from django.http import JsonResponse
from django.utils import timezone
from django.utils.timezone import localtime

def get_ranking_data(request):
    task_id = request.GET.get('task_id')
    if not task_id:
        return JsonResponse({'error': 'Task ID is required.'}, status=400)

    try:
        task = Task.objects.get(id=task_id)
    except Task.DoesNotExist:
        return JsonResponse({'error': 'Task not found.'}, status=404)

    rankings = Ranking.objects.filter(task=task).order_by('-date_time')

    # 시작 순위 가져오기
    start_rank = task.start_rank if task.start_rank else 1000

    rankings_list = []
    previous_rank = None
    for ranking in rankings:
        # 날짜와 시간을 한국 시간대로 변환
        local_date_time = localtime(ranking.date_time)
        date_time_str = local_date_time.strftime('%Y-%m-%d %H:%M:%S')

        rank = ranking.rank if ranking.rank else 1000

        # 등락폭 계산
        if previous_rank is not None:
            difference = rank - previous_rank
        else:
            difference = 0

        # 순위 변동을 숫자로만 반환
        # 양수: 순위가 내려감 (숫자 커짐)
        # 음수: 순위가 올라감 (숫자 작아짐)
        difference_value = difference

        # 시작 대비 등락폭 계산
        start_difference = rank - start_rank

        rankings_list.append({
            'date_time': date_time_str,
            'keyword': ranking.keyword.name,
            'rank': rank,
            'difference': difference_value,           # 숫자 값으로만 반환
            'start_difference': start_difference,     # 숫자 값으로만 반환
        })

        previous_rank = rank

    return JsonResponse({'rankings': rankings_list})

from .api_clients import get_rel_keywords,get_data_lab_trend

def get_estimated_search_volume(keyword):
    import datetime
    """
    1) 최근 한 달(30일) 범위를 계산.
    2) DataLab API: time_unit='date'로 일별 ratio 조회.
    3) 광고 API: 해당 키워드의 monthlyPcQcCnt(월간 검색수).
    4) ratio * (monthlyPcQcCnt / 100) => 일별 추정 검색량.
    
    반환 예시:
    {
      'keyword': '방석',
      'startDate': '2023-02-08',
      'endDate': '2023-03-10',
      'results': [
          {'period': '2023-02-08', 'ratio': 40.5, 'estimatedVolume': 16200},
          {'period': '2023-02-09', 'ratio': 100.0, 'estimatedVolume': 40000},
          ...
      ]
    }
    """
    # 1) 최근 한 달 범위
    today = datetime.date.today()
    start_date = (today - datetime.timedelta(days=30)).strftime("%Y-%m-%d")
    end_date = today.strftime("%Y-%m-%d")
    
    # 2) DataLab API (일별 ratio)
    datalab_data = get_data_lab_trend([keyword], start_date, end_date, time_unit='date', device='pc')
    if not datalab_data or 'results' not in datalab_data:
        logger.error("DataLab API 실패 또는 결과 없음.")
        return None
    
    # 3) 광고 API
    ad_data = get_rel_keywords([keyword])
    if ad_data is None or ad_data.empty:
        logger.error("광고 API 실패 또는 결과 없음.")
        return None
    
    # 해당 키워드 row만 필터링
    df_filtered = ad_data[ad_data['relKeyword'] == keyword]
    if df_filtered.empty:
        logger.error("광고 API 결과 중 '%s' 키워드가 없음.", keyword)
        monthly_pc_cnt = 0
    else:
        monthly_pc_cnt = int(df_filtered.iloc[0]['monthlyPcQcCnt'])
    
    # 4) ratio * (monthly_pc_cnt / 100)
    result_list = []
    for result_item in datalab_data['results']:
        # 보통 results에 1개의 item만 들어있음 (키워드 1개 호출)
        data_arr = result_item.get('data', [])
        for row in data_arr:
            period_str = row.get('period')
            try:
                ratio_val = float(row.get('ratio'))
            except (ValueError, TypeError):
                ratio_val = 0.0
            estimated = int(ratio_val * (monthly_pc_cnt / 100.0))
            result_list.append({
                'period': period_str,
                'ratio': ratio_val,
                'estimatedVolume': estimated
            })
    
    return {
        'keyword': keyword,
        'startDate': start_date,
        'endDate': end_date,
        'results': result_list
    }
    
@login_required
def get_keyword_volume_recent_month(request):
    """
    모달창 등에서 GET 파라미터로 'keyword'를 받아,
    최근 한 달 기준으로 일별 검색량을 추정하여 반환.
    
    호출 예:
    GET /traffic/api/get_keyword_volume_recent_month/?keyword=방석
    """
    raw = request.GET.get('keyword', '')
    # 모든 공백문자를 제거
    keyword = re.sub(r'\s+', '', raw)
    if not keyword:
        return JsonResponse({'error': 'No keyword provided.'}, status=400)
    
    result = get_estimated_search_volume(keyword)
    if not result:
        return JsonResponse({'error': 'Failed to compute search volume.'}, status=500)
    
    return JsonResponse({'success': True, 'data': result})

    
@login_required
def download_selected_products_excel(request):
    product_ids = request.GET.get('product_ids')
    if not product_ids:
        messages.error(request, '상품이 선택되지 않았습니다.')
        return redirect('rankings:product_list')

    ids = [int(pk) for pk in product_ids.split(',') if pk.isdigit()]
    products = Product.objects.filter(id__in=ids)

    # 오늘 날짜 기준
    today = date.today()
    start_date = today + timedelta(days=1)
    end_date   = today + timedelta(days=10)

    # 엑셀 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '작업 등록'

    # 헤더 작성
    headers = ['상품ID', '상품명', '카테고리', '순위조회키워드', 'URL 타입', 'URL', '메모', '트래픽명', '이용권 수', '이용가능 시작일자', '이용가능 종료일자','단일','원부']
    ws.append(headers)



    # 데이터 작성
    for idx, prod in enumerate(products, start=1):
        row = [
            prod.id,
            prod.name,
            prod.category or '',            # ← 카테고리가 str 이면 바로 사용
            prod.search_keyword or '',
            '',
            '',
            '',                             # 메모
            '',                             # 트래픽명
            1,                              # 이용권 수
            start_date.strftime('%Y-%m-%d'),
            end_date.strftime('%Y-%m-%d'),
            prod.single_product_link or '',
            prod.original_link       or '',
        ]
        ws.append(row)


    # 컬럼 너비 조절
    for col_idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # 엑셀 파일 응답
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = 'task_template.xlsx'
    response['Content-Disposition'] = f'attachment; filename={smart_str(filename)}'
    wb.save(response)
    return response

@login_required
def upload_excel_data(request):
    if request.method != 'POST':
        messages.error(request, '잘못된 요청입니다.')
        return redirect('rankings:product_list')

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, '엑셀 파일을 선택해주세요.')
        return redirect('rankings:product_list')

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # 1) 헤더 읽기
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        expected_headers = [
            '번호', '상품명', '카테고리', '순위조회키워드',
            '트래픽명',     # <- 엑셀 템플릿에 반드시 이 이름으로 컬럼 추가
            'URL 타입', 'URL', '메모',
            '이용권 수', '이용가능 시작일자', '이용가능 종료일자',
        ]
        if headers != expected_headers:
            messages.error(request, '엑셀 헤더가 올바르지 않습니다. “트래픽명” 컬럼이 빠졌거나 순서가 다를 수 있습니다.')
            return redirect('rankings:product_list')

        # 2) 데이터 읽어서 task_data 로 변환
        task_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            task_data.append(dict(zip(headers, row)))
        if not task_data:
            messages.error(request, '엑셀에 데이터가 없습니다.')
            return redirect('rankings:product_list')

        # 3) products, initial_data 준비
        products = []
        initial_data = {}
        for idx, data in enumerate(task_data, start=1):
            prod_name = data.get('상품명')
            if not prod_name:
                continue
            try:
                product = Product.objects.get(name=prod_name)
            except Product.DoesNotExist:
                continue
            products.append(product)

            # --- 여기서 트래픽명으로 조회 ---
            traffic_name = (data.get('트래픽명') or '').strip()
            traffic_id = None
            if traffic_name:
                try:
                    traffic = Traffic.objects.get(name=traffic_name)
                    traffic_id = traffic.id
                except Traffic.DoesNotExist:
                    messages.warning(request, f'행 {idx}: “{traffic_name}” 트래픽을 찾을 수 없습니다.')
            # --------------------------------

            initial_data[product.id] = {
                'keyword':              data.get('순위조회키워드', ''),
                'url_type':             data.get('URL 타입', ''),
                'url':                  data.get('URL', ''),
                'memo':                 data.get('메모', ''),
                'traffic_name':         traffic_name,
                'traffic_id':           traffic_id or '',
                'ticket_count':         data.get('이용권 수', ''),
                'available_start_date': data.get('이용가능 시작일자', ''),
                'available_end_date':   data.get('이용가능 종료일자', ''),
            }

            

        traffics = Traffic.objects.all()
        return render(request, 'rankings/task_register.html', {
            'products': products,
            'traffics': traffics,
            'initial_data': initial_data,
        })

    except Exception as e:
        messages.error(request, f'엑셀 처리 중 오류: {e}')
        return redirect('rankings:product_list')



    
@login_required
def task_register_from_excel(request):
    if request.method == 'POST':
        task_data_json = request.POST.get('task_data')
        if not task_data_json:
            return JsonResponse({'success': False, 'error': '엑셀 데이터를 찾을 수 없습니다.'})

        try:
            task_data = json.loads(task_data_json)
            request.session['uploaded_task_data'] = task_data
            return JsonResponse({'success': True})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': '엑셀 데이터 파싱 중 오류가 발생했습니다.'})

    else:
        task_data = request.session.get('uploaded_task_data')
        if not task_data:
            messages.error(request, '엑셀 데이터를 찾을 수 없습니다.')
            return redirect('rankings:product_list')

        # task_data를 사용하여 products와 initial_data 구성
        products = []
        initial_data = {}
        for idx, data in enumerate(task_data):
            product_name = data.get('상품명')
            if not product_name:
                continue  # 상품명이 없으면 건너뜁니다.
            try:
                product = Product.objects.get(name=product_name)
            except Product.DoesNotExist:
                continue
            products.append(product)
            product_id = product.id
            initial_data[product_id] = {
                'keyword': data.get('순위조회키워드', ''),
                'url_type': data.get('URL 타입', ''),
                'url': data.get('URL', ''),
                'memo': data.get('메모', ''),
                'product_name': product_name,
                'traffic_id': '',  # 필요 시 추가
                'ticket_count': data.get('이용권 수', ''),
                'available_start_date': data.get('이용가능 시작일자', ''),
                'available_end_date': data.get('이용가능 종료일자', ''),
            }

        traffics = Traffic.objects.all()
        context = {
            'products': products,
            'traffics': traffics,
            'initial_data': initial_data,
        }
        return render(request, 'rankings/task_register.html', context)

#모니터링 시작 
from .models import RankingMonitoring, KeywordRanking

    

from django.db.models import Max


@login_required
def ranking_monitoring_list(request):
    from sales_management.models import NaverAdShoppingProduct
    # 0) 검색어(q) 가져오기
    q = request.GET.get('q', '').strip()

    # 1) 랭킹 객체 가져오기 & 검색 필터
    rankings = RankingMonitoring.objects.all().prefetch_related('keywords')
    if q:
        rankings = rankings.filter(product_name__icontains=q)

    # 2) 이미지 URL 맵
    product_ids = [r.product_id for r in rankings]
    img_qs = NaverAdShoppingProduct.objects.filter(
        product_id_of_mall__in=product_ids
    ).values('product_id_of_mall', 'product_image_url')
    img_map = {
        obj['product_id_of_mall']: obj['product_image_url']
        for obj in img_qs
    }

    # 3) 상품 링크 맵
    prod_qs = Product.objects.filter(single_product_mid__in=product_ids).values(
        'single_product_mid', 'single_product_link'
    )
    link_map = {
        o['single_product_mid']: o['single_product_link']
        for o in prod_qs
    }

    # 4) 날짜 컬럼 수집
    all_dates = set()
    for r in rankings:
        for kw in r.keywords.all():
            all_dates.add(kw.update_at.strftime("%Y-%m-%d"))
    date_columns = sorted(all_dates, reverse=True)
    first_date = date_columns[0] if date_columns else None

    # 5) 각 랭킹 객체에 추가 데이터 세팅
    for r in rankings:
        # 이미지 & 링크
        r.product_img  = img_map.get(r.product_id, '/static/placeholder.svg')
        r.product_link = link_map.get(r.product_id, '#')

        # 피벗 준비
        pivot_data = {}
        seen = set()
        for kwobj in r.keywords.all():
            date_str = kwobj.update_at.strftime("%Y-%m-%d")
            if (kwobj.keyword, date_str) in seen:
                continue
            seen.add((kwobj.keyword, date_str))

            entry = pivot_data.setdefault(
                kwobj.keyword,
                {'search_volume': kwobj.search_volume, 'ranks': {}}
            )
            entry['ranks'][date_str] = kwobj.rank

        # 순위 계산
        for info in pivot_data.values():
            if date_columns:
                latest = date_columns[0]
                info['latest_rank'] = info['ranks'].get(latest, 0)

                if len(date_columns) > 1:
                    prev = date_columns[1]
                    prev_rank = info['ranks'].get(prev, 0)
                    info['previous_rank']   = prev_rank
                    info['rank_change']     = prev_rank - info['latest_rank']
                    info['abs_rank_change'] = abs(info['rank_change'])
                else:
                    info['previous_rank'] = info['rank_change'] = info['abs_rank_change'] = 0
            else:
                info['latest_rank'] = info['previous_rank'] = info['rank_change'] = info['abs_rank_change'] = 0

        # 정렬된 피벗 데이터 (검색량 기준)
        r.sorted_pivot_data = sorted(
            pivot_data.items(),
            key=lambda item: item[1]['search_volume'],
            reverse=True
        )

        # 5.1) 최신 순위만 뽑아서 매핑
        latest_date = r.keywords.aggregate(latest=Max('update_at'))['latest']
        if latest_date:
            recent_qs = r.keywords.filter(update_at=latest_date)
            r.latest_ranks = {kr.keyword: kr.rank for kr in recent_qs}
        else:
            r.latest_ranks = {}

    # 6) 컨텍스트 전달
    context = {
        'rankings':     rankings,
        'date_columns': date_columns,
        'first_date':   first_date,
        'search_query': q,
    }
    return render(request, 'rankings/ranking_monitoring.html', context)









@login_required
def add_monitoring(request):
    if request.method == 'POST':
        selected_products_str = request.POST.get('selected_products', '')
        monitoring_keywords   = request.POST.get('monitoring_keywords', '')

        selected_product_ids = [x.strip() for x in selected_products_str.split(',') if x.strip()]
        keyword_list         = [kw.strip() for kw in monitoring_keywords.split(',') if kw.strip()]

        for product_pk in selected_product_ids:
            try:
                product = Product.objects.get(pk=product_pk)
            except Product.DoesNotExist:
                continue

            ranking_obj, created = RankingMonitoring.objects.get_or_create(
                product_id=product.single_product_mid,
                defaults={
                    'product_url':  product.single_product_link,
                    'product_name': product.name,
                }
            )
            if not created:
                ranking_obj.product_url  = product.single_product_link
                ranking_obj.product_name = product.name
                ranking_obj.save()

            unique_keywords = list(dict.fromkeys(keyword_list))
            best_rank       = 1001
            best_link       = product.single_product_link

            for kw in unique_keywords:
                if ranking_obj.keywords.filter(keyword=kw).exists():
                    continue

                # single_link 순위 조회
                try:
                    rank_single = get_naver_rank(kw, product.single_product_link)
                except TypeError:
                    messages.error(request, f"링크 형식 오류: [{product.name}] 단일 링크를 알림에 등록해 주세요.")
                    rank_single = -1

                # original_link 순위 조회
                try:
                    rank_original = get_naver_rank(kw, product.original_link)
                except TypeError:
                    messages.error(request, f"링크 형식 오류: [{product.name}] 오리지널 링크를 알림에 등록해 주세요.")
                    rank_original = -1

                # 최종 순위 결정
                if rank_single == -1 and rank_original == -1:
                    final_rank          = 1001
                    is_original_better  = False
                elif rank_single == -1:
                    final_rank          = rank_original
                    is_original_better  = True
                elif rank_original == -1:
                    final_rank          = rank_single
                    is_original_better  = False
                else:
                    if rank_original < rank_single:
                        final_rank         = rank_original
                        is_original_better = True
                    else:
                        final_rank         = rank_single
                        is_original_better = False

                if final_rank < best_rank:
                    best_rank = final_rank
                    best_link = product.original_link if is_original_better else product.single_product_link

                KeywordRanking.objects.create(
                    ranking           = ranking_obj,
                    keyword           = kw,
                    rank              = final_rank,
                    is_original_better= is_original_better
                )

            ranking_obj.product_url = best_link
            ranking_obj.save()

            # 메인 키워드 (최대 3개) 설정
            if unique_keywords:
                ranking_obj.main_keyword1      = unique_keywords[0] if len(unique_keywords) >= 1 else ''
                ranking_obj.main_keyword2      = unique_keywords[1] if len(unique_keywords) >= 2 else ''
                ranking_obj.main_keyword3      = unique_keywords[2] if len(unique_keywords) >= 3 else ''
                ranking_obj.main_keyword1_rank = 0 if ranking_obj.main_keyword1 else None
                ranking_obj.main_keyword2_rank = 0 if ranking_obj.main_keyword2 else None
                ranking_obj.main_keyword3_rank = 0 if ranking_obj.main_keyword3 else None
                ranking_obj.save()

        messages.success(request, "모니터링 등록이 완료되었습니다.")
        return redirect('rankings:product_list')

    return redirect('rankings:product_list')



@login_required
def update_monitoring_search(request):
    """
    등록된 모든 KeywordRanking 중
    키워드별로 한 번씩만 get_rel_keywords() 호출하여
    한 달 검색량을 업데이트.
    """
    # 1) 중복 제거된 키워드 리스트만 뽑기
    unique_keywords = KeywordRanking.objects\
        .values_list('keyword', flat=True)\
        .distinct()

    updated_count = 0

    for kw in unique_keywords:
        # API 호출은 키워드당 한 번만!
        df = get_rel_keywords([kw])
        if df is None or df.empty:
            continue

        total_search = int(df.iloc[0].get('totalSearchCount', 0))

        # 같은 키워드를 가진 레코드 중, 실제 값이 바뀌는 것만 한꺼번에 업데이트
        changed = KeywordRanking.objects\
            .filter(keyword=kw)\
            .exclude(search_volume=total_search)\
            .update(search_volume=total_search)

        updated_count += changed

    messages.success(request, f"검색량 업데이트 완료! ({updated_count}건 반영)")
    return redirect('rankings:ranking_monitoring_list')


@login_required
def update_monitoring_rank(request):
    """
    오늘 날짜에 대한 순위만 갱신.
    (이전 날짜의 레코드는 건드리지 않음)
    """
    from django.utils import timezone
    today = timezone.now().date()

    updated_count = 0
    # (1) (키워드, 상품URL) 쌍을 저장할 세트
    seen_pairs = set()

    # 모든 RankingMonitoring + 해당 키워드
    rankings = RankingMonitoring.objects.all().prefetch_related('keywords')

    for r in rankings:
        for kwobj in r.keywords.all():
            # 키워드 문자열 & 현재 모니터링의 상품URL을 합쳐 유니크한 쌍으로 만든다
            pair = (kwobj.keyword, r.product_url)

            # (2) 이미 같은 (키워드, URL)을 처리했다면 건너뛰기
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)

            # (3) 실제 API 호출
            new_rank = get_naver_rank(kwobj.keyword, r.product_url)
            if new_rank == -1:
                
                # 수정: 그래도 DB에 (rank=-1, update_at=today)로 insert
                KeywordRanking.objects.create(
                    ranking=r,
                    keyword=kwobj.keyword,
                    rank=-1,
                    update_at=today
                )
                updated_count += 1
                continue

            # (4) 오늘 날짜 레코드만 찾아서 업데이트 or 생성
            existing_today = KeywordRanking.objects.filter(
                ranking=r, 
                keyword=kwobj.keyword,
                update_at=today
            ).first()

            if existing_today:
                # 오늘 날짜 레코드가 있으면 rank 업데이트
                existing_today.rank = new_rank
                existing_today.save()
            else:
                # 오늘 날짜 레코드가 없으면 새로 생성
                KeywordRanking.objects.create(
                    ranking=r,
                    keyword=kwobj.keyword,
                    rank=new_rank,
                    update_at=today
                )
            updated_count += 1

    messages.success(request, f"순위 업데이트 완료! ({updated_count}건 반영)")
    return redirect('rankings:ranking_monitoring_list')


# views.py
import json
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from .models import RankingMonitoring

@login_required
@require_POST
def update_main_keywords(request):
    """
    AJAX POST로 전달된 상품의 메인키워드 목록을 DB에 저장합니다.
    요청 JSON 예시:
    {
       "product_id": "P001",
       "main_keywords": [
           {"keyword": "키워드1", "rank": 3},
           {"keyword": "키워드2", "rank": 5},
           {"keyword": "키워드3", "rank": 1}
       ]
    }
    """
    try:
        data = json.loads(request.body)
        product_id = data.get("product_id")
        main_keywords = data.get("main_keywords", [])  # 최대 3개

        # RankingMonitoring 레코드 찾기
        ranking_obj = RankingMonitoring.objects.get(product_id=product_id)

        # 최대 3개까지 업데이트
        # (없으면 "", rank=None)
        ranking_obj.main_keyword1 = main_keywords[0]["keyword"] if len(main_keywords) >= 1 else ""
        ranking_obj.main_keyword1_rank = main_keywords[0]["rank"] if len(main_keywords) >= 1 else None

        ranking_obj.main_keyword2 = main_keywords[1]["keyword"] if len(main_keywords) >= 2 else ""
        ranking_obj.main_keyword2_rank = main_keywords[1]["rank"] if len(main_keywords) >= 2 else None

        ranking_obj.main_keyword3 = main_keywords[2]["keyword"] if len(main_keywords) >= 3 else ""
        ranking_obj.main_keyword3_rank = main_keywords[2]["rank"] if len(main_keywords) >= 3 else None

        ranking_obj.save()

        return JsonResponse({"status": "success"})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})
    

@login_required
def ranking_monitoring_detail_list(request):
    from sales_management.models import NaverAdShoppingProduct  # import 추가

    product_id = request.GET.get('product_id')
    ranking_obj = get_object_or_404(RankingMonitoring, product_id=product_id)

    # 1) 날짜 컬럼 수집
    all_dates = {
        kw.update_at.strftime("%Y-%m-%d")
        for kw in ranking_obj.keywords.all()
    }
    date_columns = sorted(all_dates, reverse=True)

    # 2) 피벗 데이터 구성 (키워드 → {search_volume, ranks: {date: rank}})
    pivot = {}
    for kw in ranking_obj.keywords.all():
        dt = kw.update_at.strftime("%Y-%m-%d")
        entry = pivot.setdefault(kw.keyword, {
            'search_volume': kw.search_volume,
            'ranks': {}
        })
        entry['ranks'][dt] = kw.rank

    # 3) rank_list 만들기
    rank_list = []
    for keyword, data in pivot.items():
        history = [
            data['ranks'].get(dt, '—')
            for dt in date_columns
        ]
        rank_list.append({
            'keyword':        keyword,
            'search_volume':  data['search_volume'],
            'history':        history
        })

    # 4) 외부 모델에서 이미지 URL 가져오기
    nav_prod = NaverAdShoppingProduct.objects.filter(
        product_id_of_mall=product_id
    ).first()
    image_url = nav_prod.product_image_url if nav_prod else "/static/placeholder.svg"
    category_path = nav_prod.category_path if nav_prod and nav_prod.category_path else ""
    prod_obj = Product.objects.filter(single_product_mid=product_id).first()
    if prod_obj:
        product_link = prod_obj.single_product_link
    else:
        product_link = ranking_obj.product_url

    # 5) product info
    product = {
        'id':        ranking_obj.product_id,
        'title':     ranking_obj.product_name,
        'image_url': image_url,
        'category_path':  category_path,      # ← 여기에 넣어줌
        'tags':      [],  # 필요시 채우세요
        'link':      product_link,   # ← 새로 추가
    }

    # 6) 기존 키워드(<input> 초기값)
    existing_keywords = list(pivot.keys())

    return render(request, "rankings/ranking_monitoring_detail.html", {
        'product':           product,
        'existing_keywords': existing_keywords,
        'date_columns':      date_columns,
        'rank_list':         rank_list,
    })

@login_required
def add_monitoring_detail(request):
    # single_mid 얻기
    single_mid = (
        request.POST.get('selected_products', '').strip()
        or request.GET.get('product_id', '').strip()
    )
    product = get_object_or_404(Product, single_product_mid=single_mid)

    ranking_obj, created = RankingMonitoring.objects.get_or_create(
        product_id=product.single_product_mid,
        defaults={
            'product_url':  product.single_product_link,
            'product_name': product.name,
        }
    )
    if not created:
        ranking_obj.product_url  = product.single_product_link
        ranking_obj.product_name = product.name
        ranking_obj.save()

    if request.method == 'POST':
        # ① 삭제된 키워드만 지우기
        deleted_str = request.POST.get('deleted_keywords', '')
        deleted_list = [kw.strip() for kw in deleted_str.split(',') if kw.strip()]
        if deleted_list:
            KeywordRanking.objects.filter(
                ranking=ranking_obj,
                keyword__in=deleted_list
            ).delete()

        # ② 새로 추가할 키워드 처리
        monitoring_keywords = request.POST.get('monitoring_keywords', '')
        new_list = [kw.strip() for kw in monitoring_keywords.split(',') if kw.strip()]
        unique_new = list(dict.fromkeys(new_list))

        best_rank = 1001
        best_link = product.single_product_link

        for kw in unique_new:
            # 이미 있으면 건너뜀
            if ranking_obj.keywords.filter(keyword=kw).exists():
                continue

            # 단일/오리지널 순위 조회
            rank_single = get_naver_rank(kw, product.single_product_link)
            if product.original_link:
                rank_original = get_naver_rank(kw, product.original_link)
            else:
                rank_original = -1

            # final_rank 결정
            if rank_single == -1 and rank_original == -1:
                final_rank, is_orig = 1001, False
            elif rank_single == -1:
                final_rank, is_orig = rank_original, True
            elif rank_original == -1:
                final_rank, is_orig = rank_single, False
            else:
                if rank_original < rank_single:
                    final_rank, is_orig = rank_original, True
                else:
                    final_rank, is_orig = rank_single, False

            # best_link 업데이트
            if final_rank < best_rank:
                best_rank, best_link = final_rank, (
                    product.original_link if is_orig
                    else product.single_product_link
                )

            # DB에 생성
            KeywordRanking.objects.create(
                ranking=ranking_obj,
                keyword=kw,
                rank=final_rank,
                is_original_better=is_orig
            )

        # ③ best_link (& 필요시 메인키워드) 업데이트
        ranking_obj.product_url = best_link
        ranking_obj.save()

        messages.success(request, "모니터링 키워드가 업데이트되었습니다.")
        return redirect(f"/traffic/ranking-monitoring-detail/?product_id={single_mid}")

    # GET 은 그대로 상세 페이지로
    return redirect(f"/traffic/ranking-monitoring-detail/?product_id={single_mid}")


@login_required
def delete_monitoring_detail(request):
    if request.method == 'POST':
        pid = request.POST.get('product_id', '').strip()

        # 1) pid가 곧 RankingMonitoring.product_id라면 바로 삭제 시도
        try:
            ranking_obj = RankingMonitoring.objects.get(product_id=pid)
        except RankingMonitoring.DoesNotExist:
            # 2) pid가 Product PK일 수도 있으니 한 번 더 시도
            product = get_object_or_404(Product, pk=pid)
            ranking_obj = get_object_or_404(
                RankingMonitoring,
                product_id=product.single_product_mid
            )

        ranking_obj.delete()
        messages.success(request, "모니터링 상품이 삭제되었습니다.")

    # 삭제 완료 후 모니터링 목록 페이지로 이동
    return redirect('rankings:ranking_monitoring_list')




@login_required
def update_monitoring_search_detail(request):
    """
    단일 상품(product_id)에 묶여 있는 KeywordRanking 중
    중복 키워드를 제거하고 한 번씩만 검색량을 업데이트
    """
    product_id = request.POST.get('product_id') or request.GET.get('product_id')
    ranking_obj = get_object_or_404(RankingMonitoring, product_id=product_id)
    keyword_qs = ranking_obj.keywords.all()

    updated_count = 0
    processed = set()  # 이미 처리한 키워드 집합

    logger.info(f"시작: product_id={product_id}, total 키워드 건수={keyword_qs.count()}")
    for kwobj in keyword_qs:
        kw = kwobj.keyword
        if kw in processed:
            logger.debug(f"스킵(중복): keyword={kw}")  # 이미 한 번 처리한 키워드
            continue

        processed.add(kw)
        logger.debug(f"처리 중: keyword={kw}")

        try:
            df = get_rel_keywords([kw])
        except Exception:
            logger.exception(f"get_rel_keywords 호출 실패: keyword={kw}")
            continue

        if df is None or df.empty:
            logger.warning(f"검색량 데이터 없음: keyword={kw}")
            continue

        total_search = int(df.iloc[0].get('totalSearchCount', 0))
        logger.debug(f"조회결과: keyword={kw}, total_search_count={total_search}")

        # 해당 키워드를 가진 모든 레코드 한 번에 업데이트
        objs = KeywordRanking.objects.filter(ranking=ranking_obj, keyword=kw)
        # 변경 전후 비교해서 실제 바뀌는 레코드 수 집계
        changed = objs.exclude(search_volume=total_search)\
                      .update(search_volume=total_search)
        if changed:
            updated_count += changed
            logger.info(f"업데이트: keyword={kw}, 변경건수={changed}")

    logger.info(f"완료: 총 {updated_count}건 반영 (product_id={product_id})")
    messages.success(
        request,
        f"상품(ID={product_id}) 검색량 업데이트 완료! ({updated_count}건 반영)"
    )
    return redirect(
        f"{reverse('rankings:ranking_monitoring_detail_list')}?product_id={product_id}"
    )