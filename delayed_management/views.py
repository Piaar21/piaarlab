import io
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .forms import DelayedOrderUploadForm, OptionStoreMappingUploadForm  
from .models import DelayedOrder, ProductOptionMapping ,OptionStoreMapping, DelayedShipment,DelayedShipmentGroup,ExternalPlatformMapping,ExternalProductItem,ExternalProductOption, OptionMapping,OptionPlatformDetail,OutOfStockMapping,OutOfStock,OutOfStockCheck
from django.core.paginator import Paginator
from .api_clients import get_exchangeable_options,get_option_info_by_code,get_all_options_by_product_name,get_options_detail_by_codes,get_inventory_by_option_codes,fetch_coupang_seller_product_with_options,NAVER_ACCOUNTS,COUPANG_ACCOUNTS,fetch_coupang_all_seller_products,fetch_naver_products_with_details,get_coupang_seller_product_info
from .spreadsheet_utils import read_spreadsheet_data
from datetime import date, timedelta, datetime
from django.conf import settings
from .spreadsheet_utils import get_gspread_client
import uuid
import hmac
import hashlib
from decouple import config
import logging
import requests
import gspread
from django.http import HttpResponse
from django.db.models import Q
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.utils import timezone
import json
from django.views.decorators.http import require_POST
from openpyxl import Workbook



logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)  # 필요 시 레벨 지정

# ==========================
# 1) "업로드" + "리스트 업로드" 함수
# ==========================
def upload_delayed_orders(request):
    print("=== DEBUG: upload_delayed_orders 뷰 진입 ===")

    # (A) 엑셀 업로드 (임시 세션 저장)
    if request.method == 'POST' and 'upload_excel' in request.POST:
        form = DelayedOrderUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active

            temp_orders = []
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx == 0:
                    # 첫 행(헤더) 스킵
                    continue
                if not row or len(row) < 11:
                    # 컬럼 부족 등 스킵
                    continue

                option_code         = (row[0] or "").strip()
                customer_name       = (row[1] or "").strip()
                customer_contact    = (row[2] or "").strip()
                order_product_name  = (row[3] or "").strip()
                order_option_name   = (row[4] or "").strip()
                quantity            = (row[5] or "").strip()
                seller_product_name = (row[6] or "").strip()
                seller_option_name  = (row[7] or "").strip()
                order_number_1      = (row[8] or "").strip()
                order_number_2      = (row[9] or "").strip()  # 필요하다면 남김
                store_name_raw      = (row[10] or "").strip()

                # 옵션코드 / 주문번호1이 반드시 있어야 한다고 가정
                if not option_code:
                    print("=== DEBUG: 옵션코드 없음, 스킵 ===")
                    continue
                if not order_number_1:
                    print("=== DEBUG: 주문번호1 없음, 스킵 ===")
                    continue

                order_data = {
                    'option_code': option_code,
                    'customer_name': customer_name,
                    'customer_contact': customer_contact,
                    'order_product_name': order_product_name,
                    'order_option_name': order_option_name,
                    'quantity': quantity,
                    'seller_product_name': seller_product_name,
                    'seller_option_name': seller_option_name,
                    'order_number_1': order_number_1,
                    'order_number_2': order_number_2,  # 필요에 따라 사용
                    'store_name': store_name_raw,
                }
                temp_orders.append(order_data)

            # 세션에 임시저장
            request.session['delayed_orders_temp'] = temp_orders
            messages.success(request, f"{len(temp_orders)}건이 임시로 업로드되었습니다.")
            return redirect('upload_delayed_orders')
        else:
            messages.error(request, "파일 업로드 실패.")
            return redirect('upload_delayed_orders')

    # (B) 최종 DB 저장 + 자동처리
    elif request.method == 'POST' and 'finalize' in request.POST:
        temp_orders = request.session.get('delayed_orders_temp', [])
        if not temp_orders:
            messages.error(request, "임시 데이터가 없습니다.")
            return redirect('upload_delayed_orders')

        # 1) 기존 DB에서 'order_number_1' 목록 추출 → 중복 체크용
        existing_orders = set(
            DelayedShipment.objects.values_list('order_number_1', flat=True)
        )
        print(f"=== DEBUG: DB 내 이미 존재하는 주문번호1 개수: {len(existing_orders)}")

        success_count = 0
        fail_count = 0
        newly_created_ids = []

        # 업로드 실패 항목만 남길 임시배열
        remaining_temp_orders = []

        from collections import defaultdict
        group_token_map = defaultdict(lambda: str(uuid.uuid4()))

        # 2) temp_orders 순회, 중복 검사
        for od in temp_orders:
            order_num_1 = od.get('order_number_1', '')

            if not order_num_1:
                # 주문번호1 없으면 실패 처리
                fail_count += 1
                remaining_temp_orders.append(od)
                continue

            if order_num_1 in existing_orders:
                # 이미 DB에 존재하면 실패 처리
                fail_count += 1
                remaining_temp_orders.append(od)
                continue

            # 3) 중복 아니면 DB 저장
            #    그룹핑 기준: (주문번호1 + customer_name + customer_contact)
            group_key = (
                od['order_number_2'].strip(),
                od['customer_name'].strip(),
                od['customer_contact'].strip()
            )
            group_token = group_token_map[group_key]

            shipment = DelayedShipment.objects.create(
                option_code         = od['option_code'],
                customer_name       = od['customer_name'],
                customer_contact    = od['customer_contact'],
                order_product_name  = od['order_product_name'],
                order_option_name   = od['order_option_name'],
                quantity            = od['quantity'],
                seller_product_name = od['seller_product_name'],
                seller_option_name  = od['seller_option_name'],
                order_number_1      = od['order_number_1'],
                order_number_2      = od['order_number_2'],  # 필요 시 저장
                store_name          = od['store_name'],
                token               = group_token,
            )
            newly_created_ids.append(shipment.id)
            success_count += 1

            # 새로 등록된 order_number_1 추가
            existing_orders.add(order_num_1)

        # 4) 세션에는 실패 항목만 남긴다
        request.session['delayed_orders_temp'] = remaining_temp_orders

        print(f"=== DEBUG: 업로드 성공 {success_count}건 / 실패 {fail_count}건 ===")

        # (자동 처리) 옵션추출 + 스토어매핑 + 재입고동기화
        if newly_created_ids:
            extract_options_for_ids(newly_created_ids)
            store_mapping_for_ids(newly_created_ids)
            update_restock_from_sheet(request)  # <= "스프레드시트 기반 상태/날짜 동기화"

        # 5) 결과 메시지를 표시하고 템플릿을 다시 렌더
        messages.success(
            request,
            (
                f"총 {len(temp_orders)}건 중 저장 성공 {success_count}건, 실패 {fail_count}건\n"
                "(실패 항목은 테이블에 남아있음)"
            )
        )
        return render(request, 'delayed_management/upload_delayed_orders.html', {
            'form': DelayedOrderUploadForm(),
            'temp_orders': remaining_temp_orders,  # 실패한 항목
        })

    # (C) 임시 데이터 중 하나 삭제
    elif request.method == 'POST' and 'delete_item' in request.POST:
        temp_orders = request.session.get('delayed_orders_temp', [])
        index_to_delete = request.POST.get('delete_index')
        if index_to_delete is not None:
            try:
                idx = int(index_to_delete)
                if 0 <= idx < len(temp_orders):
                    deleted_item = temp_orders[idx]
                    del temp_orders[idx]
                    request.session['delayed_orders_temp'] = temp_orders
                    messages.success(request, f"{deleted_item['option_code']} 항목이 삭제되었습니다.")
                else:
                    messages.error(request, "잘못된 인덱스입니다.")
            except ValueError:
                messages.error(request, "잘못된 요청입니다.")
        return redirect('upload_delayed_orders')
    
    # **(C-2) 임시 데이터 전체 삭제** ← 새로 추가할 부분
    elif request.method == 'POST' and 'delete_all_items' in request.POST:
        request.session['delayed_orders_temp'] = []
        messages.success(request, "테이블 전체가 삭제되었습니다.")
        return redirect('upload_delayed_orders')


    # (D) GET: 업로드 폼 표시
    else:
        temp_orders = request.session.get('delayed_orders_temp', [])
        return render(request, 'delayed_management/upload_delayed_orders.html', {
            'form': DelayedOrderUploadForm(),
            'temp_orders': temp_orders,
        })




# ==========================
# 2) 하위 처리 함수들
# ==========================
def extract_options_for_ids(shipment_ids):
    qs = DelayedShipment.objects.filter(id__in=shipment_ids)
    if not qs.exists():
        return

    for shipment in qs:
        # (1) needed_qty 계산
        try:
            needed_qty = int(shipment.quantity or '1')
        except ValueError:
            needed_qty = 1

        # (2) get_exchangeable_options() → [{"optionName":..., "optionCode":..., "stock":...}, ...]
        exchange_list = get_exchangeable_options(shipment.option_code, needed_qty=needed_qty)

        if not exchange_list:
            # 결과가 아예 없으면 '상담원 문의'
            shipment.exchangeable_options = "상담원 문의"
            shipment.save()
            continue

        # (3) JSON 배열 → optionName 들만 뽑아서 콤마 문자열 생성
        #     예: ['애쉬그레이(R)', '샌드카키(L)', ...]
        name_list = [item["optionName"] for item in exchange_list]
        final_str = ",".join(name_list)  # "애쉬그레이(R),샌드카키(L),..."

        # (4) DB 저장
        shipment.exchangeable_options = final_str
        shipment.save()



def store_mapping_for_ids(shipment_ids):
    """
    새로 생성된 DelayedShipment 레코드들(ID 목록)에 대해
    store_name 매핑을 수행 (option_code → OptionStoreMapping).
    """
    qs = DelayedShipment.objects.filter(id__in=shipment_ids)
    if not qs.exists():
        return  # 아무것도 없으면 끝

    for shipment in qs:
        code = shipment.option_code
        try:
            osm = OptionStoreMapping.objects.get(option_code=code)
            shipment.store_name = osm.store_name
        except OptionStoreMapping.DoesNotExist:
            shipment.store_name = ""
        shipment.save()


def update_restock_from_sheet(request):
    service_account_file = settings.SERVICE_ACCOUNT_FILE
    spreadsheet_id = "xxxxxx..."
    client = get_gspread_client(service_account_file)
    sh = client.open_by_key(spreadsheet_id)

    tabs = ["구매된상품들", "배송중", "도착완료", "서류작성", "선적중"]
    today = date.today()
    total_updated = 0

    for tab_name in tabs:
        try:
            worksheet = sh.worksheet(tab_name)
        except gspread.WorksheetNotFound:
            print(f"=== 시트 '{tab_name}' 없음, 건너뜁니다.")
            continue

        data = worksheet.get_all_values()
        for idx, row in enumerate(data[1:], start=2):
            if not row or len(row) < 1:
                continue

            option_code = (row[0] or "").strip()
            if not option_code:
                continue

            # 스프레드시트 탭 → DB status
            status_code = map_status(tab_name)

            # ETA 계산 (min_days만 사용)
            min_days, max_days = ETA_RANGES.get(status_code, (0,0))
            calc_date = today + timedelta(days=min_days) if min_days else None

            # 기존처럼 get() → MultipleObjectsReturned 가능
            # → filter()로 전부 반영
            qs = DelayedShipment.objects.filter(option_code=option_code)
            if not qs.exists():
                continue

            # 상태+expected_restock_date (bulk update)
            qs.update(
                status=status_code,
                expected_restock_date=calc_date,
            )

            # restock_date가 None인 경우만 새로 설정
            for ship in qs:
                if ship.restock_date is None:
                    ship.restock_date = calc_date
                    ship.save()

            total_updated += qs.count()

    messages.success(request, f"{total_updated}건 동기화 완료")
    return redirect('restock_management')




def upload_store_mapping(request):
    # 1) 검색어(GET 파라미터) 확인
    search_option_code = request.GET.get('search_option_code', '').strip()
    found_store_name = None

    if search_option_code:
        try:
            # 검색 옵션코드가 존재하면 store_name 가져오기
            mapping = OptionStoreMapping.objects.get(option_code=search_option_code)
            found_store_name = mapping.store_name
        except OptionStoreMapping.DoesNotExist:
            # 검색 실패 시 None
            found_store_name = None

    if request.method == 'POST':
        form = OptionStoreMappingUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active

            row_count = 0
            codes_seen = set()
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx == 0:
                    continue
                if not row or len(row) < 2:
                    continue

                option_code = (row[0] or "").strip()
                store_name = (row[1] or "").strip()

                if not option_code:
                    continue
                if option_code in codes_seen:
                    continue
                codes_seen.add(option_code)

                OptionStoreMapping.objects.update_or_create(
                    option_code=option_code,
                    defaults={'store_name': store_name}
                )
                row_count += 1

            messages.success(request, f"{row_count}건의 스토어명 매핑 정보가 업로드/갱신되었습니다.")
            return redirect('upload_store_mapping')
    else:
        form = OptionStoreMappingUploadForm()

    # 페이지네이션
    mappings_qs = OptionStoreMapping.objects.all().order_by('option_code')
    paginator = Paginator(mappings_qs, 30)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # 페이지네이션 범위
    visible_pages = 5
    current = page_obj.number
    total = page_obj.paginator.num_pages

    half = visible_pages // 2
    start_page = current - half
    if start_page < 1:
        start_page = 1
    end_page = start_page + visible_pages - 1
    if end_page > total:
        end_page = total
        start_page = max(end_page - visible_pages + 1, 1)

    page_range_custom = range(start_page, end_page + 1)

    context = {
        'form': form,
        'mappings': page_obj,
        'page_obj': page_obj,
        'page_range_custom': page_range_custom,
        'search_option_code': search_option_code,
        'found_store_name': found_store_name,
    }
    return render(request, 'delayed_management/upload_store_mapping.html', context)


def delete_store_mapping(request, mapping_id):
    if request.method == 'POST':
        mapping = get_object_or_404(OptionStoreMapping, id=mapping_id)
        mapping.delete()
        messages.success(request, "삭제 완료")
    return redirect('upload_store_mapping')


def add_or_update_store_mapping(request):
    if request.method == 'POST':
        option_code = request.POST.get('option_code', '').strip()
        store_name = request.POST.get('store_name', '').strip()

        if option_code and store_name:
            # 이미 존재하면 업데이트
            OptionStoreMapping.objects.update_or_create(
                option_code=option_code,
                defaults={'store_name': store_name}
            )
            messages.success(request, f"옵션코드 {option_code} 스토어명이 업데이트/추가되었습니다.")
        else:
            messages.error(request, "옵션코드와 스토어명을 모두 입력해주세요.")
    return redirect('upload_store_mapping')



def delayed_shipment_list(request):
    qs = DelayedShipment.objects.filter(flow_status='pre_send').order_by('-created_at')

    today = date.today()
    for s in qs:
        min_days, max_days = ETA_RANGES.get(s.status, (0,0))

        # 고정 안내날짜 range
        if s.restock_date:
            s.range_start = s.restock_date + timedelta(days=min_days)
            s.range_end   = s.restock_date + timedelta(days=max_days)
        else:
            s.range_start = None
            s.range_end   = None

        # 동적 예상날짜 range
        if s.expected_restock_date:
            s.expected_start = s.expected_restock_date + timedelta(days=min_days)
            s.expected_end   = s.expected_restock_date + timedelta(days=max_days)
        else:
            s.expected_start = None
            s.expected_end   = None

    return render(request, 'delayed_management/delayed_shipment_list.html', {
        'shipments': qs
    })


########################################################
# 교환가능옵션 / 스토어매핑 / 선택삭제 / 문자발송 모두를 처리하는 뷰
########################################################

def change_exchangeable_options(request):
    """
    하나의 폼에서 4개 버튼(action)을 구분:
      - extract_options (옵션추출)
      - store_mapping (스토어 매핑)
      - delete_multiple (선택 삭제)
      - send_sms (문자 발송)
    """
    if request.method == 'POST':
        action = request.POST.get('action', '')

        # 1) 옵션추출 로직
        if action == 'extract_options':
            shipment_ids = request.POST.getlist('shipment_ids', [])
            if not shipment_ids:
                messages.error(request, "옵션추출할 항목이 선택되지 않았습니다.")
                return redirect('delayed_shipment_list')

            # id__in 으로 필터
            shipments = DelayedShipment.objects.filter(id__in=shipment_ids)
            updated = 0
            for s in shipments:
                # 교환가능 옵션 조회
                needed_qty = int(s.quantity or 1)
                exchange_list = get_exchangeable_options(s.option_code, needed_qty)
                # 문자열화
                s.exchangeable_options = ",".join(exchange_list) if exchange_list else "상담원 문의"
                s.save()
                updated += 1

            messages.success(request, f"{updated}건 옵션추출 완료.")
            return redirect('delayed_shipment_list')

        # 2) 스토어 매핑 로직
        elif action == 'store_mapping':
            shipment_ids = request.POST.getlist('shipment_ids', [])
            if not shipment_ids:
                messages.error(request, "스토어 매핑할 항목이 선택되지 않았습니다.")
                return redirect('delayed_shipment_list')

            shipments = DelayedShipment.objects.filter(id__in=shipment_ids)
            updated = 0
            for s in shipments:
                # s.option_code 로 OptionStoreMapping 조회
                try:
                    osm = OptionStoreMapping.objects.get(option_code=s.option_code)
                    s.store_name = osm.store_name
                    s.save()
                    updated += 1
                except OptionStoreMapping.DoesNotExist:
                    # 매핑이 없으면 넘어감 (필요시 처리)
                    pass

            messages.success(request, f"{updated}건 스토어 매핑 완료.")
            return redirect('delayed_shipment_list')

        # 3) 선택 삭제 로직
        elif action == 'delete_multiple':
            shipment_ids = request.POST.getlist('shipment_ids', [])
            if not shipment_ids:
                messages.error(request, "삭제할 항목이 선택되지 않았습니다.")
                return redirect('delayed_shipment_list')

            qs = DelayedShipment.objects.filter(id__in=shipment_ids)
            deleted_info = qs.delete()
            total_deleted = deleted_info[0]  # (삭제된 객체 수, {model: 개수})
            return redirect('delayed_shipment_list')

        # 4) 문자 발송 로직
        # 예시) change_exchangeable_options(request) 내부 혹은 send_sms 로직 등에서
        # 'store_name'이 없으면 발송 대상에서 제외하는 로직을 추가

        # (1) 문자 발송 로직
        elif action == 'send_sms':
            logger.debug("=== DEBUG: 문자 발송 로직 진입 ===")

            shipment_ids = request.POST.getlist('shipment_ids', [])
            if not shipment_ids:
                messages.error(request, "문자 발송할 항목이 선택되지 않았습니다.")
                return redirect('delayed_shipment_list')

            shipments = DelayedShipment.objects.filter(id__in=shipment_ids)
            if not shipments.exists():
                messages.error(request, "발송할 데이터가 없습니다.")
                return redirect('delayed_shipment_list')

            messages_list = []
            local_fail_ids = []

            for s in shipments:
                if not s.store_name:
                    local_fail_ids.append(s.id)
                    s.send_status = 'failed'  # 혹은 'FAIL'
                    s.save()
                    continue

                # pf_id → 알림톡 유무
                channel_name = map_store_to_channel(s.store_name)
                pf_id = get_pfId_by_channel(channel_name)
                template_id = get_templateId_by_channel(channel_name)

                # 임시로 DB에 'SENDING' (or just set send_status='pending' here?)
                if pf_id:
                    s.send_type = 'kakao'
                else:
                    s.send_type = 'sms'
                s.send_status = 'pending'  # 대기중
                s.save()

                # 치환변수 등
                min_days, max_days = ETA_RANGES.get(s.status, (0, 0))
                if s.restock_date:
                    start_d = s.restock_date + timedelta(days=min_days)
                    end_d   = s.restock_date + timedelta(days=max_days)
                    발송일자_str = f"{start_d.strftime('%Y.%m.%d')} ~ {end_d.strftime('%m.%d')}"
                else:
                    발송일자_str = "상담원문의"

                url_thx = f"piaarlab.store/delayed/customer-action?action=wait&token={s.token}"
                url_change = f"piaarlab.store/delayed/option-change?action=change&token={s.token}"

                variables = {
                    '#{고객명}': s.customer_name or "",
                    '#{상품명}': s.order_product_name or "",
                    '#{옵션명}': s.order_option_name or "",
                    '#{발송일자}': 발송일자_str,
                    '#{교환옵션명}': s.exchangeable_options or "",
                    '#{채널명}': s.store_name or "",
                    '#{url}': "example.com",
                    '#{url_thx}': url_thx,
                    '#{url_change}': url_change,
                }
                msg = {
                    "to": s.customer_contact or "",
                    "from": SOLAPI_SENDER_NUMBER,
                }
                if pf_id:
                    msg["kakaoOptions"] = {
                        "pfId": pf_id,
                        "templateId": template_id,
                        "variables": variables
                    }
                else:
                    msg["text"] = (
                        f"[문자 테스트]\n"
                        f"안녕하세요 {s.customer_name}님..."
                    )

                # customFields => shipment_id
                if "customFields" not in msg:
                    msg["customFields"] = {}
                msg["customFields"]["shipment_id"] = str(s.id)

                messages_list.append((s, msg))

            if not messages_list and not local_fail_ids:
                messages.warning(request, "발송할 대상이 없습니다.")
                return redirect('delayed_shipment_list')

            try:
                logger.debug("=== DEBUG: solapi_send_messages() 호출 전 ===")
                success_list, fail_list, group_id = solapi_send_messages(messages_list)
                logger.debug(f"=== DEBUG: solapi_send_messages() => success={success_list}, fail={fail_list}, group_id={group_id}")

                # local_fail_ids도 합침
                fail_list.extend(local_fail_ids)

                # groupId 저장
                all_sent_ids = [s.id for (s, _) in messages_list]
                DelayedShipment.objects.filter(id__in=all_sent_ids).update(solapi_group_id=group_id)

                # (A) 성공건 => send_status='pending'
                DelayedShipment.objects.filter(id__in=success_list).update(flow_status='sent', send_status='pending')
                # (B) 실패건 => send_status='failed'
                DelayedShipment.objects.filter(id__in=fail_list).update(flow_status='pre_send', send_status='failed')

                messages.success(request, f"문자 발송 완료 (groupId={group_id}). 대기중 상태로 표시됩니다.")

            except Exception as e:
                logger.exception("=== DEBUG: 문자 발송 오류 ===")
                messages.error(request, f"문자 발송 오류: {str(e)}")

            return redirect('delayed_shipment_list')


        
        # 5) 옵션매핑 전송 로직
        # (C) "옵션매핑 전송" 버튼 (send_option_mapping)
        elif action == 'send_option_mapping':
            shipment_ids = request.POST.getlist('shipment_ids', [])
            logger.debug(f"[send_option_mapping] shipment_ids={shipment_ids}")

            if not shipment_ids:
                messages.error(request, "전송할 항목이 선택되지 않았습니다.")
                return redirect('delayed_shipment_list')

 

            # 1) DelayedShipment 쿼리셋
            shipments = DelayedShipment.objects.filter(id__in=shipment_ids)
            if not shipments.exists():
                messages.error(request, "선택된 항목이 없거나 이미 삭제되었습니다.")
                return redirect('delayed_shipment_list')

            updated_count = 0

            # (A) DelayedShipment -> OutOfStock 업데이트 (복수건)
            for s in shipments:
                code = s.option_code
                product_name = s.seller_product_name
                option_name = s.seller_option_name
                expected_dt = s.expected_restock_date

                if not code:
                    logger.warning(f"[send_option_mapping] DelayedShipment(id={s.id}) 에 option_code가 없습니다.")
                    continue

                # OutOfStock 에서 해당 option_code를 가진 여러 레코드를 필터
                out_qs = OutOfStock.objects.filter(option_code=code)
                if not out_qs.exists():
                    logger.warning(f"[send_option_mapping] OutOfStock(option_code={code}) 없음.")
                    continue

                # 여러 건에 대해 일괄 업데이트 (loop or update() 사용 가능)
                # 여기서는 개별 필드 처리를 위해 loop 사용
                count_this_code = 0
                for out_item in out_qs:
                    out_item.seller_product_name = product_name
                    out_item.seller_option_name = option_name
                    out_item.expected_restock_date = expected_dt
                    out_item.status = 1  # 상태값=1 로 설정
                    out_item.save()
                    count_this_code += 1

                updated_count += count_this_code

            messages.success(request, f"{updated_count}건의 OutOfStock 업데이트 완료 (option_code 매핑).")

            # (B) 셀러툴 API 호출 -> seller_tool_stock 동기화
            #     1) 코드 목록만 추출
            code_list = shipments.values_list('option_code', flat=True)\
                                 .exclude(option_code__exact='')\
                                 .distinct()
            code_list = list(code_list)

            if not code_list:
                messages.warning(request, "옵션코드가 없어 셀러툴 재고 업데이트를 진행할 수 없습니다.")
                return redirect('out_of_stock_management')

            # 2) 셀러툴에서 재고 조회
            stock_map = get_inventory_by_option_codes(code_list)
            logger.debug(f"[send_option_mapping] stock_map={stock_map}")  # {option_code: 재고수량}

            # 3) OutOfStock 테이블에 반영
            #    동일 option_code가 여러 건이 있을 수 있으니 update() or loop
            for code in code_list:
                st_stock = stock_map.get(code, 0)
                # filter(...) 후 update()로 일괄적용
                OutOfStock.objects.filter(option_code=code).update(seller_tool_stock=st_stock)

            messages.success(request, f"셀러툴재고 {len(code_list)}건 동기화 완료")
            return redirect('out_of_stock_management')

    # 기본 리턴
    return redirect('delayed_shipment_list')



def delayed_exchange_options_view(request):

    # 예: GET 파라미터로 option_code를 받는다고 가정
    option_code = request.GET.get('option_code')

    if not option_code:
        return render(request, 'error.html', {'message': 'option_code 파라미터가 필요합니다.'})

    exchangeable_options = get_exchangeable_options(option_code)

    return render(request, 'delayed_management/exchangeable_options.html', {
        'option_code': option_code,
        'exchangeable_options': exchangeable_options
    })


# 상태별 ETA 범위 (예시)
ETA_RANGES = {
    'purchase':  (14, 21),   # 구매된상품들: 14~21일
    'shipping':  (10, 14),   # 배송중
    'arrived':   (7, 10),    # 도착완료
    'document':  (5, 7),     # 서류작성
    'loading':   (1, 4),     # 선적중
}

def restock_list(request):
    """
    재입고/배송지연 목록을 보여주는 예시 뷰.
    - restock_date(고정) + min/max
    - expected_restock_date(동적) + min/max
    """
    shipments = DelayedShipment.objects.all().order_by('-created_at')

    for s in shipments:
        min_days, max_days = ETA_RANGES.get(s.status, (0, 0))

        # 1) 고정 안내날짜 range
        if s.restock_date:
            start1 = s.restock_date + timedelta(days=min_days)
            end1   = s.restock_date + timedelta(days=max_days)
            # 템플릿에서 {{ s.range_start }} / {{ s.range_end }} 로 접근
            s.range_start = start1
            s.range_end   = end1
        else:
            s.range_start = None
            s.range_end   = None

        # 2) 동적 예상날짜 range
        if s.expected_restock_date:
            start2 = s.expected_restock_date + timedelta(days=min_days)
            end2   = s.expected_restock_date + timedelta(days=max_days)
            s.expected_start = start2
            s.expected_end   = end2
        else:
            s.expected_start = None
            s.expected_end   = None

    return render(request, 'delayed_management/restock_management.html', {
        'shipments': shipments,
    })




tabs = ["구매된상품들", "배송중", "도착완료", "서류작성", "선적중"]



def map_status(raw_status):
    if raw_status == "구매된상품들":
        return "purchase"
    elif raw_status == "배송중":
        return "shipping"
    elif raw_status == "도착완료":
        return "arrived"
    elif raw_status == "서류작성":
        return "document"
    elif raw_status == "선적중":
        return "loading"
    return "nopurchase"

def update_restock_from_sheet(request):
    service_account_file = settings.SERVICE_ACCOUNT_FILE
    spreadsheet_id = "1qQfo2Pp-odUuYwKmQXNPv1phzK6Gi2JtlJYaaz3T1F0"

    client = get_gspread_client(service_account_file)
    sh = client.open_by_key(spreadsheet_id)

    tabs = ["구매된상품들", "배송중", "도착완료", "서류작성", "선적중"]
    today = date.today()
    total_updated = 0

    for tab_name in tabs:
        try:
            worksheet = sh.worksheet(tab_name)
        except gspread.WorksheetNotFound:
            print(f"=== 시트 '{tab_name}' 없음, 건너뜁니다.")
            continue

        data = worksheet.get_all_values()
        for idx, row in enumerate(data[1:], start=2):
            # row[0]에 option_code가 있다고 가정
            if not row or len(row) < 1:
                continue

            option_code = (row[0] or "").strip()
            if not option_code:
                continue

            # 스프레드시트 탭 이름으로 상태 매핑
            status_code = map_status(tab_name)

            # ETA 계산 (min_days만 사용)
            min_days, max_days = ETA_RANGES.get(status_code, (0, 0))
            calc_date = today + timedelta(days=min_days) if min_days else None

            # ▼ 변경 후 코드: filter(...) 사용 (get(...) 대신)
            qs = DelayedShipment.objects.filter(option_code=option_code)
            if not qs.exists():
                # 해당 option_code가 없으면 스킵
                continue

            # 1) status, expected_restock_date bulk update
            qs.update(
                status=status_code,
                expected_restock_date=calc_date
            )

            # 2) restock_date는 None인 경우에만 세팅
            for ship in qs:
                if ship.restock_date is None:
                    ship.restock_date = calc_date
                    ship.save()

            total_updated += qs.count()

    messages.success(
        request,
        f"{total_updated}건 동기화 완료 (restock_date는 최초 1회만 설정, expected_restock_date는 매번 갱신)."
    )
    return redirect('restock_management')






# Solapi API 환경변수
SOLAPI_API_KEY = config('SOLAPI_API_KEY', default=None)
SOLAPI_API_SECRET = config('SOLAPI_API_SECRET', default=None)
SOLAPI_SENDER_NUMBER = config('SOLAPI_SENDER_NUMBER', default='')  # 발신번호


def send_message_list(request):
    # 기본 QS
    qs = DelayedShipment.objects.filter(flow_status='sent')
    
    # 검색 파라미터가 있으면 필터링
    recipient_name = request.GET.get('recipient_name')
    if recipient_name:
        qs = qs.filter(customer_name__icontains=recipient_name)

    # 최신순 정렬
    qs = qs.order_by('-created_at')

    today = date.today()
    for s in qs:
        min_days, max_days = ETA_RANGES.get(s.status, (0,0))

        # 고정 안내날짜 range
        if s.restock_date:
            s.range_start = s.restock_date + timedelta(days=min_days)
            s.range_end   = s.restock_date + timedelta(days=max_days)
        else:
            s.range_start = None
            s.range_end   = None

        # 동적 예상날짜 range
        if s.expected_restock_date:
            s.expected_start = s.expected_restock_date + timedelta(days=min_days)
            s.expected_end   = s.expected_restock_date + timedelta(days=max_days)
        else:
            s.expected_start = None
            s.expected_end   = None

    return render(request, 'delayed_management/send_message_list.html', {
        'shipments': qs
    })

def send_message_process(request):
    if request.method == 'POST':
        action = request.POST.get('action', '')
        logger.debug(f"=== DEBUG: send_message_process() action={action}")

        shipment_ids = request.POST.getlist('shipment_ids', [])
        logger.debug(f"=== DEBUG: shipment_ids={shipment_ids}")

        # (A) 선택 삭제(delete_multiple)
        if action == 'delete_multiple':
            if not shipment_ids:
                messages.error(request, "삭제할 항목이 선택되지 않았습니다.")
                return redirect('send_message_list')

            qs = DelayedShipment.objects.filter(id__in=shipment_ids)
            deleted_info = qs.delete()
            total_deleted = deleted_info[0]  # (삭제된 객체 수, { 'app.Model': count })
            return redirect('send_message_list')
        
        elif action == 'check_status':
            logger.debug("=== DEBUG: 전송상태 확인(check_status) 진입 ===")
            shipment_ids = request.POST.getlist('shipment_ids', [])
            if not shipment_ids:
                messages.error(request, "전송상태 확인할 항목이 선택되지 않았습니다.")
                return redirect('send_message_list')

            shipments = DelayedShipment.objects.filter(id__in=shipment_ids)
            if not shipments.exists():
                messages.error(request, "해당 데이터가 존재하지 않습니다.")
                return redirect('send_message_list')

            group_ids = set(
                shipments.values_list('solapi_group_id', flat=True)
                .exclude(solapi_group_id__isnull=True)
                .exclude(solapi_group_id__exact='')
            )
            logger.debug(f"=== DEBUG: group_ids={group_ids}")

            if not group_ids:
                messages.warning(request, "해당 항목에 groupId가 없습니다 (미발송 상태이거나 오류).")
                return redirect('send_message_list')

            for gid in group_ids:
                if not gid:
                    continue
                try:
                    logger.debug(f"=== DEBUG: check_solapi_group_status_and_update(gid={gid}) 호출")
                    check_solapi_group_status_and_update(gid)
                except Exception as e:
                    logger.exception(f"=== DEBUG: 상태조회 groupId={gid} 에러 => {e}")

            messages.success(request, f"총 {len(group_ids)}개 groupId 상태조회 완료.")
            return redirect('send_message_list')


        # (B) 문자 발송(send_sms)
        elif action == 'resend_sms':
            logger.debug("=== DEBUG: 문자 재발송 로직(resend_sms) 진입 ===")

            # 1) 체크된 항목이 있는지 확인
            if not shipment_ids:
                messages.error(request, "재발송할 항목이 선택되지 않았습니다.")
                return redirect('send_message_list')

            # 2) DB에서 Shipment 목록 조회
            shipments = DelayedShipment.objects.filter(id__in=shipment_ids)
            if not shipments.exists():
                messages.error(request, "발송할 데이터가 없습니다.")
                logger.debug("=== DEBUG: DB 조회결과 없음 ===")
                return redirect('send_message_list')

            # 3) 메시지 목록 및 로컬 실패 리스트
            messages_list = []
            local_fail_ids = []

            for s in shipments:
                # (A) 스토어명 없는 경우 → 실패 처리 후 continue
                if not s.store_name:
                    logger.debug(f"=== DEBUG: store_name 없음, shipment_id={s.id}")
                    local_fail_ids.append(s.id)
                    s.send_status = 'FAIL'
                    s.save()
                    continue

                # (B) store_name → pf_id/template_id
                channel_name = map_store_to_channel(s.store_name)
                pf_id = get_pfId_by_channel(channel_name)
                template_id = get_templateId_by_channel(channel_name)

                # (C) 발송유형, 상태 임시 설정
                if pf_id:
                    s.send_type = 'KAKAO'
                else:
                    s.send_type = 'SMS'
                s.send_status = 'SENDING'
                s.save()

                # (D) 재입고 안내날짜 계산
                min_days, max_days = ETA_RANGES.get(s.status, (0, 0))
                if s.restock_date:
                    start_d = s.restock_date + timedelta(days=min_days)
                    end_d   = s.restock_date + timedelta(days=max_days)
                    발송일자_str = f"{start_d.strftime('%Y.%m.%d')} ~ {end_d.strftime('%m.%d')}"
                else:
                    발송일자_str = "상담원문의"

                # (E) URL 치환변수
                url_thx = f"piaarlab.store/delayed/customer-action?action=wait&token={s.token}"
                url_change = f"piaarlab.store/delayed/option-change?action=change&token={s.token}"

                # (F) 알림톡/문자용 치환 변수
                variables = {
                    '#{고객명}': s.customer_name or "",
                    '#{상품명}': s.order_product_name or "",
                    '#{옵션명}': s.order_option_name or "",
                    '#{발송일자}': 발송일자_str,
                    '#{교환옵션명}': s.exchangeable_options or "",
                    '#{채널명}': s.store_name or "",
                    '#{url}': "example.com",
                    '#{url_thx}': url_thx,
                    '#{url_change}': url_change,
                }

                # (G) Solapi 메시지 오브젝트
                msg_obj = {
                    "to": s.customer_contact or "",
                    "from": SOLAPI_SENDER_NUMBER,   # 예: "023389465"
                }

                if pf_id:
                    # (알림톡) kakaoOptions
                    msg_obj["kakaoOptions"] = {
                        "pfId": pf_id,
                        "templateId": template_id,
                        "variables": variables
                    }
                else:
                    # (SMS) text
                    # 여기서 [재발송] 문구만 추가
                    msg_obj["text"] = (
                        f"[재발송]\n"
                        f"안녕하세요 {s.customer_name or ''}님,\n"
                        f"{s.order_product_name or ''} 제품이 품절되어 연락드렸습니다.\n"
                        f"구매하신 스토어로 연락주시면 상품 변경 및 취소처리를 "
                        f"빠르게 도와드릴 수 있도록 하겠습니다. 감사합니다."
                    )

                # (H) messages_list 에 (shipment, msg) 형태로 추가
                messages_list.append((s, msg_obj))

            # 4) 만약 실제 발송할 대상이 전혀 없다면(스토어명 없는 항목만 골랐거나 등)
            if not messages_list and not local_fail_ids:
                logger.debug("=== DEBUG: 재발송 대상이 전혀 없음 ===")
                messages.warning(request, "재발송 대상이 없습니다.")
                return redirect('send_message_list')

            # 5) Solapi 발송
            try:
                logger.debug("=== DEBUG: solapi_send_messages() 호출 (resend_sms) ===")
                success_list, fail_list, group_id = solapi_send_messages(messages_list)

                # (A) store_name 없는 항목 등 미발송 → fail_list 합치기
                fail_list.extend(local_fail_ids)

                # (B) DB 업데이트 (성공 / 실패)
                DelayedShipment.objects.filter(id__in=success_list).update(
                    flow_status='sent',
                    send_status='SUCCESS'
                )
                DelayedShipment.objects.filter(id__in=fail_list).update(
                    flow_status='pre_send',
                    send_status='FAIL'
                )

                messages.success(request, f"재발송 완료 (groupId={group_id})")
                logger.debug("=== DEBUG: 문자 재발송 완료 ===")

            except Exception as e:
                logger.exception("=== DEBUG: 문자 재발송 오류 ===")
                messages.error(request, f"문자 재발송 오류: {str(e)}")

            # 6) 끝나면 목록 리다이렉트
            return redirect('send_message_list')


    # 그 외(GET 등) 잘못된 요청이면 리스트 페이지로
    return redirect('send_message_list')



def solapi_send_messages(messages_list):
    """
    HTTP 200 -> success_list
    else -> fail_list
    """
    success_list = []
    fail_list = []
    group_id = None

    body = {"messages": []}
    for s, msg_obj in messages_list:
        # 이미 customFields / shipment_id 를 세팅했으므로 OK
        body["messages"].append(msg_obj)


    now_iso = datetime.utcnow().replace(tzinfo=timezone.utc).isoformat(timespec='seconds')
    salt = str(uuid.uuid4())
    signature = make_solapi_signature(now_iso, salt, SOLAPI_API_SECRET)
    auth_header = f"HMAC-SHA256 apiKey={SOLAPI_API_KEY}, date={now_iso}, salt={salt}, signature={signature}"

    headers = {
        "Content-Type": "application/json",
        "Authorization": auth_header
    }
    url = "https://api.solapi.com/messages/v4/send-many"
    res = requests.post(url, json=body, headers=headers, timeout=10)
    if res.status_code == 200:
        res_json = res.json()
        group_id = res_json.get("groupId")
        for s, _ in messages_list:
            s.send_status = 'SUCCESS'
            s.message_sent_at = timezone.now()
            s.save()
            success_list.append(s.id)
    else:
        for s, _ in messages_list:
            s.send_type = 'NONE'
            s.send_status = 'FAIL'
            s.save()
            fail_list.append(s.id)

    return success_list, fail_list, group_id


def check_solapi_group_status_and_update(group_id):
    """
    Solapi에서 최종 상태를 가져와 (알림톡 성공/문자 성공/실패) DB에 반영.
    여기서는 /messages/v4/groups/{groupId}?withMessageList=true 이용.
    """
    logger.debug("=== DEBUG: check_solapi_group_status_and_update() start ===")
    logger.debug(f"=== DEBUG: group_id={group_id}")

    # 1) /messages/v4/groups/{groupId}?withMessageList=true
    url = f"https://api.solapi.com/messages/v4/groups/{group_id}?withMessageList=true"
    logger.debug(f"=== DEBUG: groups endpoint URL => {url}")

    # 인증
    now_iso = datetime.utcnow().replace(tzinfo=timezone.utc).isoformat(timespec='seconds')
    salt = str(uuid.uuid4())
    signature = make_solapi_signature(now_iso, salt, SOLAPI_API_SECRET)
    auth_header = f"HMAC-SHA256 apiKey={SOLAPI_API_KEY}, date={now_iso}, salt={salt}, signature={signature}"

    headers = {
        "Content-Type": "application/json",
        "Authorization": auth_header
    }

    res = requests.get(url, headers=headers, timeout=10)
    logger.debug(f"=== DEBUG: check_solapi_group_status code={res.status_code}, body[:500]={res.text[:500]}")
    if res.status_code != 200:
        logger.warning(f"=== DEBUG: groupId={group_id} 상태조회 실패 => {res.text}")
        return

    data = res.json()
    # messageList 안에 각 메시지별 상태가 들어 있을 수 있음
    message_list = data.get("messageList", [])
    logger.debug(f"=== DEBUG: messageList size={len(message_list)}")

    if not message_list:
        logger.debug("=== DEBUG: messageList 비어있음 => 처리 불가 (아직 SENDING 중일 가능성)")
        return

    # 2) 메시지별 처리
    for i, msg_item in enumerate(message_list):
        logger.debug(f"=== DEBUG: [#{i}] msg_item => {msg_item}")

        cf = msg_item.get("customFields", {})
        shipment_id_str = cf.get("shipment_id")
        if not shipment_id_str:
            logger.debug("=== DEBUG: shipment_id 없음 => 건너뜀")
            continue

        try:
            shipment_id = int(shipment_id_str)
            s = DelayedShipment.objects.get(id=shipment_id)
        except (ValueError, DelayedShipment.DoesNotExist):
            logger.debug(f"=== DEBUG: invalid shipment_id={shipment_id_str}")
            continue

        # 알림톡 성공 여부
        # Solapi 문서에 따라 "statusCode"가 "2000"이면 성공
        main_code = msg_item.get("statusCode")  # 알림톡 상태
        # 대체발송 문자
        failover = msg_item.get("failover", {})
        failover_code = failover.get("statusCode")  # "2000"이면 문자 성공

        logger.debug(f"=== DEBUG: shipment<{s.id}> main_code={main_code}, failover_code={failover_code}")

        # 실제 로직
        if main_code == "2000":
            # 알림톡 성공
            s.send_type = 'KAKAO'
            s.send_status = 'SUCCESS'
            logger.debug(f"   => 알림톡 성공 처리 (shipment_id={s.id})")
        elif failover_code == "2000":
            # 알림톡 실패, 문자 성공
            s.send_type = 'SMS'
            s.send_status = 'SUCCESS'
            logger.debug(f"   => 문자 성공 처리 (shipment_id={s.id})")
        else:
            # 둘 다 실패
            s.send_type = 'SMS'
            s.send_status = 'FAIL'
            logger.debug(f"   => 둘 다 실패 처리 (shipment_id={s.id})")

        s.save()

    logger.debug("=== DEBUG: check_solapi_group_status_and_update() end ===")


@csrf_exempt
def solapi_webhook_message(request):
    """
    Solapi "메시지 리포트" 웹훅
    예: [ { "messageId":"...", "type":"ATA", "statusCode":"4000", "customFields":{"shipment_id":"298"}, ... }, ... ]
    """
    if request.method != 'POST':
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    try:
        body_str = request.body.decode('utf-8')
        data_list = json.loads(body_str)  # [ {...}, {...} ]
    except Exception as e:
        logger.debug(f"=== WEBHOOK ERROR: invalid JSON => {str(e)}")
        return JsonResponse({"error": f"invalid json: {str(e)}"}, status=400)

    # data_list 루프
    for item in data_list:
        # 1) 로그 출력
        logger.debug("=== WEBHOOK ITEM START ===")
        logger.debug(f"messageId={item.get('messageId')}")
        logger.debug(f"groupId={item.get('groupId')}")
        logger.debug(f"type={item.get('type')}  statusCode={item.get('statusCode')}")
        logger.debug(f"customFields={item.get('customFields', {})}")
        logger.debug("=== WEBHOOK ITEM END ===")

        # 2) 필요한 필드 파싱
        message_id  = item.get("messageId")
        group_id    = item.get("groupId")
        type_       = item.get("type")         # "ATA" or "LMS" etc.
        status_code = item.get("statusCode")   # "4000" 등
        custom_fields = item.get("customFields", {})

        # 3) DB 식별
        shipment_id_str = custom_fields.get("shipment_id")
        if not shipment_id_str:
            logger.debug("No shipment_id => skip this item")
            continue

        try:
            shipment_id = int(shipment_id_str)
            s = DelayedShipment.objects.get(id=shipment_id)
        except (ValueError, DelayedShipment.DoesNotExist):
            logger.debug(f"Invalid or not-found shipment_id={shipment_id_str}")
            continue

        # 4) 분기 로직
        if type_ == "ATA" and status_code == "4000":
            # 알림톡(ATA) 성공
            s.send_type = "KAKAO"
            s.send_status = "SUCCESS"
            logger.debug(f"shipment<{s.id}> => 알림톡 SUCCESS")
        elif type_ == "ATA" and status_code == "1026":
            # 문자(LMS) 성공
            s.send_type = "dupli"
            s.send_status = "FAIL"
            logger.debug(f"shipment<{s.id}> => 알림톡 FAIL (duplication)")
        elif type_ == "LMS" and status_code == "4000":
            # 문자(LMS) 성공
            s.send_type = "SMS"
            s.send_status = "SUCCESS"
            logger.debug(f"shipment<{s.id}> => 문자(LMS) SUCCESS")
        else:
            # 나머지 -> 발송 실패
            s.send_type = "NONE"
            s.send_status = "FAIL"
            logger.debug(f"shipment<{s.id}> => FAIL (type={type_}, code={status_code})")

        s.save()

    # (5) 반드시 200 응답
    return JsonResponse({"message":"ok"}, status=200)

##########################################
# 2) 그룹 단위로 카톡을 보내는 예시 함수
##########################################
def send_kakao_for_group(group_id):
    """
    DelayedShipmentGroup 하나의 id를 받아,
    그 그룹 내에 포함된 DelayedShipment 여러 개를 한꺼번에 안내.
    """
    group = get_object_or_404(DelayedShipmentGroup, id=group_id)

    # 만약 group.token 이 없으면 새로 발행
    if not group.token:
        group.token = str(uuid.uuid4())
        group.save()

    # 그룹 대표 연락처
    contact = group.contact

    # 그룹에 속한 주문
    shipments = group.shipments.all()
    # 실제론 여러 주문에 대해 안내문을 만들거나, 첫 번째 주문만 참조할 수도 있음
    # 여기서는 첫 Shipment만 예시
    first_s = shipments.first()

    if not first_s:
        raise ValueError("이 그룹에는 DelayedShipment가 하나도 없습니다.")

    # store_name(혹은 brand)에 따라 pfId, templateId 매핑
    store_name = first_s.store_name or "기본값"
    pf_id = get_pfId_by_channel(store_name)        # 프로젝트에 맞게 구현
    template_id = get_templateId_by_channel(store_name)

    # URL: 그룹 단위로 “기다리기 / 옵션변경”을 처리할 엔드포인트
    # 예: /delayed/customer-group-action?token=...
    domain = getattr(settings, 'MY_SERVER_DOMAIN', 'https://34.64.123.206.com',"https://piaarlab.store/")
    confirm_url = f"{domain}/delayed/customer-group-action?token={group.token}"

    # 치환변수 구성
    # 만약 여러 상품명을 한꺼번에 보여주려면
    # product_summary = ", ".join([s.order_product_name for s in shipments])
    product_summary = first_s.order_product_name or "(상품명)"

    variables = {
        "#{url}": confirm_url,
        "#{상품명}": product_summary,
        "#{고객명}": first_s.customer_name or "고객님"
    }

    msg = {
        "to": contact,
        "from": getattr(settings, 'SOLAPI_SENDER_NUMBER', ''),  # 발신번호
        "kakaoOptions": {
            "pfId": pf_id,
            "templateId": template_id,
            "variables": variables
        }
    }

    # 전송
    result = solapi_send_messages([msg])  # 여러 건이면 messages 배열
    # 필요시 result를 반환 / DB 저장 처리
    return result


def make_solapi_signature(now_iso, salt, secret_key):
    # HMAC-SHA256(now_iso+salt, secret_key)
    data = (now_iso + salt).encode('utf-8')
    secret = secret_key.encode('utf-8')
    sig = hmac.new(secret, data, hashlib.sha256).hexdigest()
    return sig

def map_store_to_channel(store_name: str) -> str:
    """
    store_name → channel_name
    예) "니뜰리히", "수비다", "노는 개 최고양", "아르빙" 등
    """

    return store_name or ""

def get_pfId_by_channel(channel_name):
    """
    채널명 → pfId
    """
    pf_map = {
        "니뜰리히": "KA01PF241004055851440a91cCwGs47a",
        "수비다": "KA01PF241016104213982rcJgXgeHWd8",
        "수비다 SUBIDA": "KA01PF241016104213982rcJgXgeHWd8",
        "노는 개 최고양": "KA01PF241016104250071z8FktOivA0r",
        "노는개최고양": "KA01PF241016104250071z8FktOivA0r",
        "아르빙": "KA01PF241016104450781C2ssOn2lCCM",
    }
    return pf_map.get(channel_name, "")

def get_templateId_by_channel(channel_name):
    """
    채널명 → templateId
    """
    tmpl_map = {
        "니뜰리히": "KA01TP241230063811612F7QtAXfIYxu",
        "수비다": "KA01TP241230063811612F7QtAXfIYxu",
        "수비다 SUBIDA": "KA01TP241230063811612F7QtAXfIYxu",
        "노는 개 최고양": "KA01TP241230063811612F7QtAXfIYxu",
        "노는개최고양": "KA01TP241230063811612F7QtAXfIYxu",
        "아르빙": "KA01TP241230063811612F7QtAXfIYxu",
    }
    return tmpl_map.get(channel_name, "")

def confirm_token_view(request):
    token = request.GET.get('token', '')
    if not token:
        return HttpResponse("잘못된 접근", status=400)

    try:
        shipment = DelayedShipment.objects.get(token=token)
        shipment.confirmed = True
        shipment.save()
        return HttpResponse("확인 처리되었습니다. 감사합니다.")
    except DelayedShipment.DoesNotExist:
        return HttpResponse("유효하지 않은 토큰", status=404)
    

def customer_action_view(request):
    action = request.GET.get('action', '')
    token = request.GET.get('token', '')
    logger.debug(f"=== DEBUG: action={action}, token={token}")

    if not token:
        return HttpResponse("유효하지 않은 요청: token 없음", status=400)

    shipment = DelayedShipment.objects.filter(token=token).first()
    logger.debug(f"=== DEBUG: found shipment={shipment} for token={token}")
    if not shipment:
        return HttpResponse("유효하지 않은 토큰", status=404)

    # 이미 waiting/changed_option 이 있다면 더 이상 변경 불가
    if shipment.waiting or shipment.changed_option:
        logger.debug(f"=== DEBUG: shipment already waiting={shipment.waiting} or changed_option={shipment.changed_option}")
        return HttpResponse("이미 처리된 내역이 존재합니다. 더 이상 변경할 수 없습니다.")

    if action == 'wait':
        shipment.waiting = True
        shipment.confirmed = True
        # 발송흐름 상태도 '확인완료'로
        logger.debug(f"=== DEBUG: Before: flow_status={shipment.flow_status}, waiting={shipment.waiting}")
        shipment.flow_status = 'confirmed'
        shipment.save()
        return redirect(f'/delayed/thank-you/?token={token}')

    elif action == 'change':
        # TODO: flow_status='confirmed' 는 옵션변경 후 최종 저장(옵션변경 완료 시점)에 세팅
        return redirect(f"/delayed/option-change/?token={token}")

    else:
        return HttpResponse("잘못된 action", status=400)

    
    
def thank_you_view(request):
    
    """
    "기다리기" 버튼을 클릭한 뒤 띄울 페이지.
    간단한 감사 메시지를 렌더링 + 제품도착예상 날짜 표시.
    """
    token = request.GET.get('token', '')
    logger.debug("thank_you_view() token = %s", token)
    shipments = DelayedShipment.objects.filter(token=token)
    if not shipments.exists():
        return HttpResponse("유효하지 않은 토큰", status=404)

    # 각 Shipment마다 예상날짜 계산
    for s in shipments:
        min_days, max_days = ETA_RANGES.get(s.status, (0, 0))
        if s.expected_restock_date:
            s.expected_start = s.expected_restock_date + timedelta(days=min_days)
            s.expected_end   = s.expected_restock_date + timedelta(days=max_days)
        else:
            s.expected_start = None
            s.expected_end   = None

    return render(request, 'delayed_management/thank_you.html', {
        'shipments': shipments,
    }) 

def option_change_view(request):
    token = request.GET.get('token', '')
    logger.debug("option_change_view() token = %s", token)
    shipments = DelayedShipment.objects.filter(token=token)
    if not shipments.exists():
        return HttpResponse("유효하지 않은 토큰", status=404)

    # 1) ETA_RANGES 등 상태별 날짜 계산
    for s in shipments:
        min_days, max_days = ETA_RANGES.get(s.status, (0, 0))

        # (1) 고정 안내날짜
        if s.restock_date:
            s.range_start = s.restock_date + timedelta(days=min_days)
            s.range_end   = s.restock_date + timedelta(days=max_days)
        else:
            s.range_start = None
            s.range_end   = None

        # (2) 동적 예상날짜
        if s.expected_restock_date:
            s.expected_start = s.expected_restock_date + timedelta(days=min_days)
            s.expected_end   = s.expected_restock_date + timedelta(days=max_days)
        else:
            s.expected_start = None
            s.expected_end   = None

    shipments_data = []
    for s in shipments:
        # (A) DB에 있는 '교환옵션' 문자열 → splitted
        #    예: "애쉬그레이(R),샌드카키(L),오트밀베이지(R)"
        if s.exchangeable_options:
            splitted = [opt.strip() for opt in s.exchangeable_options.split(',') if opt.strip()]
        else:
            splitted = []

        # (B) 실제 API 통해 **optionCode** 포함한 전체 dict list 가져오기
        try:
            needed_qty = int(s.quantity or 1)
        except ValueError:
            needed_qty = 1

        # 예: [{optionName:'애쉬그레이(R)', optionCode:'4qom1...', stock:2}, ...] 
        full_list = get_exchangeable_options(s.option_code, needed_qty)

        # (C) splitted(=DB 저장된 옵션명 목록)와 일치하는 옵션만 선별
        #     → 최종 real_options
        real_options = []
        for opt_info in full_list:  
            # opt_info 예: {"optionName":"애쉬그레이(R)","optionCode":"..."}
            if opt_info["optionName"] in splitted:
                real_options.append(opt_info)

        # (주의) 만약 splitted에 없는 **재고가능 옵션**도 다 보여주길 원한다면,
        #       그냥 full_list 전체를 사용해도 좋습니다.
        #       여기서는 "DB에 저장된 목록"만 제한적으로 뿌린다고 가정.

        shipments_data.append({
            "shipment": s,
            "exchangeable_options": real_options,  # => 템플릿에선 opt.optionName / opt.optionCode 사용
        })

    return render(request, 'delayed_management/change_option.html', {
        "shipments_data": shipments_data,
        "token": token,
    })


def option_change_process(request):
    if request.method != 'POST':
        return HttpResponse("잘못된 접근입니다.", status=405)

    token = request.POST.get('token', '')
    shipments = DelayedShipment.objects.filter(token=token)
    if not shipments.exists():
        return HttpResponse("유효하지 않은 토큰", status=404)

    # 이미 waiting=True or changed_option 이 있으면 수정 불가
    any_already = shipments.filter(Q(waiting=True) | ~Q(changed_option='')).exists()
    logger.debug(f"=== DEBUG: [option_change_process] token={token}")
    logger.debug(f"=== DEBUG: shipments.count={shipments.count()}")
    logger.debug(f"=== DEBUG: any_already={any_already}")
    if any_already:
        return HttpResponse("이미 처리된 내역이 존재합니다.", status=400)

    for s in shipments:
        new_opt_key  = f"new_option_for_{s.id}"
        new_code_key = f"new_option_code_for_{s.id}"
        new_qty_key  = f"new_qty_for_{s.id}"  # 수량 변경 고려 시

        new_option     = request.POST.get(new_opt_key, '')
        new_option_code= request.POST.get(new_code_key, '')
        new_qty_str    = request.POST.get(new_qty_key, '1')

        try:
            new_qty = int(new_qty_str)
        except ValueError:
            new_qty = 1

        logger.debug(f"=== DEBUG: shipment.id={s.id}, new_option={new_option}, new_option_code={new_option_code}, new_qty_str={new_qty_str}")

        # 선택된 옵션이 없거나 "옵션을 선택하세요" 그대로면 처리
        stripped_option = new_option.strip()
        if not stripped_option:
            logger.debug("=== DEBUG: 선택된 옵션이 없거나 '옵션을 선택하세요' 그대로임")
            new_option = "변경 가능한 옵션이 없습니다. 상담원에게 문의해주세요."
            new_option_code = ""

        logger.debug(f"=== DEBUG: [SAVE] shipment.id={s.id} => changed_option={new_option}, changed_option_code={new_option_code}")

        s.changed_option = new_option
        s.changed_option_code = new_option_code
        s.quantity = new_qty
        s.confirmed = True
        s.waiting   = False
        s.flow_status = 'confirmed'
        s.save()

    logger.debug("=== DEBUG: 모든 shipment에 대해 옵션변경 완료 → option_change_done redirect")
    return redirect('option_change_done')





def option_change_done(request):
    return render(request, 'delayed_management/option_change_done.html')

def customer_group_action_view(request):
    token = request.GET.get('token', '')
    if not token:
        return HttpResponse("token이 없습니다.", status=400)

    group = DelayedShipmentGroup.objects.filter(token=token).first()
    if not group:
        return HttpResponse("유효하지 않은 token.", status=404)

    # group 내 모든 Shipment
    shipments = group.shipments.all()
    # 템플릿에서 "옵션별로 기다리기 or 변경"을 표시하도록
    return render(request, 'delayed_management/customer_group_action.html', {
        'group': group,
        'shipments': shipments
    })





def process_confirmed_shipments(request):
    """
    confirmed_list 템플릿에서 POST로 전달된 shipment_ids 에 대해
    - action = "complete_process" → 출고완료(flow_status='shipped' 등)
    - action = "delete_multiple"  → 선택 삭제
    """
    if request.method == 'POST':
        action = request.POST.get('action', '')
        shipment_ids = request.POST.getlist('shipment_ids', [])

        if not shipment_ids:
            messages.warning(request, "선택된 항목이 없습니다.")
            return redirect('confirmed_list')

        qs = DelayedShipment.objects.filter(id__in=shipment_ids)

        if action == 'complete_process':
            # 출고완료 버튼을 눌렀을 때
            count = qs.update(flow_status='shipped')  # 필요에 따라 send_status 등 다른 필드도 업데이트
            messages.success(request, f"{count}건 출고완료로 변경했습니다.")
            return redirect('confirmed_list')

        elif action == 'delete_multiple':
            # 선택 삭제 버튼을 눌렀을 때
            deleted_info = qs.delete()
            total_deleted = deleted_info[0]  # (삭제된 객체 수, { 'app.Model': count })
            return redirect('confirmed_list')

        elif action == 'revert_waiting':
            """
            '기다리기로 변경' 로직:
            - 기존에 changed_option (옵션변경)이나 confirmed=True 였던 것을 다시 'waiting=True'로 돌림
            - changed_option=''(초기화), changed_option_code=''(초기화)
            - confirmed=False (다시 미확인 상태)
            - waiting=True
            - flow_status=... (선택: pre_send? confirmed? 여기서 '기다리는 중'을 어떤 flow_status로 둘지 결정)
            """
            count = 0
            for s in qs:
                s.changed_option = ""
                s.changed_option_code = ""
                s.waiting = True
                s.confirmed = True
                s.flow_status = 'confirmed'  # 또는 'pre_send' 등, 원하는 흐름상태
                s.save()
                count += 1

            messages.success(request, f"{count}건 기다리기로 변경했습니다.")
            return redirect('confirmed_list')
        
        else:
            # 정의되지 않은 action일 경우
            messages.error(request, "알 수 없는 요청입니다.")
            return redirect('confirmed_list')

    # GET 혹은 다른 메소드로 접근 시 목록 페이지로
    return redirect('confirmed_list')




def shipped_list_view(request):
    """
    flow_status='shipped' 인 데이터들만 보여주는 출고완료 페이지
    """
    shipments = DelayedShipment.objects.filter(flow_status='shipped').order_by('-created_at')
    return render(request, 'delayed_management/shipped_list.html', {'shipments': shipments})


def process_shipped_shipments(request):
    """
    '출고완료 목록' 템플릿에서 POST로 전달된 shipment_ids에 대해
    - action = "delete_multiple"      → 선택 삭제
    - action = "revert_to_confirmed" → 확인완료로 되돌리기(flow_status='confirmed')
    """
    if request.method == 'POST':
        action = request.POST.get('action', '')
        shipment_ids = request.POST.getlist('shipment_ids', [])

        if not shipment_ids:
            messages.warning(request, "선택된 항목이 없습니다.")
            return redirect('shipped_list')  # 실제 템플릿 이름/URL에 맞게 변경

        qs = DelayedShipment.objects.filter(id__in=shipment_ids)

        if action == 'delete_multiple':
            # 예: 실제로는 delete가 아니라 is_deleted 처리 등으로 구현할 수도 있음
            count = qs.delete()[0]  # delete()는 (삭제개수, {테이블정보}) 형태로 반환
            messages.success(request, f"{count}건을 삭제했습니다.")
            return redirect('shipped_list')

        elif action == 'revert_to_confirmed':
            # 확인완료로 되돌리기
            count = qs.update(flow_status='confirmed')
            messages.success(request, f"{count}건을 확인완료 상태로 되돌렸습니다.")
            return redirect('shipped_list')

    # GET 등 다른 요청이면 그냥 목록 페이지로
    return redirect('shipped_list')

def confirmed_list(request):
    """
    flow_status='confirmed' 인 레코드를 기본으로 가져오되,
    GET 파라미터 filter에 따라 waiting / changed_option 필터
    """
    qs = DelayedShipment.objects.filter(flow_status='confirmed')

    filter_val = request.GET.get('filter', 'all')  # 디폴트 'all'
    if filter_val == 'waiting':
        qs = qs.filter(waiting=True)
    elif filter_val == 'changed':
        qs = qs.filter(~Q(changed_option=''))

    shipments = qs.order_by('-created_at')

    # ▼ 안내날짜(range_start, range_end), 예상날짜(expected_start, expected_end) 계산
    for s in shipments:
        # status에 따른 min_days, max_days
        min_days, max_days = ETA_RANGES.get(s.status, (0, 0))

        # (1) 안내날짜(고정) range
        if s.restock_date:
            s.range_start = s.restock_date + timedelta(days=min_days)
            s.range_end   = s.restock_date + timedelta(days=max_days)
        else:
            s.range_start = None
            s.range_end   = None

        # (2) 예상날짜(매번 갱신) range
        if s.expected_restock_date:
            s.expected_start = s.expected_restock_date + timedelta(days=min_days)
            s.expected_end   = s.expected_restock_date + timedelta(days=max_days)
        else:
            s.expected_start = None
            s.expected_end   = None

        # (3) changed_option(문자열) 파싱 (한 개만 있는 상황 가정)
        #     예: "애쉬그레이(R)|4qom1cc59a2c2f3e3i"
        #          또는 "애쉬그레이(R)" (옵션코드 없는 경우)
        parsed_option = {
            "option_name": "",
            "option_code": ""
        }
        if s.changed_option:
            # 만약 실제로 ','가 들어있다면(여러 개로 구분) → 
            # "한 개만"이라는 전제하에 제일 앞의 토큰만 처리 or 그대로 “,”가 없다고 가정
            tk = s.changed_option.strip()

            if '|' in tk:
                nm, cd = tk.split('|', 1)
                nm = nm.strip()
                cd = cd.strip()
                parsed_option["option_name"] = nm
                parsed_option["option_code"] = cd
            else:
                # 옵션코드 없는 경우
                parsed_option["option_name"] = tk

        # 파싱 결과를 모델 인스턴스에 임시로 넣어두면, 템플릿에서 s.changed_option_parsed 로 접근 가능
        s.changed_option_parsed = parsed_option

    return render(request, 'delayed_management/confirmed_list.html', {
        'shipments': shipments,
        'filter_val': filter_val,  # 템플릿에서 현재 필터값 체크용
    })


def get_exchange_options_api(request, shipment_id):
    """
    GET /delayed/api/exchange-options/<shipment_id>/
      → { status: "success", options: [...] }
    """
    try:
        ship = DelayedShipment.objects.get(id=shipment_id)
    except DelayedShipment.DoesNotExist:
        return JsonResponse({"status":"error","message":"Shipment not found"}, status=404)

    if not ship.exchangeable_options:
        return JsonResponse({"status":"success","options":[]})
    arr = [x.strip() for x in ship.exchangeable_options.split(',') if x.strip()]
    return JsonResponse({"status":"success","options": arr})


def get_seller_tool_options_api(request, shipment_id):
    """
    GET /delayed/api/sellertool-options/<shipment_id>/
      → {
          "status":"success",
          "allOptions":[
             {"optionName":"애쉬그레이(R)", "optionCode":"4qom1cc59a2c2f3e3i", "stock":2},
             ...
          ],
          "current":["애쉬그레이(R)", "샌드카키(L)", ...]
        }
    """
    try:
        ship = DelayedShipment.objects.get(id=shipment_id)
    except DelayedShipment.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Shipment not found"}, status=404)

    # (A) DB에 저장된 교환옵션 → 리스트
    if ship.exchangeable_options:
        current_list = [x.strip() for x in ship.exchangeable_options.split(',') if x.strip()]
    else:
        current_list = []

    # (B) 기존 API Clients
    info = get_option_info_by_code(ship.option_code)
    if not info or not info.get('productName'):
        return JsonResponse({"status": "error", "message": "상품 정보가 없습니다."}, status=400)

    product_name = info['productName']

    # (C) 동일 상품의 모든 옵션 조회
    codes = get_all_options_by_product_name(product_name)
    inventory_map = get_inventory_by_option_codes(codes)  # { code: stockUnit, ...}
    detail_map = get_options_detail_by_codes(codes)       # { code: {optionName, salesPrice, ...}, ...}

    all_options = []
    base_sales_price = info['salesPrice']

    # (D) 전체 옵션 나열
    for code in codes:
        det = detail_map.get(code)
        if not det:
            continue
        stock = inventory_map.get(code, 0)
        # (선택) 만약 "판매가 동일 & 재고 >= 0" 조건만 표시 etc.
        if det['salesPrice'] == base_sales_price:
            all_options.append({
                "optionName": det['optionName'],
                "optionCode": code,
                "stock": stock
            })

    return JsonResponse({
        "status": "success",
        "allOptions": all_options,   # [{optionName, optionCode, stock}, ...]
        "current": current_list      # ["애쉬그레이(R)", "샌드카키(L)", ... ]
    })

@csrf_exempt
def add_exchange_option_api(request, shipment_id):
    """
    POST /delayed/exchange-options/<shipment_id>/add/
    FormData: option_code, option_name
    """
    if request.method != 'POST':
        return JsonResponse({"status":"error","message":"POST only"}, status=405)

    try:
        ship = DelayedShipment.objects.get(id=shipment_id)
    except DelayedShipment.DoesNotExist:
        return JsonResponse({"status":"error","message":"Shipment not found"}, status=404)

    opt_code = request.POST.get('option_code','').strip()
    opt_name = request.POST.get('option_name','').strip()
    if not opt_code or not opt_name:
        return JsonResponse({"status":"error","message":"option_code/option_name required"}, status=400)

    # 추가
    old_str = ship.exchangeable_options or ""
    arr = [x.strip() for x in old_str.split(',') if x.strip()]
    if opt_name in arr:
        return JsonResponse({"status":"error","message":"이미 존재하는 옵션명입니다."}, status=400)

    arr.append(opt_name)
    ship.exchangeable_options = ",".join(arr)
    ship.save()
    return JsonResponse({"status":"success"})

@csrf_exempt
def remove_exchange_option_api(request, shipment_id):
    """
    POST /delayed/exchange-options/<shipment_id>/remove/
      FormData: option_name
    """
    if request.method != 'POST':
        return JsonResponse({"status":"error","message":"POST only"}, status=405)

    try:
        ship = DelayedShipment.objects.get(id=shipment_id)
    except DelayedShipment.DoesNotExist:
        return JsonResponse({"status":"error","message":"Shipment not found"}, status=404)

    opt_name = request.POST.get('option_name','').strip()
    if not opt_name:
        return JsonResponse({"status":"error","message":"option_name required"}, status=400)

    if not ship.exchangeable_options:
        return JsonResponse({"status":"error","message":"no exchangeable_options"}, status=400)

    arr = [x.strip() for x in ship.exchangeable_options.split(',') if x.strip()]
    if opt_name not in arr:
        return JsonResponse({"status":"error","message":"해당 옵션이 없습니다."}, status=400)

    arr.remove(opt_name)
    ship.exchangeable_options = ",".join(arr)
    ship.save()

    return JsonResponse({"status":"success"})



def save_seller_tool_options_api(request, shipment_id):
    """
    POST /delayed/api/save-sellertool-options/<shipment_id>/
      FormData: selected_options_json = '["빨강_M", "파랑_L"]'
    """
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    import json
    from .models import DelayedShipment

    if request.method != 'POST':
        return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)

    # DelayedShipment 조회
    shipment = get_object_or_404(DelayedShipment, id=shipment_id)

    # 1) JSON 파싱
    selected_options_json = request.POST.get('selected_options_json', '[]')
    try:
        selected_options = json.loads(selected_options_json)  # ex) ["빨강_M","파랑_L"]
    except:
        return JsonResponse({"status": "error", "message": "JSON parsing error"}, status=400)

    # 2) DB에 저장할 문자열
    new_opts_str = ', '.join(selected_options)
    shipment.exchangeable_options = new_opts_str
    shipment.save()

    return JsonResponse({"status": "success"})

def upload_platform_mapping(request):
    """
    1. 엑셀 파일 파싱
    2. 각 행에 대해 ExternalPlatformMapping 저장
    """
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            # 에러 처리
            ...
        
        # 1) 엑셀 파싱 로직
        #    예: pandas, openpyxl, etc.
        df = pd.read_excel(file)
        
        # 2) 각 행을 돌면서 DB 저장
        for idx, row in df.iterrows():
            option_code = row.get('option_code', '')
            platform_name = row.get('platform_name', '')
            product_id = row.get('platform_product_id', '')
            option_id = row.get('platform_option_id', '')

            # create or update
            obj, created = ExternalPlatformMapping.objects.update_or_create(
                option_code=option_code,
                platform_name=platform_name,
                defaults={
                    'platform_product_id': product_id,
                    'platform_option_id': option_id
                }
            )
        # 업로드 완료 후 리다이렉트 or 메시지
        return redirect('out_of_stock_management')  # 품절관리 페이지로 이동
    
    return render(request, 'upload_platform_mapping.html') 












def api_option_list_view(request):
    """
    ExternalProductOption 테이블을 조회하여,
    - 상품명(product__product_name) or
    - 옵션코드(seller_manager_code)
    로 검색할 수 있는 리스트 페이지.

    (option_id, option_name1/2, stock_quantity 등도 함께 표시)
    """
    search_query = request.GET.get('search_query', '').strip()

    # 1) ExternalProductOption 목록 (product select_related)
    qs = ExternalProductOption.objects.select_related('product').order_by('-id')

    # 2) 검색 조건 적용
    #    상품명 or 옵션코드
    if search_query:
        qs = qs.filter(
            Q(product__product_name__icontains=search_query) |
            Q(seller_manager_code__icontains=search_query) |
            Q(option_id__icontains=search_query)
        )

    # 3) 페이지네이션 (페이지당 50개)
    paginator = Paginator(qs, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # elided_page_range (페이지 버튼이 많아질 때 간략히 표시)
    page_range_custom = page_obj.paginator.get_elided_page_range(
        page_obj.number, on_each_side=2, on_ends=1
    )

    # 4) 템플릿 컨텍스트
    context = {
        "options": page_obj.object_list,       # 실제 옵션 목록
        "page_obj": page_obj,
        "page_range_custom": page_range_custom,
        "search_query": search_query,          # 템플릿에서 검색어 유지
    }
    return render(request, "delayed_management/api_list.html", context)


def update_naver_option_list(request):
    """
    네이버 API로부터 상품/옵션 리스트를 받아, OutOfStock 모델에 매핑해서 업데이트한다.
    """
    account_name = request.GET.get("account_name", "").strip()
    if not account_name:
        return JsonResponse({"error": "account_name이 없습니다."}, status=400)

    # 1) 네이버 계정 찾기
    account_info = None
    for acct in NAVER_ACCOUNTS:
        if account_name in acct["names"]:
            account_info = acct
            break

    if not account_info:
        msg = f"[update_naver_option_list] 계정 '{account_name}'을 NAVER_ACCOUNTS에서 찾을 수 없습니다."
        logger.warning(msg)
        return JsonResponse({"error": msg}, status=400)

    # 2) 네이버 상품/옵션 목록 가져오기
    is_ok, detailed_list = fetch_naver_products_with_details(account_info)
    if not is_ok:
        logger.error(f"[update_naver_option_list] fetch fail: {detailed_list}")
        return JsonResponse({"error": "네이버 상품목록 조회 실패", "detail": detailed_list}, status=400)

    logger.debug(f"[update_naver_option_list] detailed_list count={len(detailed_list)}")

    # 업데이트된 OutOfStock 레코드(또는 생성된 레코드) 수
    updated_count = 0

    for prod_data in detailed_list:
        # 공통 영역(상품 레벨)
        origin_no      = prod_data.get("originProductNo")            # 상품ID
        rep_image_dict = prod_data.get("representativeImage", {})     # 대표이미지
        if isinstance(rep_image_dict, dict):
            rep_img_url = rep_image_dict.get("url", "")
        else:
            rep_img_url = ""

        # 네이버에서 사용하는 '정가(salePrice)'와 '할인가(discountedPrice)'
        naver_sale_price       = prod_data.get("salePrice", 0)        # 정가
        naver_discounted_price = prod_data.get("discountedPrice", 0)  # 할인가
        product_name           = prod_data.get("productName", "")     # 주문상품명(주문 시 노출되는 상품명)

        # (옵션 목록 파싱)
        option_combos = prod_data.get("optionCombinations", [])
        if not option_combos:
            # 옵션이 없는 단일상품인 경우도 처리 가능
            # 여기서는 단일상품이라면 option_id를 임의로 처리하거나 생략할 수 있습니다.
            # 예시) option_id를 origin_no 로 대체
            single_option_id_stock = prod_data.get("stockQuantity", 0)

            _, created = OutOfStock.objects.update_or_create(
                option_id=str(origin_no),  # 옵션ID가 꼭 있어야 unique한 항목이므로
                defaults={
                    "platform_name":          account_name,     # 플랫폼명
                    "representative_image":   rep_img_url,      # 대표이미지
                    "product_id":             origin_no,        # 상품ID
                    "option_id_stock":        single_option_id_stock,
                    "expected_restock_date":  None,             # API상 별도 필드가 없으므로 None
                    "order_product_name":     product_name,
                    "order_option_name_01":   "",               # 단일상품이므로 옵션명 비어있음
                    "order_option_name_02":   "",               # 단일상품이므로 옵션명 비어있음
                    "option_code":            "",               # 단일상품이므로 (or sellerManagerCode가 없으므로)
                    "original_price":         naver_sale_price,
                    "sale_price":             naver_discounted_price,
                    "add_option_price":       0,                # 단일상품이라 추가옵션가 없음
                    # out_of_stock_at, updated_at 은 자동 or 상황에 따라 세팅
                }
            )
            updated_count += 1
            continue

        # 옵션이 여러 개 존재하는 경우
        for combo in option_combos:
            opt_id      = combo.get("id")                     # 옵션ID
            opt_stk     = combo.get("stockQuantity", 0)       # 옵션ID재고
            opt_name1   = combo.get("optionName1", "")        # 주문옵션명01
            opt_name2   = combo.get("optionName2", "")        # 주문옵션명02
            opt_code    = combo.get("sellerManagerCode", "")  # 옵션코드
            add_opt_prc = combo.get("price", 0)               # 추가옵션가

            # OutOfStock 업데이트 (option_id 가 unique)
            _, created = OutOfStock.objects.update_or_create(
                option_id=str(opt_id),  # 필드 unique=True
                defaults={
                    "platform_name":         account_name,       # 플랫폼명
                    "representative_image":  rep_img_url,        # 대표이미지
                    "product_id":            origin_no,          # 상품ID
                    "option_id_stock":       opt_stk,            # 옵션ID재고
                    "expected_restock_date": None,               # 현재 API 데이터 없음
                    "order_product_name":    product_name,       # 주문상품명 = productName
                    "order_option_name_01":  opt_name1,          # 주문옵션명01
                    "order_option_name_02":  opt_name2,          # 주문옵션명02 (추가)
                    "option_code":           opt_code,           # 옵션코드 = sellerManagerCode
                    "original_price":        naver_sale_price,   # 정가 = salePrice
                    "sale_price":            naver_discounted_price,  # 할인가 = discountedPrice
                    "add_option_price":      add_opt_prc,        # 추가옵션가 = combo.price
                    # out_of_stock_at, updated_at → 필요 시 추가 로직
                }
            )
            updated_count += 1

    return JsonResponse({
        "message": f"{account_name} 상품/옵션 {updated_count}건 업데이트 완료",
        "count": updated_count,
    })


def update_coupang_option_list(request):
    """
    쿠팡 API로부터 상품/옵션 데이터를 가져와 OutOfStock 테이블에 매핑.
    seller_product_id=14447739495, 13639740382 가 실제로 big_list에 들어오는지 여부를 확인(로그)한다.
    """

    account_name = request.GET.get("account_name", "").strip()
    if not account_name:
        return JsonResponse({"error": "account_name이 없습니다."}, status=400)

    # 1) 쿠팡 계정 찾기
    account_info = None
    for acct in COUPANG_ACCOUNTS:
        if account_name in acct["names"]:
            account_info = acct
            break

    if not account_info:
        msg = f"[update_coupang_option_list] '{account_name}' 계정을 COUPANG_ACCOUNTS에서 찾을 수 없습니다."
        logger.warning(msg)
        return JsonResponse({"error": msg}, status=400)

    logger.debug(f"[update_coupang_option_list] account_name={account_name}, account_info={account_info}")

    # 2) 등록상품 목록 조회
    is_ok, big_list = fetch_coupang_all_seller_products(account_info, max_per_page=50)
    if not is_ok:
        logger.error(f"[update_coupang_option_list] 쿠팡 상품목록 조회 실패: {big_list}")
        return JsonResponse({"error": "쿠팡 상품목록 조회 실패", "detail": big_list}, status=400)

    product_count = len(big_list)
    logger.info(f"[update_coupang_option_list] {account_name}: productCount={product_count}")

    # (A) 집계용 변수
    fetched_options_count = 0
    new_count = 0
    updated_count = 0
    failed_count = 0
    upload_fail_count = 0

    # 우리가 확인하고 싶은 상품ID 목록
    SPECIAL_IDS = {14447739495, 13639740382}

    # 3) 상품 목록 순회
    for product_data in big_list:
        seller_product_id = product_data.get("sellerProductId")
        if not seller_product_id:
            continue

        # (B) 확인 로깅: 이 productId 가 우리가 찾는 셋에 있는지?
        if seller_product_id in SPECIAL_IDS:
            logger.info("[DEBUG] Special seller_product_id=%s 가 big_list에 포함되어 있습니다!", seller_product_id)

        # (C) 옵션/재고 정보 수집
        is_ok2, flattened_list = fetch_coupang_seller_product_with_options(account_info, seller_product_id)
        if not is_ok2:
            logger.warning(f"[update_coupang_option_list] 상품ID={seller_product_id} 옵션 조회 실패 → skip")
            continue
        if not flattened_list:
            continue

        # 전체 옵션 개수 누적
        fetched_options_count += len(flattened_list)

        # (D) 각 옵션 저장
        for row in flattened_list:
            vendor_item_id     = row.get("optionID", "")
            rep_img_url        = row.get("representativeImage", "")
            product_name       = row.get("productName", "")
            opt_name1          = row.get("optionName1", "")
            opt_name2          = row.get("optionName2", "")
            option_stock       = row.get("optionStock", 0)
            option_code        = row.get("optionSellerCode", "")
            sale_price_val     = row.get("salePrice", 0)

            if not vendor_item_id:
                upload_fail_count += 1
                logger.warning(f"[update_coupang_option_list] vendor_item_id 없음 → skip row={row}")
                continue

            defaults_data = {
                "platform_name":         account_name,
                "representative_image":  rep_img_url,
                "product_id":            seller_product_id,
                "option_id_stock":       option_stock,
                "order_product_name":    product_name,
                "order_option_name_01":  opt_name1,
                "order_option_name_02":  opt_name2,
                "option_code":           option_code,
                "original_price":        sale_price_val,
                "sale_price":            sale_price_val,
                "add_option_price":      0,
            }

            try:
                obj, created = OutOfStock.objects.update_or_create(
                    option_id=str(vendor_item_id),
                    defaults=defaults_data
                )
                if created:
                    new_count += 1
                else:
                    updated_count += 1

            except Exception as e:
                failed_count += 1
                logger.exception(
                    f"[update_coupang_option_list] update_or_create 예외 row={row}, error={e}"
                )

    # 5) 최종 로그
    logger.info(
        "[update_coupang_option_list] %s: productCount=%d, fetchedOptions=%d, new=%d, updated=%d, "
        "failed=%d, uploadFail=%d",
        account_name, product_count, fetched_options_count, new_count, updated_count,
        failed_count, upload_fail_count
    )

    # 6) 최종 응답
    msg = (
        f"{account_name} 상품(옵션) 업데이트 완료"
        f"(상품={product_count}, 옵션={fetched_options_count}, "
        f"신규={new_count}, 업데이트={updated_count}, "
        f"예외={failed_count}, 스킵={upload_fail_count}"
    )

    return JsonResponse({
        "message": msg,
        "product_count": product_count,
        "fetched_options_count": fetched_options_count,
        "new_count": new_count,
        "updated_count": updated_count,
        "failed_count": failed_count,
        "upload_fail_count": upload_fail_count,
    })





def outofstock_list_view(request):
    """
    OutOfStock 테이블을 조회하여,
    (상품ID / 옵션ID / 옵션코드 / 주문상품명 등) 검색 & 페이지네이션 리스트 페이지
    """
    search_query = request.GET.get("search_query", "").strip()

    # 1) OutOfStock 목록
    qs = OutOfStock.objects.all().order_by("-id")

    # 2) 검색 조건 적용
    if search_query:
        qs = qs.filter(
            Q(product_id__icontains=search_query) |
            Q(option_id__icontains=search_query) |
            Q(option_code__icontains=search_query) |
            Q(order_product_name__icontains=search_query)
        )

    # 3) 페이지네이션
    paginator = Paginator(qs, 50)  # 한 페이지당 50개
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    # 페이지 번호들 (많을 경우 elided)
    page_range_custom = page_obj.paginator.get_elided_page_range(
        number=page_obj.number,
        on_each_side=2,
        on_ends=1
    )

    context = {
        "options": page_obj.object_list,
        "page_obj": page_obj,
        "page_range_custom": page_range_custom,
        "search_query": search_query,
    }
    return render(request, "delayed_management/outofstock_list.html", context)

def out_of_stock_management_view(request):
    """
    OutOfStock 테이블 중에서 '옵션매핑 전송'된(예: status=1) 레코드만 조회.
    이후 (A) 품절여부(outofstock/instock), (B) 검색, (C) 페이지네이션 적용
    """
    from django.db.models import Q
    from django.core.paginator import Paginator
    from django.shortcuts import render
    from .models import OutOfStock

    filter_val = request.GET.get('filter', 'all').strip()  # 'outofstock' / 'instock' / 'all'
    search_query = request.GET.get('search_query', '').strip()

    # 1) OutOfStock 중 status=1 (옵션매핑 전송된 것만 표시)
    base_qs = OutOfStock.objects.filter(status=1).order_by('-updated_at')

    # 2) 필터 적용
    # outofstock → option_id_stock <= 0 or isnull
    # instock → option_id_stock > 0
    # all → (status=1 전체)
    if filter_val == 'outofstock':
        base_qs = base_qs.filter(Q(option_id_stock__lte=0) | Q(option_id_stock__isnull=True))
    elif filter_val == 'instock':
        base_qs = base_qs.filter(option_id_stock__gt=0)
    # 'all' 일 경우: 필터 없이 그대로

    # 3) 검색 (플랫폼명, 주문상품명, 옵션코드, 옵션ID 등)
    if search_query:
        lower_query = search_query.lower()
        base_qs = base_qs.filter(
            Q(platform_name__icontains=lower_query) |
            Q(order_product_name__icontains=lower_query) |
            Q(option_code__icontains=lower_query) |
            Q(option_id__icontains=lower_query)
        )

    # 4) 페이지네이션
    paginator = Paginator(base_qs, 100)  # 페이지당 100개
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    page_range_custom = page_obj.paginator.get_elided_page_range(
        page_obj.number, on_each_side=2, on_ends=1
    )

    context = {
        "details": page_obj.object_list,       # 템플릿에서 'details'로 반복문
        "page_obj": page_obj,
        "page_range_custom": page_range_custom,
        "filter_val": filter_val,
        "search_query": search_query,
    }
    return render(request, "delayed_management/out_of_stock_management.html", context)



def out_of_stock_delete_all_view(request):
    """
    '전체 삭제' 버튼 클릭 시 동작:
    - 실제로 삭제하지 않고, OutOfStock의 status를 0으로 변경
      (즉, status=1 => status=0 으로 하여 품절관리에서 제외)
    - 완료 후 품절관리 페이지로 리다이렉트
    """
    from django.shortcuts import redirect
    from django.contrib import messages
    from .models import OutOfStock

    # (A) OutOfStock의 status=1 → status=0 으로 업데이트
    OutOfStock.objects.filter(status=1).update(status=0)

    # (B) 완료 후 메시지 표시
    messages.success(request, "전체 품절 데이터의 상태값을 0으로 바꿨습니다. (실제 삭제는 아님)")
    return redirect('out_of_stock_management')

@require_POST
def update_seller_tool_stock(request):
    """
    (1) OutOfStock 중 status=1인 레코드에서 option_code를 수집
    (2) option_code들을 100개씩 나누어 셀러툴 API(get_inventory_by_option_codes) 조회
    (3) OutOfStock.seller_tool_stock 갱신
    (4) 결과: “업데이트 완료 X건, 변경없음 Y건”을 반환(JSON)
    """


    logger = logging.getLogger(__name__)

    try:
        # (A) OutOfStock 중 status=1인 레코드만 조회
        outofstock_qs = OutOfStock.objects.filter(status=1)
        if not outofstock_qs.exists():
            return JsonResponse({
                "success": False,
                "message": "status=1 인 OutOfStock 데이터가 없어 업데이트할 항목이 없습니다."
            }, status=400)

        # (B) option_code 수집 (공백/None 제외)
        code_list = []
        for o in outofstock_qs:
            code = (o.option_code or "").strip()
            if code:
                code_list.append(code)
        code_list = list(set(code_list))  # 중복 제거

        if not code_list:
            return JsonResponse({
                "success": False,
                "message": "status=1 레코드 중 option_code가 없어 업데이트 불가"
            }, status=400)

        # (C) 여러번 나누어 셀러툴 API 호출
        CHUNK_SIZE = 100
        stock_map = {}

        def chunked(seq, size):
            for i in range(0, len(seq), size):
                yield seq[i : i + size]

        for chunk in chunked(code_list, CHUNK_SIZE):
            partial_map = get_inventory_by_option_codes(chunk)
            # partial_map은 {option_code: 재고수량} 형태라고 가정
            stock_map.update(partial_map)

        # (D) DB 갱신
        updated_count = 0
        unchanged_count = 0

        for out_item in outofstock_qs:
            code = (out_item.option_code or "").strip()
            if not code:
                continue

            old_val = out_item.seller_tool_stock or 0
            new_val = stock_map.get(code, 0)

            if new_val != old_val:
                out_item.seller_tool_stock = new_val
                out_item.save()
                updated_count += 1
            else:
                unchanged_count += 1

        return JsonResponse({
            "success": True,
            "message": (
                f"[update_seller_tool_stock] 셀러툴 재고 업데이트 완료! "
                f"(업데이트: {updated_count}건, 변경없음: {unchanged_count}건)"
            )
        })

    except Exception as e:
        logger.exception("[update_seller_tool_stock] 예외발생")
        return JsonResponse({
            "success": False,
            "message": str(e),
        }, status=500)


@require_POST
def option_id_stock_update_view(request):
    """
    { "ids": [3,5,7, ...] } 형태로 넘어온 OutOfStock pk 목록에 대해
    네이버/쿠팡 API를 호출해 'option_id_stock' 등 재고 정보를 갱신한다.
    """
    import logging, json
    logger = logging.getLogger(__name__)

    from django.http import JsonResponse
    from .models import OutOfStock
    from .api_clients import (
        NAVER_ACCOUNTS, COUPANG_ACCOUNTS,
        fetch_naver_option_stock,  # 네이버 옵션 재고 조회 (originNo 기반)
        get_coupang_item_inventories  # 쿠팡 옵션 재고 조회 (vendorItemId 기반)
    )

    try:
        data = json.loads(request.body)
        pk_list = data.get("ids", [])
        if not pk_list:
            return JsonResponse({"success": False, "message": "선택된 항목이 없습니다."}, status=400)
    except:
        return JsonResponse({"success": False, "message": "Invalid JSON"}, status=400)

    out_qs = OutOfStock.objects.filter(pk__in=pk_list)
    if not out_qs.exists():
        return JsonResponse({"success": False, "message": "대상 OutOfStock이 없습니다."}, status=400)

    updated_count = 0

    for out_item in out_qs:
        platform = (out_item.platform_name or "").strip()
        origin_no = (out_item.product_id or "").strip()  # 네이버 originNo / 쿠팡 sellerItemId?
        vendor_id = (out_item.option_id or "").strip()   # 네이버 combo.id / 쿠팡 vendorItemId

        # [1] Naver 계정?
        account_info = None
        for acct in NAVER_ACCOUNTS:
            if platform in acct.get("names", []):
                account_info = acct
                break

        if account_info:
            # 네이버
            is_ok, combos = fetch_naver_option_stock(account_info, origin_no)
            if not is_ok:
                logger.debug(f"[option_id_stock_update] 네이버 API 실패: {combos}")
                continue

            # combos = [{"id":..., "stockQuantity":..., "price":...}, ...]
            combo = next((c for c in combos if str(c["id"]) == vendor_id), None)
            if not combo:
                logger.debug(f"OutOfStock id={out_item.id}, vendor_id={vendor_id} → 옵션 못찾음")
                continue

            old_stk = out_item.option_id_stock
            new_stk = combo.get("stockQuantity", 0)
            if new_stk != old_stk:
                out_item.option_id_stock = new_stk
                # 필요 시 out_item.sale_price = combo["price"]
                out_item.save()
                updated_count += 1
            continue

        # [2] Coupang 계정?
        account_info = None
        for acct in COUPANG_ACCOUNTS:
            if platform in acct.get("names", []):
                account_info = acct
                break

        if account_info:
            # 쿠팡
            is_ok, item_data = get_coupang_item_inventories(account_info, vendor_id)
            if not is_ok:
                logger.debug(f"[option_id_stock_update] 쿠팡 API 실패: {item_data}")
                continue

            old_stk = out_item.option_id_stock
            new_stk = item_data.get("amountInStock", 0)
            if new_stk != old_stk:
                out_item.option_id_stock = new_stk
                # 필요 시 out_item.sale_price = item_data["salePrice"]
                out_item.save()
                updated_count += 1
            continue

        # [3] Unknown platform
        logger.debug(f"OutOfStock id={out_item.id}: Unknown platform='{platform}' skip")

    return JsonResponse({
        "success": True,
        "updated_count": updated_count,
        "message": f"옵션ID재고 업데이트 완료: {updated_count}건 변경됨."
    })


@require_POST
def do_out_of_stock_view(request):
    """
    1) JSON { "detail_ids": [1,2,3...] } 로 들어온 OutOfStock PK들에 대해,
       (네이버) 정가/옵션추가금과 함께 품절처리 (new_stock=0) API 호출
       (쿠팡) 재고=0 만 전송
    2) 성공 시 out_item.option_id_stock=0, out_item.out_of_stock_at = now
    """
    import logging
    import json
    from django.http import JsonResponse
    from django.utils import timezone
    from django.views.decorators.http import require_POST

    from .models import OutOfStock
    from .api_clients import (
        naver_update_option_stock,    # (가정) → (origin_no, option_id, new_stock, keep_price, base_sale_price, ...)
        coupang_update_item_stock     # (가정) → (vendor_item_id, new_stock, ...)
    )

    logger = logging.getLogger(__name__)

    if request.method != "POST":
        return JsonResponse({
            "success": False,
            "message": "Only POST method is allowed."
        }, status=405)

    try:
        body = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON body"}, status=400)

    raw_list = body.get("detail_ids", [])
    # ["3","7","15"] 등 → int 변환
    pk_list = []
    for val in raw_list:
        try:
            pk_list.append(int(val))
        except ValueError:
            logger.debug(f"[do_out_of_stock_view] pk 변환 실패: {val}")

    if not pk_list:
        return JsonResponse({"success": False, "message": "선택된 옵션이 없습니다."}, status=400)

    out_qs = OutOfStock.objects.filter(pk__in=pk_list)
    if not out_qs.exists():
        return JsonResponse({"success": False, "message": "대상 OutOfStock이 없습니다."}, status=400)

    updated_count = 0
    now_time = timezone.now()

    for out_item in out_qs:
        platform_name = (out_item.platform_name or "").strip()
        origin_no     = (out_item.product_id or "").strip()  # 네이버 originNo / 쿠팡 sellerProductId?
        opt_id        = (out_item.option_id or "").strip()   # 네이버 optionID / 쿠팡 vendorItemId?

        if not origin_no or not opt_id:
            logger.warning(f"[do_out_of_stock_view] OutOfStock(pk={out_item.pk}) origin_no/option_id가 없습니다 → skip")
            continue

        # 네이버인지 쿠팡인지 확인
        if platform_name in ["니뜰리히","수비다","수비다 SUBIDA","노는개최고양","노는 개 최고양","아르빙"]:
            # (A) 네이버 품절처리
            # new_stock=0, base_sale_price=out_item.original_price, keep_price=out_item.add_option_price
            try:
                is_ok, msg = naver_update_option_stock(
                    origin_no=origin_no,
                    option_id=opt_id,
                    new_stock=0,
                    base_sale_price=out_item.original_price,  # 정가
                    keep_price=out_item.add_option_price,      # 옵션추가금
                    platform_name=platform_name
                )
                # ── 여기서 msg 안에 네이버 측 응답이 들어있다면 아래처럼 로깅
                logger.debug(
                    f"[NAVER API 응답] pk={out_item.pk}, origin_no={origin_no}, "
                    f"option_id={opt_id}, is_ok={is_ok}, msg={msg}"
                )

                if is_ok:
                    out_item.option_id_stock = 0
                    out_item.out_of_stock_at = now_time
                    out_item.save()
                    updated_count += 1
                else:
                    logger.warning(f"[NAVER 품절실패] pk={out_item.pk}, msg={msg}")

            except Exception as e:
                # 네이버 서버 통신 과정에서 예외가 발생하면 로그를 남김
                logger.error(f"[NAVER API 연동 중 예외 발생] pk={out_item.pk}, error={e}", exc_info=True)

        elif platform_name in ["쿠팡01", "쿠팡02"]:
            # (B) 쿠팡 품절처리 → 재고=0 만 전송
            try:
                is_ok, msg = coupang_update_item_stock(
                    vendor_item_id=opt_id,
                    new_stock=0,
                    platform_name=platform_name
                )
                # 쿠팡 쪽 응답도 동일하게 로깅 가능
                logger.debug(
                    f"[쿠팡 API 응답] pk={out_item.pk}, vendor_item_id={opt_id}, "
                    f"is_ok={is_ok}, msg={msg}"
                )

                if is_ok:
                    out_item.option_id_stock = 0
                    out_item.out_of_stock_at = now_time
                    out_item.save()
                    updated_count += 1
                else:
                    logger.warning(f"[쿠팡 품절실패] pk={out_item.pk}, msg={msg}")

            except Exception as e:
                logger.error(f"[쿠팡 API 연동 중 예외 발생] pk={out_item.pk}, error={e}", exc_info=True)

        else:
            logger.debug(f"[do_out_of_stock_view] Unknown platform='{platform_name}', skip")

    return JsonResponse({
        "success": True,
        "message": f"{updated_count}건 품절처리 완료.",
        "updated_count": updated_count
    })


@require_POST
def add_stock_9999_view(request):
    """
    체크된 OutOfStock 레코드들에 대해,
    - 네이버: put_naver_option_stock_9999 (base_sale_price=original_price, keep_price=add_option_price)
    - 쿠팡: put_coupang_option_stock_9999 (vendor_item_id=option_id)
    후 재고=9999로 설정
    """
    import logging, json
    logger = logging.getLogger(__name__)

    from django.http import JsonResponse
    from django.views.decorators.http import require_POST
    from django.utils import timezone

    # 모델/함수 임포트
    from .models import OutOfStock
    from .api_clients import (
        put_naver_option_stock_9999,
        put_coupang_option_stock_9999,
    )

    # 1) 요청 메소드 체크
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Only POST method allowed"}, status=405)

    # 2) 요청 바디 파싱
    try:
        body = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON body"}, status=400)

    # 3) PK 목록
    raw_ids = body.get('detail_ids', [])  # 예: [3, 5, 10]
    # 문자열/숫자 섞여 있을 수 있으니 정수 변환
    pk_list = []
    for val in raw_ids:
        try:
            pk_list.append(int(val))
        except ValueError:
            logger.debug(f"[add_stock_9999_view] '{val}' → int 변환 실패, skip")

    if not pk_list:
        return JsonResponse({"success":False,"message":"선택된 항목이 없습니다."}, status=400)

    # 4) OutOfStock 쿼리
    out_qs = OutOfStock.objects.filter(pk__in=pk_list)
    if not out_qs.exists():
        return JsonResponse({"success":False,"message":"해당 OutOfStock이 없습니다."}, status=400)

    updated_count = 0

    # 5) 반복 처리
    for out_item in out_qs:
        platform   = (out_item.platform_name or "").strip()
        origin_no  = (out_item.product_id or "").strip()  # 네이버 originNo / 쿠팡 sellerProductId? (필요시 수정)
        option_id  = (out_item.option_id or "").strip()   # 네이버 option ID / 쿠팡 vendorItemId?

        # (A) 네이버의 경우 base_sale_price=original_price, keep_price=add_option_price
        if platform in ["니뜰리히", "수비다 SUBIDA", "노는 개 최고양", "아르빙"]:
            is_ok, msg = put_naver_option_stock_9999(
                origin_no=origin_no,
                option_id=option_id,
                platform_name=platform,
                base_sale_price=out_item.original_price,  # 정가
                keep_price=out_item.add_option_price      # 추가옵션가
            )
            if is_ok:
                out_item.option_id_stock = 9999
                out_item.save()
                updated_count += 1
            else:
                logger.warning(
                    f"[add_stock_9999_view] NAVER 재고 9999 실패: pk={out_item.pk}, msg={msg}"
                )

        # (B) 쿠팡의 경우 just put_coupang_option_stock_9999
        elif platform in ["쿠팡01","쿠팡02"]:
            is_ok, msg = put_coupang_option_stock_9999(
                vendor_item_id=option_id,
                platform_name=platform
            )
            if is_ok:
                out_item.option_id_stock = 9999
                out_item.save()
                updated_count += 1
            else:
                logger.warning(
                    f"[add_stock_9999_view] 쿠팡 재고 9999 실패: pk={out_item.pk}, msg={msg}"
                )
        else:
            logger.debug(f"[add_stock_9999_view] Unknown platform={platform}, skip")

    return JsonResponse({
        "success": True,
        "updated_count": updated_count,
        "message": f"{updated_count}건 재고를 9999로 변경했습니다."
    })


def update_seller_tool_and_increase_stock_view(request):
    """
    1) OutOfStock에서 모든 option_code 수집 (또는 특정 조건이 있다면 filter)
    2) 여러 번 나눠 get_inventory_by_option_codes(...) 로 셀러툴 재고 확인
    3) old_seller_tool != new_seller_tool 이면 DB 업데이트
    4) new_seller_tool - old_seller_tool = diff > 0 → 실제 옵션재고 늘림
       - diff <= 10 → (재고 += diff)
       - diff > 10  → 재고=9999
    5) 네이버/쿠팡 API도 같이 호출 (정가+옵션가)
    """
    import logging
    logger = logging.getLogger(__name__)
    from django.shortcuts import redirect
    from django.contrib import messages

    from .models import OutOfStock
    from .api_clients import (
        get_inventory_by_option_codes,
        naver_update_option_stock,      
        coupang_update_item_stock,      
        put_naver_option_stock_9999,    
        put_coupang_option_stock_9999
    )

    def chunked(seq, size):
        for i in range(0, len(seq), size):
            yield seq[i : i + size]

    # (A) OutOfStock에서 option_code 수집
    out_qs = (
        OutOfStock.objects
        .filter(status=1)  # 여기서 status=1 조건 추가
        .exclude(option_code__isnull=True)
        .exclude(option_code__exact="")
    )
    if not out_qs.exists():
        messages.warning(request, "OutOfStock에 옵션코드가 없습니다.")
        return redirect('out_of_stock_management')

    code_list = []
    for out_item in out_qs:
        code_list.append(out_item.option_code.strip())

    code_list = list(set(code_list))
    if not code_list:
        messages.warning(request, "옵션코드가 없어 재고 업데이트 불가.")
        return redirect('out_of_stock_management')

    logger.debug(f"[update_seller_tool_and_increase_stock_view] code_list={code_list}")

    # (B) 여러 번 나눠서 셀러툴 API 조회
    CHUNK_SIZE = 100
    stock_map = {}
    for chunk in chunked(code_list, CHUNK_SIZE):
        partial_map = get_inventory_by_option_codes(chunk)
        stock_map.update(partial_map)

    # (C) OutOfStock 반복
    updated_st_count = 0
    action_count = 0

    for out_item in out_qs:
        code = out_item.option_code.strip()
        old_seller_tool = out_item.seller_tool_stock
        new_seller_tool = stock_map.get(code, 0)  # 셀러툴 API에서 가져온 재고

        # (C1) 셀러툴재고 업데이트
        if new_seller_tool != old_seller_tool:
            out_item.seller_tool_stock = new_seller_tool
            out_item.save()
            updated_st_count += 1

            diff = new_seller_tool - old_seller_tool
            logger.debug(
                f"[seller_tool diff] code={code}, old={old_seller_tool}, "
                f"new={new_seller_tool}, diff={diff}"
            )

            # (C2) diff>0 이면 실제 옵션 재고 늘림
            if diff > 0:
                platform     = (out_item.platform_name or "").strip()
                origin_no    = (out_item.product_id or "").strip()
                opt_id       = (out_item.option_id or "").strip()
                base_price   = out_item.original_price or 0    # 정가
                keep_price   = out_item.add_option_price or 0  # 옵션추가금

                if diff <= 10:
                    # (C2a) diff <= 10 => 재고 += diff
                    new_local_stock = (out_item.option_id_stock or 0) + diff

                    if platform in ["니뜰리히","수비다 SUBIDA","노는 개 최고양","아르빙"]:
                        ok, msg = naver_update_option_stock(
                            origin_no=origin_no,
                            option_id=opt_id,
                            new_stock=new_local_stock,
                            platform_name=platform,
                            keep_price=keep_price,
                            base_sale_price=base_price
                        )
                        if ok:
                            out_item.option_id_stock = new_local_stock
                            out_item.save()
                            action_count += 1
                        else:
                            logger.warning(f"[naver_update_option_stock fail] {msg}")

                    elif platform in ["쿠팡01","쿠팡02"]:
                        ok, msg = coupang_update_item_stock(
                            vendor_item_id=opt_id,
                            new_stock=new_local_stock,
                            platform_name=platform
                        )
                        if ok:
                            out_item.option_id_stock = new_local_stock
                            out_item.save()
                            action_count += 1
                        else:
                            logger.warning(f"[coupang_update_item_stock fail] {msg}")

                    else:
                        logger.debug(f"[unknown platform] {platform}")

                else:
                    # (C2b) diff >= 11 => 재고=9999
                    if platform in ["니뜰리히","수비다 SUBIDA","노는 개 최고양","아르빙"]:
                        ok, msg = put_naver_option_stock_9999(
                            origin_no=origin_no,
                            option_id=opt_id,
                            platform_name=platform,
                            base_sale_price=base_price,
                            keep_price=keep_price
                        )
                        if ok:
                            out_item.option_id_stock = 9999
                            out_item.save()
                            action_count += 1
                        else:
                            logger.warning(f"[put_naver_option_stock_9999 fail] {msg}")

                    elif platform in ["쿠팡01","쿠팡02"]:
                        ok, msg = put_coupang_option_stock_9999(
                            vendor_item_id=opt_id,
                            platform_name=platform
                        )
                        if ok:
                            out_item.option_id_stock = 9999
                            out_item.save()
                            action_count += 1
                        else:
                            logger.warning(f"[put_coupang_option_stock_9999 fail] {msg}")

                    else:
                        logger.debug(f"[unknown platform] {platform}")

    messages.success(
        request,
        f"셀러툴재고 {updated_st_count}건 업데이트 완료! 실제 옵션재고 조정 {action_count}건 수행."
    )
    return redirect('out_of_stock_management')



def coupang_product_check_view(request):
    """
    /delayed/coupang-product-check/?account_name=쿠팡01&seller_product_id=1234567
    로 들어오면, 해당 seller_product_id 상품 정보 조회
    """
    account_name = request.GET.get("account_name", "")
    seller_product_id = request.GET.get("seller_product_id", "")

    if not (account_name and seller_product_id):
        return JsonResponse({"error": "account_name, seller_product_id가 필요합니다."}, status=400)

    # (A) 쿠팡 계정 찾기
    account_info = None
    for acct in COUPANG_ACCOUNTS:
        if account_name in acct["names"]:
            account_info = acct
            break
    if not account_info:
        return JsonResponse({"error": f"해당 쿠팡 계정({account_name})을 찾을 수 없습니다."}, status=400)

    # (B) 함수 호출
    is_ok, result = get_coupang_seller_product_info(account_info, seller_product_id)
    if not is_ok:
        return JsonResponse({"error": f"조회실패: {result}"}, status=400)
    else:
        # 성공 → JSON 응답
        return JsonResponse({"product_data": result}, json_dumps_params={"ensure_ascii":False})
    

def match_option_ids_view(request):
    """
    (1) 현재 품절관리(OutOfStock)에서 status=1 인 레코드들의 option_code 수집
    (2) 그 option_code를 가진 OutOfStock 레코드는 전부 status=1 로 update
    (3) 완료 후 품절관리 페이지로 이동
    """
    # 1) status=1 인 레코드들의 option_code 집합 구하기
    code_qs = (
        OutOfStock.objects
        .filter(status=1)                           # 이미 표시된(1) 애들만
        .exclude(option_code__isnull=True)          # option_code 비어있지 않은 것
        .exclude(option_code__exact='')
        .values_list('option_code', flat=True)
    )
    code_set = set(code_qs)  # 중복제거

    # 2) code_set에 속하는 레코드를 전부 status=1 로 변경
    if code_set:
        OutOfStock.objects.filter(option_code__in=code_set).update(status=1)

    messages.success(request, "옵션ID 매칭(상태=1로 일괄적용) 작업이 완료되었습니다.")
    return redirect('out_of_stock_management')


def download_out_of_stock_excel_today(request):
    """
    오늘 품절처리된 OutOfStock 레코드들을 엑셀 파일로 다운로드.
    (금일 0시 ~ 내일 0시 범위의 out_of_stock_at)
    엑셀 열(순서): [옵션코드, 플랫폼명, 상품명, 옵션명, 품절처리시간]
    """

    # 1) 오늘 0시와 내일 0시 구하기
    now = timezone.localtime()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow_start = today_start + timedelta(days=1)

    # 2) 오늘 품절된 OutOfStock 쿼리
    out_qs = OutOfStock.objects.filter(
        out_of_stock_at__gte=today_start,
        out_of_stock_at__lt=tomorrow_start
    )

    # 3) openpyxl Workbook 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "금일 품절"

    # 4) 헤더 작성
    headers = ["옵션코드", "플랫폼명", "상품명", "옵션명", "품절처리시간"]
    ws.append(headers)

    # 5) 쿼리 결과를 엑셀 행에 추가
    for out_item in out_qs:
        option_code = out_item.option_code or ""
        platform_name = out_item.platform_name or ""
        product_name = out_item.seller_product_name or ""
        option_name = out_item.seller_option_name or ""

        out_of_stock_str = ""
        if out_item.out_of_stock_at:
            out_of_stock_str = out_item.out_of_stock_at.strftime("%Y-%m-%d %H:%M")

        ws.append([
            option_code,
            platform_name,
            product_name,
            option_name,
            out_of_stock_str,
        ])

    # 6) 메모리에 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 7) 파일 응답
    filename = "today_out_of_stock.xlsx"
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def out_of_stock_check_view(request):
    """
    품절확인 페이지:
    기존 '품절관리'와 동일한 UI/데이터 처리 로직을 가정하여 예시 구현
    """
    from django.db.models import Q
    from django.core.paginator import Paginator
    from django.shortcuts import render
    from .models import OutOfStockCheck

    # GET 파라미터 처리
    filter_val = request.GET.get('filter', 'all').strip()  # 'outofstock', 'instock', 'all'
    search_query = request.GET.get('search_query', '').strip()

    # 1) OutOfStock 중 status=1 (옵션매핑 전송된 것만 표시)
    base_qs = OutOfStockCheck.objects.filter(status=1).order_by('-updated_at')

    # 2) 필터 적용
    if filter_val == 'outofstock':
        base_qs = base_qs.filter(Q(option_id_stock__lte=0) | Q(option_id_stock__isnull=True))
    elif filter_val == 'instock':
        base_qs = base_qs.filter(option_id_stock__gt=0)
    # 'all' 은 필터 없이 전체

    # 3) 검색 (플랫폼명, 주문상품명, 옵션코드, 옵션ID 등)
    if search_query:
        lower_query = search_query.lower()
        base_qs = base_qs.filter(
            Q(platform_name__icontains=lower_query) |
            Q(order_product_name__icontains=lower_query) |
            Q(option_code__icontains=lower_query) |
            Q(option_id__icontains=lower_query)
        )

    # 4) 페이지네이션
    paginator = Paginator(base_qs, 100)  # 페이지당 100개
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    page_range_custom = page_obj.paginator.get_elided_page_range(
        page_obj.number, on_each_side=2, on_ends=1
    )

    context = {
        "details": page_obj.object_list,
        "page_obj": page_obj,
        "page_range_custom": page_range_custom,
        "filter_val": filter_val,
        "search_query": search_query,
    }

    return render(request, "delayed_management/out_of_stock_check.html", context)




def update_naver_out_of_stock_list(request):
    """
    네이버 API로부터 상품/옵션 리스트를 받아,
    품절( stockQuantity <= 0 )인 데이터만 OutOfStockCheck 모델에 기록한다.
    """
    
    account_name = request.GET.get("account_name", "").strip()
    if not account_name:
        return JsonResponse({"error": "account_name이 없습니다."}, status=400)

    # 1) 네이버 계정 찾기
    account_info = None
    for acct in NAVER_ACCOUNTS:
        if account_name in acct["names"]:
            account_info = acct
            break

    if not account_info:
        msg = f"[update_naver_out_of_stock_list] 계정 '{account_name}'을 NAVER_ACCOUNTS에서 찾을 수 없습니다."
        logger.warning(msg)
        return JsonResponse({"error": msg}, status=400)

    # 2) 네이버 상품/옵션 목록 가져오기
    is_ok, detailed_list = fetch_naver_products_with_details(account_info)
    if not is_ok:
        logger.error(f"[update_naver_out_of_stock_list] fetch fail: {detailed_list}")
        return JsonResponse({"error": "네이버 상품목록 조회 실패", "detail": detailed_list}, status=400)

    logger.debug(f"[update_naver_out_of_stock_list] detailed_list count={len(detailed_list)}")

    # 업데이트된(혹은 생성된) OutOfStockCheck 레코드 수
    updated_count = 0

    for prod_data in detailed_list:
        # (A) 공통 영역(상품 레벨)
        origin_no      = prod_data.get("originProductNo")         # 상품ID
        rep_image_dict = prod_data.get("representativeImage", {})  # 대표이미지 (dict)
        if isinstance(rep_image_dict, dict):
            rep_img_url = rep_image_dict.get("url", "")
        else:
            rep_img_url = ""

        # 네이버 '정가(salePrice)'와 '할인가(discountedPrice)'
        naver_sale_price       = prod_data.get("salePrice", 0)        # 정가
        naver_discounted_price = prod_data.get("discountedPrice", 0)  # 할인가
        product_name           = prod_data.get("productName", "")     # 주문상품명

        # (B) 옵션 목록 파싱
        option_combos = prod_data.get("optionCombinations", [])
        if not option_combos:
            # 옵션없는 단일상품인 경우 → stockQuantity <= 0 이면 품절로 간주
            single_option_id_stock = prod_data.get("stockQuantity", 0)
            if single_option_id_stock <= 0:
                # 품절상태라면 OutOfStockCheck에 업데이트
                _, created = OutOfStockCheck.objects.update_or_create(
                    option_id=str(origin_no),
                    defaults={
                        "platform_name":         account_name,
                        "representative_image":  rep_img_url,
                        "product_id":            origin_no,
                        "option_id_stock":       single_option_id_stock,
                        "expected_restock_date": None,
                        "order_product_name":    product_name,
                        "order_option_name_01":  "",     # 단일상품이므로 옵션명 X
                        "order_option_name_02":  "",
                        "option_code":           "",
                        "original_price":        naver_sale_price,
                        "sale_price":            naver_discounted_price,
                        "add_option_price":      0,
                        # out_of_stock_at, updated_at은 자동 or 필요 시 추가
                        "status": 1,   # 예: "표시상태"를 1로 설정해둘 수 있음 (옵션맵핑 전송완료 등)
                    }
                )
                updated_count += 1
            continue

        # (C) 옵션이 여러개 존재하는 경우
        for combo in option_combos:
            opt_id      = combo.get("id")                     # 옵션ID
            opt_stk     = combo.get("stockQuantity", 0)       # 옵션ID재고
            opt_name1   = combo.get("optionName1", "")        # 주문옵션명01
            opt_name2   = combo.get("optionName2", "")        # 주문옵션명02
            opt_code    = combo.get("sellerManagerCode", "")  # 옵션코드
            add_opt_prc = combo.get("price", 0)               # 추가옵션가

            # (C1) 품절(= stockQuantity <= 0)인 경우만 저장
            if opt_stk <= 0:
                # OutOfStockCheck에 저장/업데이트
                _, created = OutOfStockCheck.objects.update_or_create(
                    option_id=str(opt_id),
                    defaults={
                        "platform_name":         account_name,
                        "representative_image":  rep_img_url,
                        "product_id":            origin_no,
                        "option_id_stock":       opt_stk,
                        "expected_restock_date": None,  # API 데이터 없음
                        "order_product_name":    product_name,
                        "order_option_name_01":  opt_name1,
                        "order_option_name_02":  opt_name2,
                        "option_code":           opt_code,
                        "original_price":        naver_sale_price,
                        "sale_price":            naver_discounted_price,
                        "add_option_price":      add_opt_prc,
                        "status": 1,  # 품절확인이므로 상태값 조정 가능
                    }
                )
                updated_count += 1

    return JsonResponse({
        "message": f"{account_name} 품절 옵션 {updated_count}건 업데이트 완료",
        "count": updated_count,
    })


def update_coupang_out_of_stock_list(request):
    return JsonResponse({
        "error": "Coupang OutOfStock 업데이트는 아직 구현되지 않았습니다."
    }, status=501)


@require_POST
def update_seller_tool_stock_check(request):
    """
    (1) OutOfStockCheck 중 status=1인 레코드에서 option_code를 모아서
    (2) 100개씩 나누어 셀러툴 API(get_inventory_by_option_codes) 호출
    (3) OutOfStockCheck.seller_tool_stock 갱신
    (4) 결과: “업데이트 완료 X건, 변경없음 Y건”을 JSON으로 반환
    """
    logger = logging.getLogger(__name__)

    try:
        # (A) OutOfStockCheck 중 status=1인 레코드만 조회
        qs = OutOfStockCheck.objects.filter(status=1)
        if not qs.exists():
            return JsonResponse({
                "success": False,
                "message": "status=1 인 OutOfStockCheck 데이터가 없어 업데이트할 항목이 없습니다."
            }, status=400)

        # (B) option_code 수집
        codes = {
            (o.option_code or "").strip()
            for o in qs
            if (o.option_code or "").strip()
        }
        if not codes:
            return JsonResponse({
                "success": False,
                "message": "status=1 레코드 중 option_code가 없어 업데이트 불가"
            }, status=400)

        # (C) 100개씩 나누어 API 호출
        CHUNK_SIZE = 100
        stock_map = {}

        def chunked(seq, size):
            for i in range(0, len(seq), size):
                yield seq[i : i + size]

        for chunk in chunked(list(codes), CHUNK_SIZE):
            partial = get_inventory_by_option_codes(chunk)
            # partial: { option_code: 재고수량 }
            stock_map.update(partial)

        # (D) DB 갱신
        updated = 0
        unchanged = 0

        for item in qs:
            code = (item.option_code or "").strip()
            if not code:
                continue

            old_stock = item.seller_tool_stock or 0
            new_stock = stock_map.get(code, 0)

            if new_stock != old_stock:
                item.seller_tool_stock = new_stock
                item.save(update_fields=['seller_tool_stock'])
                updated += 1
            else:
                unchanged += 1

        return JsonResponse({
            "success": True,
            "message": (
                f"[update_seller_tool_stock_check] 셀러툴 재고 업데이트 완료! "
                f"(업데이트: {updated}건, 변경없음: {unchanged}건)"
            )
        })

    except Exception as e:
        logger.exception("[update_seller_tool_stock_check] 예외 발생")
        return JsonResponse({
            "success": False,
            "message": str(e),
        }, status=500)
    
def match_option_ids_check_view(request):
    """
    (1) 현재 품절관리(OutOfStock)에서 status=1 인 레코드들의 option_code 수집
    (2) 그 option_code를 가진 OutOfStock 레코드는 전부 status=1 로 update
    (3) 완료 후 품절관리 페이지로 이동
    """
    # 1) status=1 인 레코드들의 option_code 집합 구하기
    code_qs = (
        OutOfStockCheck.objects
        .filter(status=1)                           # 이미 표시된(1) 애들만
        .exclude(option_code__isnull=True)          # option_code 비어있지 않은 것
        .exclude(option_code__exact='')
        .values_list('option_code', flat=True)
    )
    code_set = set(code_qs)  # 중복제거

    # 2) code_set에 속하는 레코드를 전부 status=1 로 변경
    if code_set:
        OutOfStockCheck.objects.filter(option_code__in=code_set).update(status=1)

    messages.success(request, "옵션ID 매칭(상태=1로 일괄적용) 작업이 완료되었습니다.")
    return redirect('out_of_stock_check')




from .api_clients import (
    put_naver_option_stock_9999,
    put_coupang_option_stock_9999,
)

@require_POST
def add_stock_9999_check_view(request):
    """
    체크된 OutOfStockCheck 레코드들에 대해,
    - 네이버: put_naver_option_stock_9999 (base_sale_price=original_price, keep_price=add_option_price)
    - 쿠팡: put_coupang_option_stock_9999 (vendor_item_id=option_id)
    후 재고=9999로 설정
    """
    logger = logging.getLogger(__name__)

    # 1) 요청 바디 파싱
    try:
        body = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON body"}, status=400)

    # 2) PK 목록 수집
    raw_ids = body.get('detail_ids', [])
    pk_list = []
    for val in raw_ids:
        try:
            pk_list.append(int(val))
        except (ValueError, TypeError):
            logger.debug(f"[add_stock_9999_check_view] invalid id '{val}', skip")

    if not pk_list:
        return JsonResponse({"success": False, "message": "선택된 항목이 없습니다."}, status=400)

    # 3) OutOfStockCheck 쿼리
    qs = OutOfStockCheck.objects.filter(pk__in=pk_list)
    if not qs.exists():
        return JsonResponse({"success": False, "message": "해당 OutOfStockCheck이 없습니다."}, status=400)

    updated_count = 0

    # 4) 반복 처리
    for item in qs:
        platform  = (item.platform_name or "").strip()
        origin_no = (item.product_id or "").strip()
        option_id = (item.option_id or "").strip()

        # 네이버 계정
        if platform in ["니뜰리히", "수비다 SUBIDA", "노는 개 최고양", "아르빙"]:
            is_ok, msg = put_naver_option_stock_9999(
                origin_no=origin_no,
                option_id=option_id,
                platform_name=platform,
                base_sale_price=item.original_price,
                keep_price=item.add_option_price
            )
            if is_ok:
                item.option_id_stock = 9999
                item.save(update_fields=['option_id_stock'])
                updated_count += 1
            else:
                logger.warning(
                    f"[add_stock_9999_check_view] NAVER 재고 9999 실패: pk={item.pk}, msg={msg}"
                )

        # 쿠팡 계정
        elif platform in ["쿠팡01", "쿠팡02"]:
            is_ok, msg = put_coupang_option_stock_9999(
                vendor_item_id=option_id,
                platform_name=platform
            )
            if is_ok:
                item.option_id_stock = 9999
                item.save(update_fields=['option_id_stock'])
                updated_count += 1
            else:
                logger.warning(
                    f"[add_stock_9999_check_view] 쿠팡 재고 9999 실패: pk={item.pk}, msg={msg}"
                )
        else:
            logger.debug(f"[add_stock_9999_check_view] Unknown platform={platform}, skip")

    return JsonResponse({
        "success": True,
        "updated_count": updated_count,
        "message": f"{updated_count}건 재고를 9999로 변경했습니다."
    })
