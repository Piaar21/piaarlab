# excel_conversion/views.py
import logging
import re
from django.shortcuts import render, redirect
from django.urls import reverse
from django.http import HttpResponse
import pandas as pd
import io
from datetime import datetime

# 로거 설정
logger = logging.getLogger(__name__)

# 매핑 테이블: 원본 코드 → 내부 코드
code_map = {
    "55415499":"p0l84ir4cx72ulg274","55415500":"p1l84ist4e72vlmfek","55415501":"4xmt1dec8ca758g75u",
    "55415502":"ND22H01-S40CDGR","55415503":"ND22H01-S40CDG","55415504":"ND22H01-S40CP",
    "55415505":"4xmt1dec8ca6e3cazq","55415506":"p7l84isonz72viyxfk","55415507":"4xmt1dec8ca200iedj",
    "55415508":"p3l84iquh972ufkbhy","55415509":"p5l84ira6j72uowva1","55415510":"p8l84irhwk72utieav",
    "55415511":"4xmt1dec8ca193sfnh","55415512":"ND22H01-S40CGR","55415513":"4y6e1d431bcd74ween",
    "55415514":"p3l84isd4s72vc3q1l","55415515":"ND22H01-S40CDBR","55415516":"ND22H01-S40CBR",
    "55415517":"ND22H01-S40CMI","55415518":"4y6e1d431bcd90b3dk","41728543":"4y6b1df1afb68a9br5",
    "41728584":"4y6b1df1afb6924i4m","41728616":"4y6b1df1afb6b31zbh","41728657":"4y6b1df1afb6ccvowl",
    "41728576":"4y6b1df1afb6ab33fp","41728601":"4ybo1d6510fc77voqr","41728637":"4y6b1df1afb6e7rnho",
    "41728660":"4ybo1d6510fc9bgfe9","41733239":"4ybo1d6510fca5fapv","41728537":"4y6b1df1afb69amczt",
    "41728564":"4y6b1df1afb676stgf","41728579":"4y8s1924d21d91hewt","41728582":"55f41dfe938763t6pj",
    "41728606":"4y6b1df1afb6d61jte","41728613":"4y6b1df1afb6ba23o8","41728641":"4y8s1924d21db5qiqw",
    "41728647":"4yc31f38be5cf0dl5k","41730890":"4y6b1df1afb6a3gthy","41730897":"4y6b1df1afb6de7u7k",
    "41733236":"4ybo1d6510fc91v3zp",
    "58239312": "55tm21bd845456mtfc",
    "58239710": "5czo1f8c07a603naka",
    "58239316": "55tm21bd845479bfcy",
    "58239310": "55tm21bd84545dkzco",
    "58239709": "5czo1f8c07a615boag",
    "58239313": "55tm21bd845464iwzi",
    "58239315": "55tm21bd84544e1u79",
    "58239712": "5czo1f8c07a59c7mqu",
    "58240346": "5czo1f8c07a61csnnj",
    "58239711": "5czo1f8c07a618a1vg",
    "58239311": "55tm21bd84546bhlyu",
    "58239713": "5czo1f8c07a5fezb8i",
    "58239317": "55tm21bd845473qycb",
    "58239314": "55tm21bd845437dvpo",
}

# Apps Script 동일 순서의 출력 헤더
out_headers = [
    "판매채널 주문일시","상품명 (필수)","옵션정보 (필수)","수량 (필수)","수취인명 (필수)",
    "전화번호1 (필수)","전화번호2","주소 (필수)","주소상세","판매채널","주문번호1","주문번호2",
    "상품코드","옵션코드","우편번호","택배사","배송방식","배송메세지","운송장번호",
    "판매금액","배송비","바코드",
    "[M] 상품코드","[M] 옵션코드","[M] 출고옵션코드",
    "관리메모1","관리메모2","관리메모3","관리메모4","관리메모5",
    "관리메모6","관리메모7","관리메모8","관리메모9","관리메모10"
]

# 키 정규화 함수: '(필수)' 제거, '[M]'→'M', 공백 제거
def normalize_key(key: str) -> str:
    key = re.sub(r"\s*\(.*?\)", "", key)
    key = key.replace("[M]", "M")
    return key.replace(" ", "").strip()

def normalize_phone(raw: str) -> str:
    if raw is None:
        return ''
    s = str(raw).strip()

    # 흔한 케이스 방어: 숫자로 읽혀 '.0' 이 붙은 경우 제거
    if re.fullmatch(r'\d+\.0', s):
        s = s[:-2]

    # 숫자/플러스만 남기기
    s = re.sub(r'[^\d+]', '', s)

    if s.startswith('+82'):
        rest = s[3:]
        return rest if rest.startswith('0') else ('0' + rest)

    if s.startswith('82'):
        rest = s[2:]
        return rest if rest.startswith('0') else ('0' + rest)

    return re.sub(r'\D', '', s)

def excel_upload(request):
    """
    Excel 업로드 시 Apps Script 매핑 로직을 적용해 세션에 누적 저장
    """
    data_list = request.session.get('excel_data', [])
    error_message = None
    success_message = None

    if request.method == 'POST':
        files = request.FILES.getlist('excel_files')
        if not files:
            error_message = '파일이 선택되지 않았습니다.'
            logger.warning('No files selected for upload.')
        else:
            total_new = 0
            for f in files:
                if not f.name.lower().endswith(('.xls', '.xlsx')):
                    error_message = '엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.'
                    logger.warning('Invalid file type: %s', f.name)
                    continue
                try:
                    # 전체 시트 raw 읽기
                    df = pd.read_excel(f, header=None, dtype=str)

                    # 1) 상단 고정 필드 추출
                    order_no      = str(df.iat[9, 2])  # C10
                    order_manager = str(df.iat[9, 7])  # H10
                    addr1         = str(df.iat[12, 3]) # D13
                    addr2         = str(df.iat[12, 4]) # E13
                    address       = addr1
                    detail_addr   = addr2
                    raw_phones = [p.strip() for p in re.split(r'[\/,]', str(df.iat[12, 8] or '')) if p.strip()]

                    norm_phones = []
                    for p in raw_phones:
                        np = normalize_phone(p)
                        if np:
                            norm_phones.append(np)

                    phone1 = norm_phones[0] if norm_phones else ''
                    phone2 = norm_phones[1] if len(norm_phones) > 1 else ''
                    sales_channel = '로켓배송'

                    # 2) '합계' 앞 데이터 구간 찾기
                    total_idx = None
                    for idx, row in df.iterrows():
                        if row.astype(str).str.strip().eq('합계').any():
                            total_idx = idx
                            break
                    last_idx = (total_idx - 1) if total_idx is not None else (df.shape[0] - 1)

                    # 3) 헤더 병합 (20~21행)
                    r20 = df.iloc[19].fillna('').astype(str).tolist()
                    r21 = df.iloc[20].fillna('').astype(str).tolist()
                    hdrs = [((a+' '+b).strip() if a and b else a or b) for a, b in zip(r20, r21)]

                    # 정규화된 헤더 인덱스 맵
                    norm_hdr_idx = {normalize_key(h): i for i, h in enumerate(hdrs)}

                    # 4) 짝수행(22행부터) 데이터 매핑
                    for row_idx in range(21, last_idx + 1, 2):
                        row = df.iloc[row_idx]
                        def get_val(raw_key):
                            idx = norm_hdr_idx.get(normalize_key(raw_key))
                            return row.iat[idx] if idx is not None else ''

                        # 필드 추출
                        raw = str(get_val('상품명/옵션/BARCODE') or '').strip()
                        if not raw:
                            # 상품명/옵션이 비어 있으면 이 행은 삭제(건너뛰기)
                            continue
                        parts = [p.strip() for p in raw.split(',', 1)]
                        product_name = parts[0]
                        option_info  = parts[1] if len(parts) > 1 else ''
                        orig_code    = str(get_val('상품코드') or '')
                        mapped_code  = code_map.get(orig_code, orig_code)
                        qty          = get_val('발주수량')
                        amount       = df.iat[row_idx, 12]

                        # 5) 값 구성
                        values = []
                        for hdr in out_headers:
                            nk = normalize_key(hdr)
                            if nk == normalize_key('상품명 (필수)'):      val = product_name
                            elif nk == normalize_key('옵션정보 (필수)'):    val = option_info
                            elif nk == normalize_key('수량 (필수)'):        val = qty
                            elif nk == normalize_key('수취인명 (필수)'):    val = f"{get_val('물류센터')} (로켓배송) {order_no}"
                            elif nk == normalize_key('전화번호1 (필수)'):   val = phone1
                            elif nk == normalize_key('전화번호2'):          val = phone2
                            elif nk == normalize_key('주소 (필수)'):        val = address
                            # elif nk == normalize_key('주소상세'):          val = detail_addr
                            elif nk == normalize_key('판매채널'):           val = sales_channel
                            elif nk == normalize_key('주문번호1'):          val = order_no
                            elif nk == normalize_key('주문번호2'):          val = ''
                            elif nk == normalize_key('판매금액'):           val = amount
                            elif nk in [normalize_key(x) for x in ['[M] 상품코드','[M] 옵션코드','[M] 출고옵션코드']]: val = mapped_code
                            elif nk == normalize_key('관리메모1'):         val = order_manager
                            else:                                           val = ''
                            values.append(val)

                        # 6) 키를 정규화된 상태로 저장
                        mapped_row = { normalize_key(hdr): val for hdr, val in zip(out_headers, values) }
                        data_list.append(mapped_row)
                        total_new += 1

                    logger.info('Mapped %d rows from file %s', total_new, f.name)
                except Exception as e:
                    error_message = f'파일 매핑 중 오류: {e}'
                    logger.error('Error mapping %s: %s', f.name, e)

            # 세션에 저장 및 피드백
            request.session['excel_data'] = data_list
            success_message = f"{len(files)}개 파일 처리, 총 {total_new}건 매핑되었습니다."
            logger.info(success_message)

    return render(request, 'excel_conversion/excel_upload.html', {
        'data_list': data_list,
        'error_message': error_message,
        'success_message': success_message,
    })

from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter


from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

def excel_download(request):
    data_list = request.session.get('excel_data', [])
    if not data_list:
        return redirect(reverse('excel_conversion:excel_upload'))

    # 1) DataFrame 생성 & 컬럼 이름 셋팅
    df = pd.DataFrame(data_list)
    norm_keys = [normalize_key(h) for h in out_headers]
    df = df[norm_keys]
    df.columns = out_headers

    # 2) 숫자형으로 보장 ('수량 (필수)'→K, '판매금액'→T)
    for col in ['수량 (필수)', '판매금액']:
        # 문자열 → 숫자: 콤마 제거 후 float
        df[col] = (
            df[col]
            .astype(str)
            .str.replace(',', '', regex=True)
            .pipe(pd.to_numeric, errors='coerce')
        )

    # 3) 엑셀 쓰기 + 숫자 포맷 지정
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='주문데이터')
        ws = writer.sheets['주문데이터']

        for col in ['수량 (필수)', '판매금액']:
            idx = df.columns.get_loc(col) + 1               # 1-base
            letter = get_column_letter(idx)
            # 2행부터 마지막 행까지 순회
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                     min_col=idx, max_col=idx):
                for cell in row:
                    # 천단위 콤마 포함 정수 포맷
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    output.seek(0)
    filename = f'주문데이터_{datetime.now().date().isoformat()}.xlsx'
    resp = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp

def excel_clear(request):
    if request.method == 'POST':
        request.session.pop('excel_data', None)
        logger.info('Session data cleared.')
    return redirect(reverse('excel_conversion:excel_upload'))

def excel_clear_set(request):
    if request.method == 'POST':
        request.session.pop('excel_data', None)
        logger.info('Session data cleared.')
    return redirect(reverse('excel_conversion:excel_settlement'))

def excel_shipcode_clear_set(request):
    if request.method == 'POST':
        request.session.pop('excel_shipcode_rows', None)  # ← 이 키로 저장했으니 이걸 지워야 함
    return redirect(reverse('excel_conversion:excel_shipcode'))  # ← shipcode 화면으로


new_headers = [
    "상품코드", 
    "상품명/옵션/BARCODE", 
    "매입유형/면세여부", 
    "발주유형", 
    "물류센터", 
    "발주수량", 
    "업체납품가능수량", 
    "입고수량", 
    "매입가(공급가)", 
    "공급가액(공급가)", 
    "부가세(공급가)", 
    "매입가(발주금액)", 
    "공급가액(발주금액)", 
    "부가세(발주금액)", 
    "생산년도", 
    "매입가(입고금액)", 
    "공급가액(입고금액)", 
    "부가세(입고금액)", 
    "제조일자관리", 
    "제조(수입)일자", 
    "유통(소비)기한"
]



def excel_settlement(request):
    data_list = request.session.get('excel_data', [])
    error_message = None
    success_message = None

    if request.method == 'POST':
        files = request.FILES.getlist('excel_files')
        if not files:
            error_message = '파일이 선택되지 않았습니다.'
            logger.warning('No files selected for settlement.')
        else:
            total_new = 0
            for f in files:
                if not f.name.lower().endswith(('.xls', '.xlsx')):
                    error_message = '엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.'
                    logger.warning('Invalid file type: %s', f.name)
                    continue

                try:
                    df = pd.read_excel(f, header=None)
                    product_code = str(df.iat[9, 2])  # C10

                    # '합계' 직전 행 찾기
                    total_idx = None
                    for idx, row in df.iterrows():
                        if row.astype(str).str.strip().eq('합계').any():
                            total_idx = idx
                            break
                    last_idx = (total_idx - 1) if total_idx is not None else (df.shape[0] - 1)

                    # 키 순서
                    keys = [
                        "상품코드","상품명_옵션_BARCODE","매입유형_면세여부","발주유형","물류센터",
                        "발주수량","업체납품가능수량","입고수량","매입가_공급가","공급가액_공급가",
                        "부가세_공급가","매입가_발주금액","공급가액_발주금액","부가세_발주금액",
                        "생산년도","매입가_입고금액","공급가액_입고금액","부가세_입고금액",
                        "제조일자관리","제조_수입일자","유통_소비기한"
                    ]

                    for row_idx in range(21, last_idx+1, 2):
                        odd = df.iloc[row_idx]
                        odd_values = [
                            product_code,
                            odd.iloc[1], odd.iloc[2], odd.iloc[3], odd.iloc[4],
                            odd.iloc[5], odd.iloc[6], odd.iloc[7],
                            odd.iloc[8], odd.iloc[9], odd.iloc[10],
                            odd.iloc[11], odd.iloc[12], odd.iloc[13],
                            odd.iloc[14], odd.iloc[15], odd.iloc[16],
                            odd.iloc[17], odd.iloc[18], odd.iloc[19],
                            odd.iloc[20]
                        ]
                        data_list.append(dict(zip(keys, odd_values)))
                        total_new += 1

                        # 짝수행: 상품명/옵션/BARCODE 컬럼에만 바코드를 채움
                        if row_idx+1 <= last_idx:
                            even = df.iloc[row_idx+1]
                            even_values = [""] * len(keys)
                            even_values[2] = even.iloc[2]  # 두 번째 칸에 바코드
                            data_list.append(dict(zip(keys, even_values)))
                            total_new += 1

                    logger.info(f'Mapped {total_new} rows from file {f.name}')

                except Exception as e:
                    error_message = f'파일 매핑 중 오류: {e}'
                    logger.error(f'Error mapping {f.name}: {e}')

            request.session['excel_data'] = data_list
            success_message = f"{len(files)}개 파일 처리, 총 {total_new}건 매핑되었습니다."
            logger.info(success_message)

    return render(request, 'excel_conversion/excel_settlement.html', {
        'data_list': data_list,
        'error_message': error_message,
        'success_message': success_message,
    })



# 다운로드할 엑셀의 최종 헤더 (사용자에게 보여줄 컬럼명)
SETTLEMENT_OUT_HEADERS = [
    "상품코드",
    "상품명/옵션/BARCODE",
    "매입유형/면세여부",
    "발주유형",
    "물류센터",
    "발주수량",
    "업체납품가능수량",
    "입고수량",
    "매입가(공급가)",
    "공급가액(공급가)",
    "부가세(공급가)",
    "매입가(발주금액)",
    "공급가액(발주금액)",
    "부가세(발주금액)",
    "생산년도",
    "매입가(입고금액)",
    "공급가액(입고금액)",
    "부가세(입고금액)",
    "제조일자관리",
    "제조(수입)일자",
    "유통(소비)기한",
]

# 내부적으로 DataFrame에 쓰이는 키 (앞서 매핑한 mapped_row의 키)
SETTLEMENT_KEYS = [
    "상품코드",
    "상품명_옵션_BARCODE",
    "매입유형_면세여부",
    "발주유형",
    "물류센터",
    "발주수량",
    "업체납품가능수량",
    "입고수량",
    "매입가_공급가",
    "공급가액_공급가",
    "부가세_공급가",
    "매입가_발주금액",
    "공급가액_발주금액",
    "부가세_발주금액",
    "생산년도",
    "매입가_입고금액",
    "공급가액_입고금액",
    "부가세_입고금액",
    "제조일자관리",
    "제조_수입일자",
    "유통_소비기한",
]


def excel_download_settlement(request):
    # 세션에서 정산 데이터 불러오기
    data_list = request.session.get('excel_data', [])
    if not data_list:
        # 데이터가 없으면 업로드 페이지로
        return redirect(reverse('excel_conversion:excel_settlement'))

    # DataFrame 생성
    df = pd.DataFrame(data_list)

    # 내부 키 순서로 정렬 후, 사용자 헤더로 컬럼명 변경
    df = df[SETTLEMENT_KEYS]
    df.columns = SETTLEMENT_OUT_HEADERS

    # in-memory 엑셀 생성
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='정산데이터')
    output.seek(0)

    # 응답 설정
    filename = f'정산데이터_{datetime.now().date().isoformat()}.xlsx'
    resp = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp

import logging, re
import pandas as pd
from django.shortcuts import render
from django.views.decorators.csrf import csrf_protect

logger = logging.getLogger(__name__)

# ---- 공통 유틸 ----
def _find_col(df, candidates):
    for cand in candidates:
        for c in df.columns:
            if cand in str(c):
                return c
    raise KeyError(f"컬럼을 찾을 수 없습니다: {candidates}")

def _normalize_color(text: str) -> str:
    s = str(text or "").strip()
    s = re.sub(r'^\d+\.\s*', '', s)   # ← "2.그레이" 같은 번호 접두 제거
    s = re.sub(r'\s+', '', s)         # ← 공백 제거
    return s


def _extract_color_from_sku_name(sku_name: str) -> str:
    parts = [p.strip() for p in str(sku_name or "").split(",") if p.strip()]
    for token in reversed(parts):
        if not re.search(r'\d+\s*cm', token, flags=re.I):  # 사이즈는 건너뜀
            return _normalize_color(token)
    return ""


def _collect_allowed_colors(df2, col_sku_nm) -> set:
    """
    2.xlsx '상품이름(SKU Name)'에서 색상 후보를 모두 추출해 색상 사전 구성.
    """
    colors = set()
    for v in df2[col_sku_nm].dropna().astype(str):
        c = _extract_color_from_sku_name(v)
        if c:
            colors.add(c)
    return colors

def _split_multi_items_by_delim(cell_text: str, allowed_colors: set):
    text = str(cell_text or "")
    parts = re.split(r'\s*,\s*,\s*', text.strip())
    items = []
    for seg in parts:
        s = seg.strip().strip(',')
        if not s:
            continue
        is_grid = ("격자" in s)

        # qty
        qty = 0
        m = re.search(r'-(\d+)\s*$', s) or re.search(r'(\d+)\s*$', s)
        if m: qty = int(m.group(1))

        # size
        size = None
        ms = re.search(r'(\d+)\s*cm', s, flags=re.I)
        if ms:
            size = int(ms.group(1))
        else:
            ms2 = re.search(r'\b([3-5]\d)\b', s)  # 30~59 힌트(필요시 범위 조절)
            if ms2: size = int(ms2.group(1))

        # color: 가장 긴 후보부터 찾기(부분일치 충돌 방지)
        norm_seg = _normalize_color(s)
        color = ""
        for c in sorted(allowed_colors, key=len, reverse=True):
            if c and c in norm_seg:
                color = c
                break

        items.append({"is_grid": is_grid, "color": color, "size": size, "qty": qty})
    return items



def _pick_sheet(files):
    ship_file = list_file = None
    for f in files:
        try:
            xls = pd.ExcelFile(f)
            names = set(xls.sheet_names)
            if "Sheet1" in names and ship_file is None:
                ship_file = f
            if "상품목록" in names and list_file is None:
                list_file = f
        except Exception:
            continue
    return ship_file, list_file

# ---- 메인 뷰 ----
@csrf_protect
def excel_shipcode(request):
    if request.method == "GET":
        return render(request, "excel_conversion/excel_shipcode.html", {
            "data_list": [], "error_message": None, "success_message": None,
        })

    files = request.FILES.getlist("excel_files")
    if not files or len(files) < 2:
        return render(request, "excel_conversion/excel_shipcode.html", {
            "data_list": [], "error_message": "출고데이터(1.xlsx)와 상품목록(2.xlsx) 두 개의 파일을 업로드해주세요.", "success_message": None,
        })

    try:
        ship_file, list_file = _pick_sheet(files)
        if ship_file is None or list_file is None:
            return render(request, "excel_conversion/excel_shipcode.html", {
                "data_list": [], "error_message": "시트 식별 실패: 1.xlsx는 'Sheet1', 2.xlsx는 '상품목록' 시트를 포함해야 합니다.", "success_message": None,
            })

        df1 = pd.read_excel(ship_file, sheet_name="Sheet1")   # 출고데이터
        df2 = pd.read_excel(list_file, sheet_name="상품목록")  # 상품목록

        col_recv    = _find_col(df1, ["받는분"])
        col_invoice = _find_col(df1, ["운송장번호"]) if "운송장번호" in map(str, df1.columns) \
            else _find_col(df1, ["송장번호", "운송장"])  # fallback
        col_pname   = _find_col(df1, ["상품명"])   # 멀티 품목 가능

        col_po      = _find_col(df2, ["발주번호(PO ID)"])
        col_fc      = _find_col(df2, ["물류센터(FC)"])
        col_ttype   = _find_col(df2, ["입고유형(Transport Type)"])
        col_edd     = _find_col(df2, ["입고예정일(EDD)"])
        col_sku_id  = _find_col(df2, ["상품번호(SKU ID)"])
        col_sku_bar = _find_col(df2, ["상품바코드(SKU Barcode)"])
        col_sku_nm  = _find_col(df2, ["상품이름(SKU Name)"])
        col_cqty    = _find_col(df2, ["확정수량(Confirmed Qty)"])

        allowed_colors = _collect_allowed_colors(df2, col_sku_nm)

        rows = []
        for _, r in df2.iterrows():
            po_id          = str(r[col_po]).strip()
            fc             = str(r[col_fc]).strip()
            transport_type = str(r.get(col_ttype, "")).strip()
            edd            = str(r.get(col_edd, "")).strip()
            sku_id         = str(r.get(col_sku_id, "")).strip()
            sku_barcode    = str(r.get(col_sku_bar, "")).strip()
            sku_name       = str(r.get(col_sku_nm, "")).strip()
            confirmed_raw  = str(r.get(col_cqty, "")).replace(",", "").strip()
            total_qty      = int(float(confirmed_raw)) if confirmed_raw not in ("", "nan", "None") else 0

            # 2.xlsx SKU 기준 속성
            sku_attr = {
                "is_grid": ("격자" in sku_name),
                "size":    (lambda m: int(m.group(1)) if m else None)(re.search(r'(\d+)\s*cm', sku_name, flags=re.I)),
                "color":   _extract_color_from_sku_name(sku_name),
            }

            recv_key = f"{fc} (로켓배송) {po_id}"
            df1_scope = df1[df1[col_recv].astype(str).str.strip() == recv_key].copy()

            candidate_boxes = []
            unmatched_segments = []

            for __, c in df1_scope.iterrows():
                invoice = str(c[col_invoice]).strip()
                subitems = _split_multi_items_by_delim(c[col_pname], allowed_colors)
                hit = False
                for it in subitems:
                    cond_color_ok = (not it["color"] or not sku_attr["color"] or it["color"] == _normalize_color(sku_attr["color"]))
                    cond_size_ok = (it["size"] is None or sku_attr["size"] is None or it["size"] == sku_attr["size"])
                    if it["is_grid"] == sku_attr["is_grid"] and cond_color_ok and cond_size_ok and it["qty"] > 0:
                        candidate_boxes.append((it["qty"], invoice))
                        hit = True
                if not hit:
                    unmatched_segments.append(str(c[col_pname]))

            if not candidate_boxes and total_qty > 0:
                logger.warning("NO BOX MATCH: PO=%s FC=%s SKU=%s (color=%s size=%s grid=%s) segments=%s",
                            po_id, fc, sku_name, sku_attr["color"], sku_attr["size"], sku_attr["is_grid"],
                            unmatched_segments[:3])  # 과다 로그 방지

            # 박스→수량 할당 (송장별 합산)
            remaining = total_qty
            alloc_by_invoice = {}

            for box_qty, inv in candidate_boxes:
                if remaining <= 0:
                    break
                use = min(box_qty, remaining)
                alloc_by_invoice[inv] = alloc_by_invoice.get(inv, 0) + use
                remaining -= use

            # 송장별 합계로 한 줄씩 출력
            for inv, qty in alloc_by_invoice.items():
                rows.append({
                    "poId": po_id, "fc": fc, "transportType": transport_type, "edd": edd,
                    "skuId": sku_id, "skuBarcode": sku_barcode, "skuName": sku_name,
                    "confirmedQty": str(qty),         # ← 납품수량과 동일하게
                    "invoiceNumber": inv,
                    "shippedQty": str(qty),
                })

            # 남은 수량(미배정)이 있으면 송장 공란으로 한 줄 표시 (필요 시 유지/삭제)
            if remaining > 0:
                rows.append({
                    "poId": po_id, "fc": fc, "transportType": transport_type, "edd": edd,
                    "skuId": sku_id, "skuBarcode": sku_barcode, "skuName": sku_name,
                    "confirmedQty": str(remaining),   # ← 납품수량과 동일
                    "invoiceNumber": "",
                    "shippedQty": str(remaining),
                })


        request.session["excel_shipcode_rows"] = rows
        return render(request, "excel_conversion/excel_shipcode.html", {
            "data_list": rows, "error_message": None, "success_message": f"{len(rows)}건 처리 완료",
        })

    except Exception as e:
        logger.exception("excel_shipcode 처리 오류")
        return render(request, "excel_conversion/excel_shipcode.html", {
            "data_list": [], "error_message": f"처리 중 오류: {e}", "success_message": None,
        })
